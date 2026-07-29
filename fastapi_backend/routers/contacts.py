# -*- coding: utf-8 -*-
"""
部门通讯录 API
"""
import logging
import os
import tempfile
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from database import db, db_demo

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/contacts", tags=["通讯录"])

JB_RANK = [
    ("副经理", 2), ("经理", 1),
    ("副主任", 4), ("主任", 3),
    ("副组长", 6), ("组长", 5),
]

COMPANY_CONTACTS_TABLE = "company_contacts"
COMPANY_CONTACTS_CREATE_SQL = f"""
CREATE TABLE IF NOT EXISTS {COMPANY_CONTACTS_TABLE} (
    id BIGINT NOT NULL AUTO_INCREMENT PRIMARY KEY,
    organization VARCHAR(255) NOT NULL,
    group_name VARCHAR(255) NOT NULL DEFAULT '',
    name VARCHAR(100) NOT NULL,
    position VARCHAR(500) NOT NULL DEFAULT '',
    office_phone VARCHAR(100) NOT NULL DEFAULT '',
    source_row INT NOT NULL DEFAULT 0,
    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
    KEY idx_company_contacts_organization (organization),
    KEY idx_company_contacts_name (name)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""


def _ensure_company_contacts_table() -> None:
    """Create the independent company directory table on first use."""
    result = db.execute_update(COMPANY_CONTACTS_CREATE_SQL)
    if result < 0:
        raise RuntimeError("公司通讯录数据表初始化失败")


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_company_contacts_xlsx(file_path: str) -> list[dict]:
    """Parse the supplied workbook's repeated directory-table layout."""
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("无法读取 Excel 文件，请上传有效的 .xlsx 通讯录表") from exc

    records = []
    title_suffix = "电话号码表"
    for sheet in workbook.worksheets:
        organization = ""
        group_name = ""
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [_cell_text(value) for value in row[:5]]
            if not values:
                continue
            first = values[0]
            if first.endswith(title_suffix):
                organization = first[:-len(title_suffix)].strip()
                group_name = ""
                continue
            if not organization:
                continue

            # Contact rows consistently use the first five columns:
            # serial number, group, name, position, office phone.
            serial, group, name, position, office_phone = values
            if not serial or not name or not office_phone:
                continue
            if not serial.replace(".", "", 1).isdigit():
                continue
            if group:
                group_name = group
            records.append({
                "organization": organization,
                "group_name": group_name,
                "name": name,
                "position": position,
                "office_phone": office_phone,
                "source_row": row_number,
            })

    if not records:
        raise ValueError("未从 Excel 中识别到通讯录数据，请使用“班组、姓名、岗位/职务、办公电话”格式")
    return records


def _replace_company_contacts(records: list[dict]) -> int:
    """Atomically replace the company directory after a successful import."""
    conn = db.get_connection()
    if not conn:
        raise RuntimeError("无法连接公司通讯录数据库")
    try:
        with conn.cursor() as cursor:
            cursor.execute(COMPANY_CONTACTS_CREATE_SQL)
            cursor.execute(f"DELETE FROM {COMPANY_CONTACTS_TABLE}")
            cursor.executemany(
                f"""INSERT INTO {COMPANY_CONTACTS_TABLE}
                    (organization, group_name, name, position, office_phone, source_row)
                    VALUES (%s, %s, %s, %s, %s, %s)""",
                [
                    (
                        r["organization"], r["group_name"], r["name"],
                        r["position"], r["office_phone"], r["source_row"],
                    )
                    for r in records
                ],
            )
        conn.commit()
        return len(records)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _bootstrap_company_contacts_if_empty() -> None:
    """Seed the supplied workbook once, so a fresh deployment immediately has data."""
    _ensure_company_contacts_table()
    if (db.execute_scalar(f"SELECT COUNT(*) FROM {COMPANY_CONTACTS_TABLE}") or 0) > 0:
        return
    project_root = Path(__file__).resolve().parents[2]
    candidates = sorted(project_root.glob("*电话号码表*.xlsx"))
    if not candidates:
        return
    records = _parse_company_contacts_xlsx(str(candidates[0]))
    _replace_company_contacts(records)


def _can_manage_company_contacts(name: str) -> bool:
    """Only the existing system and personnel administrators may replace the directory."""
    try:
        from routers.approvers import is_admin1_user, is_admin2_user
        return is_admin1_user(name) or is_admin2_user(name)
    except Exception:
        return False


def _jb_sort_key(jb: str) -> int:
    jb = (jb or "").strip()
    for keyword, rank in JB_RANK:
        if keyword in jb:
            return rank
    return 99


@router.get("/list")
async def get_contacts(
    department: Optional[str] = Query(None, description="筛选科室"),
    keyword: Optional[str] = Query(None, description="搜索关键字（姓名/工号/手机/座机）"),
    source: str = Query("department", description="department=部门通讯录，company=公司通讯录"),
):
    """按科室分组返回通讯录，领导优先排序"""
    if (source or "department").strip().lower() == "company":
        return _get_company_contacts(department, keyword)
    try:
        sql = (
            "SELECT name, gh, lsys, jb, rcnf, sfzh, enterprise_email "
            "FROM yggl WHERE COALESCE(zaizhi, 0) = 0 "
            "AND TRIM(lsys) NOT IN ('其他部门员工', '其他部门成员', '') "
            "AND lsys IS NOT NULL "
            "AND RIGHT(TRIM(name), 1) != '1' "
        )
        params = []
        if department:
            sql += " AND lsys = %s"
            params.append(department.strip())
        sql += " ORDER BY lsys, gh"
        rows = db.execute_query(sql, tuple(params) if params else None)
        if not rows:
            return {"success": True, "departments": [], "total": 0}

        sfzh_map = {}
        for r in rows:
            sfzh = (r.get("sfzh") or "").strip().replace(" ", "")
            if sfzh:
                sfzh_map[sfzh] = (r.get("name") or "").strip()

        phone_map = {}
        if sfzh_map:
            id_cards = list(sfzh_map.keys())
            batch_size = 100
            for i in range(0, len(id_cards), batch_size):
                batch = id_cards[i:i + batch_size]
                ph = ",".join(["%s"] * len(batch))
                try:
                    demo_rows = db_demo.execute_query(
                        f"SELECT id_card, mobile, telephone FROM employee_info WHERE id_card IN ({ph})",
                        tuple(batch),
                    )
                    for dr in demo_rows or []:
                        idc = (dr.get("id_card") or "").strip()
                        name = sfzh_map.get(idc, "")
                        if name:
                            phone_map[name] = {
                                "mobile": (dr.get("mobile") or "").strip(),
                                "telephone": (dr.get("telephone") or "").strip(),
                            }
                except Exception as e:
                    logger.warning("查询 demo 库 employee_info 失败: %s", e)

        dept_groups = {}
        for r in rows:
            name = (r.get("name") or "").strip()
            dept = (r.get("lsys") or "").strip()
            jb = (r.get("jb") or "").strip()
            gh = (r.get("gh") or "").strip()
            rcnf = r.get("rcnf")
            if hasattr(rcnf, "strftime"):
                rcnf_str = rcnf.strftime("%Y-%m-%d")
            else:
                rcnf_str = str(rcnf)[:10] if rcnf else ""
            phones = phone_map.get(name, {})
            mobile = phones.get("mobile", "")
            telephone = phones.get("telephone", "")
            email = (r.get("enterprise_email") or "").strip()

            if keyword:
                kw = keyword.strip().lower()
                searchable = f"{name}{gh}{mobile}{telephone}{email}{dept}{jb}".lower()
                if kw not in searchable:
                    continue

            person = {
                "name": name,
                "gh": gh,
                "jb": jb,
                "department": dept,
                "mobile": mobile,
                "telephone": telephone,
                "email": email,
                "rcnf": rcnf_str,
            }
            dept_groups.setdefault(dept, []).append(person)

        DEPT_ORDER = [
            "部办",
            "综合技术室", "工具技术室", "数控编程室", "智能制造技术室",
            "水发工艺室", "水轮机工艺室", "汽发工艺室", "焊接工艺室", "非标技术室",
        ]

        def dept_sort_key(dept_name):
            try:
                return DEPT_ORDER.index(dept_name)
            except ValueError:
                return len(DEPT_ORDER)

        for members in dept_groups.values():
            members.sort(key=lambda p: (_jb_sort_key(p["jb"]), p["gh"]))

        sorted_depts = sorted(dept_groups.keys(), key=dept_sort_key)
        departments = []
        total = 0
        for dept_name in sorted_depts:
            members = dept_groups[dept_name]
            total += len(members)
            departments.append({
                "name": dept_name,
                "count": len(members),
                "members": members,
            })

        return {"success": True, "departments": departments, "total": total}
    except Exception as e:
        logger.error("获取通讯录失败: %s", e)
        return {"success": False, "departments": [], "total": 0, "error": str(e)}


def _get_company_contacts(organization: Optional[str], keyword: Optional[str]) -> dict:
    """Return the imported company directory in the grouped format used by the UI."""
    try:
        _bootstrap_company_contacts_if_empty()
        sql = (
            f"SELECT organization, group_name, name, position, office_phone "
            f"FROM {COMPANY_CONTACTS_TABLE} WHERE 1=1"
        )
        params = []
        if organization:
            sql += " AND organization = %s"
            params.append(organization.strip())
        if keyword and keyword.strip():
            sql += " AND CONCAT_WS('', organization, group_name, name, position, office_phone) LIKE %s"
            params.append(f"%{keyword.strip()}%")
        sql += " ORDER BY organization, source_row, id"
        rows = db.execute_query(sql, tuple(params) if params else None)

        groups = {}
        for row in rows:
            org = (row.get("organization") or "").strip()
            if not org:
                continue
            groups.setdefault(org, []).append({
                "name": (row.get("name") or "").strip(),
                "gh": "",
                "jb": (row.get("position") or "").strip(),
                "department": org,
                "group": (row.get("group_name") or "").strip(),
                "mobile": "",
                "telephone": (row.get("office_phone") or "").strip(),
                "email": "",
                "rcnf": "",
            })
        departments = [
            {"name": name, "count": len(members), "members": members}
            for name, members in groups.items()
        ]
        return {
            "success": True,
            "source": "company",
            "departments": departments,
            "total": sum(group["count"] for group in departments),
        }
    except Exception as exc:
        logger.error("获取公司通讯录失败: %s", exc)
        return {"success": False, "departments": [], "total": 0, "error": str(exc)}


@router.get("/can-manage-company")
async def can_manage_company_contacts(name: str = Query("", description="当前用户名")):
    return {"success": True, "canManage": _can_manage_company_contacts((name or "").strip())}


@router.post("/company/import")
async def import_company_contacts(
    name: str = Query("", description="当前用户名"),
    file: UploadFile = File(...),
):
    """Replace the company directory with a new version of the standard Excel workbook."""
    operator = (name or "").strip()
    if not _can_manage_company_contacts(operator):
        raise HTTPException(status_code=403, detail="仅系统管理员或人事管理员可更新公司通讯录")
    filename = (file.filename or "").strip()
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的公司电话号码表")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传的 Excel 文件为空")
    if len(content) > 30 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Excel 文件不能超过 30MB")

    temp_path = ""
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_file.write(content)
            temp_path = temp_file.name
        records = _parse_company_contacts_xlsx(temp_path)
        imported_count = _replace_company_contacts(records)
        return {
            "success": True,
            "count": imported_count,
            "message": f"公司通讯录已更新，共导入 {imported_count} 条联系人",
        }
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("导入公司通讯录失败: %s", exc)
        raise HTTPException(status_code=500, detail="公司通讯录更新失败，请稍后重试") from exc
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except OSError:
                pass
