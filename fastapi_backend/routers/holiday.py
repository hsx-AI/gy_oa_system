# -*- coding: utf-8 -*-
"""
假期数据API路由
假期数据来源：数据库 holiday 表（year, date, type）。
"""
from fastapi import APIRouter, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional, List
from pydantic import BaseModel
from models import HolidayResponse, Holiday
from utils.holiday_loader import load_holidays_for_year
from datetime import datetime
from database import db
from routers.db_manager import _get_admin1
from io import BytesIO
import os
import json
import logging
import traceback

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

router = APIRouter(prefix="/holiday", tags=["假期数据"])


def _get_dakaman() -> Optional[str]:
    """从 webconfig 表读取 dakaman 字段（打卡管理员用户名）。"""
    try:
        rows = db.execute_query("SELECT dakaman FROM webconfig WHERE id = %s LIMIT 1", ("1",))
        if rows and rows[0].get("dakaman") is not None:
            return (rows[0]["dakaman"] or "").strip() or None
    except Exception as e:
        logger.debug(f"读取 webconfig.dakaman 失败: {e}")
    return None


# 本地大模型默认地址与模型名（当 webconfig 未配置时使用）
DEFAULT_LLM_BASE_URL = "http://10.42.60.250:11434/v1"
DEFAULT_LLM_MODEL = "qwen3:8b"


def _normalize_llm_base_url(url: str) -> str:
    """兼容用户把完整路径填进来（如 http://host:port/v1/chat/completions）。
    OpenAI SDK 只需要根路径，最多到 /v1。"""
    u = (url or "").strip()
    if not u:
        return u
    u = u.rstrip("/")
    if u.endswith("/chat/completions"):
        u = u[: -len("/chat/completions")].rstrip("/")
    return u


def _get_llm_config() -> dict:
    """
    从 webconfig 表读取本地大模型配置（URL、模型名）。
    返回: {"base_url": str, "model": str}
    """
    base_url = DEFAULT_LLM_BASE_URL
    model = DEFAULT_LLM_MODEL
    try:
        rows = db.execute_query(
            "SELECT llm_base_url, llm_model FROM webconfig WHERE id = %s LIMIT 1",
            ("1",),
        )
        if rows:
            r = rows[0]
            if r.get("llm_base_url") is not None and (r.get("llm_base_url") or "").strip():
                base_url = (r["llm_base_url"] or "").strip()
            if r.get("llm_model") is not None and (r.get("llm_model") or "").strip():
                model = (r["llm_model"] or "").strip()
    except Exception as e:
        logger.debug(f"读取 webconfig 大模型配置失败（可能无 llm_base_url/llm_model 列）: {e}")
    base_url = _normalize_llm_base_url(base_url)
    return {"base_url": base_url, "model": model}


@router.get("", response_model=HolidayResponse)
async def get_holidays(
    year: Optional[str] = Query(None, description="年份，例如：2025")
):
    """
    获取假期数据
    
    参数:
    - year: 年份（可选，默认为当前年份）
    
    返回假期配置列表。数据来源：数据库 holiday 表。
    """
    if not year:
        year = str(datetime.now().year)
    try:
        rows = load_holidays_for_year(year)
        holidays = [
            Holiday(date=r["date"], type=r["type"], festival=r.get("festival") or None)
            for r in rows if r.get("date")
        ]
        return HolidayResponse(success=True, year=year, holidays=holidays)
    except Exception as e:
        logger.error(f"读取假期数据失败: {str(e)}")
        return HolidayResponse(success=True, year=year, holidays=[])


@router.post("/save", response_model=HolidayResponse)
async def save_holidays(
    year: str = Query(..., description="年份，例如：2025"),
    current_user: str = Query(..., description="当前操作人（需为打卡管理员）"),
    holidays: List[Holiday] = None,
):
    """
    保存某一年的假期与调休设置（覆盖该年的 holiday 表）。
    仅打卡管理员或系统管理员（webconfig.admin1）可操作。
    """
    admin1 = _get_admin1()
    dakaman = _get_dakaman()
    cu = (current_user or "").strip()
    if not (admin1 and cu == admin1) and not (dakaman and cu == dakaman):
        raise HTTPException(status_code=403, detail="仅打卡管理员或系统管理员可维护假期调休设置")
    year = (year or "").strip()
    if not year:
        raise HTTPException(status_code=400, detail="年份不能为空")
    try:
        y_int = int(year)
    except ValueError:
        raise HTTPException(status_code=400, detail="年份格式不正确")

    try:
        # 简单做法：先删除该年所有记录，再按当前提交的数据重建
        db.execute_update("DELETE FROM holiday WHERE year = %s", (y_int,))
        if holidays:
            for h in holidays:
                date_str = (h.date or "").strip()
                type_str = (h.type or "").strip()
                if not date_str:
                    continue
                festival_str = _normalize_festival_name((getattr(h, "festival", None) or "").strip())
                try:
                    db.execute_update(
                        "INSERT INTO holiday (year, date, type, festival) VALUES (%s, %s, %s, %s)",
                        (y_int, date_str, type_str, festival_str),
                    )
                except Exception:
                    db.execute_update(
                        "INSERT INTO holiday (year, date, type) VALUES (%s, %s, %s)",
                        (y_int, date_str, type_str),
                    )
        # 返回最新数据
        rows = load_holidays_for_year(str(y_int))
        out = [
            Holiday(date=r["date"], type=r["type"], festival=r.get("festival") or None)
            for r in rows if r.get("date")
        ]
        return HolidayResponse(success=True, year=str(y_int), holidays=out)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"保存假期数据失败: {str(e)}")
        raise HTTPException(status_code=500, detail="保存假期数据失败，请稍后重试")


@router.get("/template")
async def download_holiday_template(
    year: str = Query(..., description="年份，例如：2025"),
    current_user: str = Query(..., description="当前操作人（需为打卡管理员）"),
):
    """
    下载某年的假期调休 Excel 模板。
    仅打卡管理员可操作。
    模板中预置元旦、春节、清明、五一、端午、中秋、国庆 7 个节日的大概日期行。
    """
    admin1 = _get_admin1()
    dakaman = _get_dakaman()
    cu = (current_user or "").strip()
    if not (admin1 and cu == admin1) and not (dakaman and cu == dakaman):
        raise HTTPException(status_code=403, detail="仅打卡管理员或系统管理员可下载假期模板")
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl，无法生成 Excel")
    year = (year or "").strip()
    try:
        y_int = int(year)
    except ValueError:
        raise HTTPException(status_code=400, detail="年份格式不正确")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{y_int}年假期模板"
        headers = ["日期", "类型"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 预置 7 个节假日的大概日期（用户可自行调整/增加），类型初始为“放假”
        approx = [
            (f"{y_int}-01-01", "放假"),
            (f"{y_int}-02-10", "放假"),
            (f"{y_int}-04-05", "放假"),
            (f"{y_int}-05-01", "放假"),
            (f"{y_int}-06-10", "放假"),
            (f"{y_int}-09-21", "放假"),
            (f"{y_int}-10-01", "放假"),
        ]
        for d, t in approx:
            ws.append([d, t])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename_ascii = f"holiday_template_{y_int}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename=\"{filename_ascii}\"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"生成假期模板失败: {str(e)}")
        raise HTTPException(status_code=500, detail="生成假期模板失败")


@router.post("/upload", response_model=HolidayResponse)
async def upload_holiday_file(
    year: str = Query(..., description="年份，例如：2025"),
    current_user: str = Query(..., description="当前操作人（需为打卡管理员）"),
    file: UploadFile = File(..., description="假期调休 Excel 模板文件"),
):
    """
    上传某年的假期调休 Excel 文件并写入 holiday 表（覆盖该年）。
    仅打卡管理员可操作。
    Excel 第一张表，前两列分别为：日期、类型。
    """
    admin1 = _get_admin1()
    dakaman = _get_dakaman()
    cu = (current_user or "").strip()
    if not (admin1 and cu == admin1) and not (dakaman and cu == dakaman):
        raise HTTPException(status_code=403, detail="仅打卡管理员或系统管理员可上传假期文件")
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl，无法读取 Excel")
    year = (year or "").strip()
    try:
        y_int = int(year)
    except ValueError:
        raise HTTPException(status_code=400, detail="年份格式不正确")

    try:
        contents = await file.read()
        buf = BytesIO(contents)
        wb = load_workbook(buf, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        holidays: List[Holiday] = []
        first = True
        from datetime import date as date_type
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue  # 跳过表头
            if not row:
                continue
            raw_date = row[0]
            raw_type = row[1] if len(row) > 1 else ""
            if not raw_date:
                continue
            # 转为 yyyy-MM-dd 字符串
            if isinstance(raw_date, (datetime, date_type)):
                d_str = raw_date.strftime("%Y-%m-%d")
            else:
                s = str(raw_date).strip()
                # 若仅有 MM-DD 或 M/D，则补全年份
                if len(s) <= 5 and "-" in s:
                    d_str = f"{y_int}-{s}"
                elif "/" in s and len(s) <= 5:
                    parts = s.split("/")
                    if len(parts) >= 2:
                        m = parts[0].zfill(2)
                        d = parts[1].zfill(2)
                        d_str = f"{y_int}-{m}-{d}"
                    else:
                        d_str = s
                else:
                    d_str = s
            t_str = str(raw_type or "").strip() or "放假"
            holidays.append(Holiday(date=d_str, type=t_str))

        # 复用保存逻辑
        return await save_holidays(year=str(y_int), current_user=current_user, holidays=holidays)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"上传假期文件失败: {str(e)}")
        raise HTTPException(status_code=500, detail="上传假期文件失败，请检查文件格式")


def _parse_llm_json_content(content: str) -> dict:
    """去掉大模型返回中的 markdown 代码块包裹并解析 JSON。"""
    if not content:
        raise json.JSONDecodeError("empty", "", 0)
    
    text = content.strip()
    
    # 0. 首先去除<think>标签（Qwen模型特有）
    if '<think>' in text.lower() and '</think>' in text.lower():
        # 提取<think>标签之后的内容
        think_end = text.lower().find('</think>') + 8  # </think>长度
        text = text[think_end:].strip()
    
    # 1. 尝试多种方法提取JSON
    
    # 方法1: 直接解析
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    
    # 方法2: 查找最长的有效JSON对象
    # 找到所有{和}的位置
    stack = []
    json_start = -1
    json_end = -1
    max_length = 0
    best_start = -1
    best_end = -1
    
    for i, char in enumerate(text):
        if char == '{':
            if not stack:
                json_start = i
            stack.append(i)
        elif char == '}':
            if stack:
                start = stack.pop()
                if not stack:  # 找到完整的JSON对象
                    json_end = i
                    length = json_end - json_start + 1
                    if length > max_length:
                        max_length = length
                        best_start = json_start
                        best_end = json_end
    
    if best_start != -1 and best_end != -1:
        json_str = text[best_start:best_end+1]
        try:
            return json.loads(json_str)
        except json.JSONDecodeError:
            pass
    
    # 方法3: 尝试去除markdown代码块
    if text.startswith("```"):
        lines = text.split('\n')
        if len(lines) > 1:
            text = '\n'.join(lines[1:])
        if text.endswith("```"):
            text = text[:-3].strip()
        else:
            text = text.replace("```", "").strip()
    
    # 方法4: 简单查找第一个{和最后一个}
    start_idx = text.find('{')
    end_idx = text.rfind('}')
    
    if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
        json_str = text[start_idx:end_idx+1]
        try:
            return json.loads(json_str)
        except json.JSONDecodeError:
            pass
    
    # 方法5: 尝试清理常见问题
    # 移除可能的中文标点后的内容
    import re
    # 查找JSON模式
    json_pattern = r'\{[^{}]*\{[^{}]*\}[^{}]*\}|\{[^{}]*\}'
    matches = re.findall(json_pattern, text, re.DOTALL)
    
    for match in matches:
        try:
            # 尝试清理匹配的文本
            clean_match = match.strip()
            # 移除开头和结尾的非JSON字符
            while clean_match and not clean_match.startswith('{'):
                clean_match = clean_match[1:]
            while clean_match and not clean_match.endswith('}'):
                clean_match = clean_match[:-1]
            
            if clean_match:
                return json.loads(clean_match)
        except json.JSONDecodeError:
            continue
    
    # 所有方法都失败
    logger.warning(f"JSON解析失败，原始内容前500字符: {content[:500]}")
    raise json.JSONDecodeError("无法解析JSON", content, 0)


def _infer_festival_from_date(date_str: str) -> str:
    """
    根据日期推断节日名称（仅处理固定日期的节日），用于大模型未返回 festival 时的兜底。
    日期格式 yyyy-MM-dd，只取 MM-dd 做匹配。
    """
    if not date_str or len(date_str) < 10:
        return ""
    try:
        md = date_str.strip()[-5:]  # "MM-dd"
        fixed = {
            "01-01": "元旦",
            "04-04": "清明",
            "04-05": "清明",
            "05-01": "劳动节",
            "10-01": "国庆节",
            "10-02": "国庆节",
            "10-03": "国庆节",
            "10-04": "国庆节",
            "10-05": "国庆节",
            "10-06": "国庆节",
            "10-07": "国庆节",
        }
        return fixed.get(md, "")
    except Exception:
        return ""


def _normalize_festival_name(name: str) -> str:
    """规范化节日名称，兼容模型返回“清明节/元旦节”等别名。"""
    text = (name or "").strip()
    if not text:
        return ""
    alias = {
        "元旦节": "元旦",
        "清明节": "清明",
        "劳动": "劳动节",
        "五一": "劳动节",
        "五一劳动节": "劳动节",
        "端午": "端午节",
        "中秋": "中秋节",
        "国庆": "国庆节",
    }
    return alias.get(text, text)


class HolidayParseRequest(BaseModel):
    year: str
    current_user: str
    text: str


def _build_holiday_prompts(y_int: int, text: str) -> tuple:
    """构建假期解析的 system/user prompt，附带 Qwen3 的 /no_think 关闭思考模式。"""
    system_prompt = (
        "你是一个假期与调休解析助手。用户会给你一整段中文放假通知，"
        "请根据其中描述，推导出这一年内每一天是“放假”还是“上班”或“正常工作日”。\n"
        "但最终只需要输出有特殊安排的日期（放假或调休上班），普通工作日不要输出。\n"
        "注意：\n"
        "1. 文中写“放假调休”通常表示一段连续假期，其中既包含法定节假日，也可能包含周末/工作日调休上班；\n"
        "2. 文中写“X 月 Y 日（周六/周日）上班”视为“上班”；\n"
        "3. 如果一句话提到了利用某个节日假期调休，例如“利用冰雪节假日调休 1 天”，"
        "   只要有明确日期，也按放假或调休上班标记到对应日期；\n"
        "4. 节日名称请尽量归纳为：元旦、春节、清明、劳动节、端午节、中秋节、国庆节、高温防暑休假 等简短中文。\n"
        "5. **重要**：不要输出任何思考过程或解释，只返回要求的 JSON 格式。不要使用 <think> 标签或其他标记。\n"
        "/no_think"
    )
    user_prompt = (
        f"年份：{y_int}\n\n"
        f"放假通知原文如下：\n{text}\n\n"
        "请按照下面 JSON 格式输出：\n"
        "{\n"
        '  \"year\": 2025,\n'
        '  \"days\": [\n'
        '    { \"date\": \"2025-01-01\", \"type\": \"放假\", \"festival\": \"元旦\" },\n'
        '    { \"date\": \"2025-01-04\", \"type\": \"上班\", \"festival\": \"元旦\" },\n'
        "    ... 只包含放假日和调休上班日 ...\n"
        "  ]\n"
        "}\n"
        "要求：\n"
        "1. date 固定为 yyyy-MM-dd 格式；\n"
        "2. type 只能是 \"放假\" 或 \"上班\" 两种；\n"
        "3. 每条记录必须包含 festival 且不能为空，为简短中文节日名称，例如 元旦/春节/清明/劳动节/端午节/中秋节/国庆节/高温防暑休假；\n"
        "4. 不要包含任何多余字段或解释，只返回一个 JSON 对象。\n"
        "/no_think"
    )
    return system_prompt, user_prompt


def _save_holidays_from_data(y_int: int, data: dict) -> List[Holiday]:
    """根据大模型解析后的 JSON 写入 holiday 表并返回最新的 Holiday 列表。"""
    days = data.get("days") or []
    holidays: List[Holiday] = []
    for d in days:
        date_str = str(d.get("date") or "").strip()
        type_str = str(d.get("type") or "").strip() or "放假"
        festival_str = _normalize_festival_name(str(d.get("festival") or "").strip())
        if not festival_str:
            festival_str = _infer_festival_from_date(date_str)
        if not date_str:
            continue
        holidays.append(Holiday(date=date_str, type=type_str, festival=festival_str or None))

    db.execute_update("DELETE FROM holiday WHERE year = %s", (y_int,))
    for h in holidays:
        date_str = (h.date or "").strip()
        type_str = (h.type or "").strip()
        if not date_str:
            continue
        festival_str = _normalize_festival_name((getattr(h, "festival", None) or "").strip())
        try:
            db.execute_update(
                "INSERT INTO holiday (year, date, type, festival) VALUES (%s, %s, %s, %s)",
                (y_int, date_str, type_str, festival_str),
            )
        except Exception:
            db.execute_update(
                "INSERT INTO holiday (year, date, type) VALUES (%s, %s, %s)",
                (y_int, date_str, type_str),
            )
    rows = load_holidays_for_year(str(y_int))
    return [
        Holiday(date=r["date"], type=r["type"], festival=r.get("festival") or None)
        for r in rows if r.get("date")
    ]


def _validate_parse_request(req: HolidayParseRequest) -> int:
    """校验解析请求并返回年份整数。"""
    admin1 = _get_admin1()
    dakaman = _get_dakaman()
    current_user = (req.current_user or "").strip()
    if not (admin1 and current_user == admin1) and not (dakaman and current_user == dakaman):
        raise HTTPException(status_code=403, detail="仅打卡管理员或系统管理员可使用大模型解析假期")
    year = (req.year or "").strip()
    try:
        return int(year)
    except ValueError:
        raise HTTPException(status_code=400, detail="年份格式不正确")


@router.post("/parse-text", response_model=HolidayResponse)
async def parse_holiday_text(req: HolidayParseRequest):
    """
    使用本地大模型解析一段放假通知文本，自动生成并保存某年的假期与调休设置。
    仅打卡管理员或系统管理员可操作。
    仅使用 webconfig.llm_base_url / llm_model 配置的本地大模型（Ollama/OpenAI 兼容接口）。
    """
    y_int = _validate_parse_request(req)

    try:
        from openai import OpenAI
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端未安装 openai SDK，无法调用大模型")

    config = _get_llm_config()
    if not config.get("base_url") or not config.get("model"):
        raise HTTPException(
            status_code=500,
            detail="未配置本地大模型：请在 webconfig 中设置 llm_base_url（如 http://host:11434/v1）和 llm_model（如 qwen3:8b）",
        )

    system_prompt, user_prompt = _build_holiday_prompts(y_int, req.text or "")

    content = ""
    try:
        local_client = OpenAI(base_url=config["base_url"], api_key="ollama")
        completion = local_client.chat.completions.create(
            model=config["model"],
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0,
            stream=False,
            extra_body={"chat_template_kwargs": {"enable_thinking": False}},
        )
        content = (completion.choices[0].message.content or "").strip()
        data = _parse_llm_json_content(content)
    except json.JSONDecodeError as e:
        logger.warning("JSON 解析失败，原始内容: %s, 错误: %s", content[:500] if content else "空响应", e)
        raise HTTPException(status_code=500, detail="大模型返回内容无法解析为 JSON，请检查模型与通知原文")
    except Exception as e:
        logger.error("本地大模型调用失败: %s", e)
        raise HTTPException(status_code=500, detail=f"本地大模型调用失败：{e}")

    try:
        out = _save_holidays_from_data(y_int, data)
        return HolidayResponse(success=True, year=str(y_int), holidays=out)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("解析或保存假期数据失败")
        raise HTTPException(status_code=500, detail=f"解析大模型返回的假期数据失败: {str(e)}")


# ==================== 流式解析接口 ====================


def _sse_event(payload: dict) -> str:
    """将字典打包为一条 SSE 事件。"""
    return f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"


@router.post("/parse-text-stream")
async def parse_holiday_text_stream(req: HolidayParseRequest):
    """
    使用本地大模型流式解析放假通知。
    以 SSE (text/event-stream) 形式返回：
    - {"type": "meta", "model": "...", "base_url": "..."}
    - {"type": "chunk", "text": "增量 token"}
    - {"type": "done", "success": true, "year": "2025", "holidays": [...]}
    - {"type": "error", "message": "..."}
    """
    y_int = _validate_parse_request(req)
    try:
        from openai import OpenAI
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端未安装 openai SDK，无法调用大模型")

    config = _get_llm_config()
    if not config.get("base_url") or not config.get("model"):
        raise HTTPException(
            status_code=500,
            detail="未配置本地大模型：请在 webconfig 中设置 llm_base_url（如 http://host:11434/v1）和 llm_model（如 qwen3:8b）",
        )

    system_prompt, user_prompt = _build_holiday_prompts(y_int, req.text or "")

    def gen():
        yield _sse_event({
            "type": "meta",
            "model": config["model"],
            "base_url": config["base_url"],
        })
        accumulated: List[str] = []
        try:
            client = OpenAI(base_url=config["base_url"], api_key="ollama")
            stream = client.chat.completions.create(
                model=config["model"],
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0,
                stream=True,
                extra_body={"chat_template_kwargs": {"enable_thinking": False}},
            )
            for event in stream:
                try:
                    delta = event.choices[0].delta
                    piece = (getattr(delta, "content", None) or "")
                except Exception:
                    piece = ""
                if not piece:
                    continue
                accumulated.append(piece)
                yield _sse_event({"type": "chunk", "text": piece})
        except Exception as e:
            logger.error("本地大模型流式调用失败: %s", e)
            yield _sse_event({"type": "error", "message": f"本地大模型调用失败：{e}"})
            return

        content = "".join(accumulated).strip()
        if not content:
            yield _sse_event({"type": "error", "message": "大模型未返回任何内容"})
            return

        try:
            data = _parse_llm_json_content(content)
        except Exception as e:
            logger.warning("JSON 解析失败，原始内容: %s, 错误: %s", content[:500], e)
            yield _sse_event({"type": "error", "message": "大模型返回内容无法解析为 JSON"})
            return

        try:
            out = _save_holidays_from_data(y_int, data)
            yield _sse_event({
                "type": "done",
                "success": True,
                "year": str(y_int),
                "holidays": [
                    {"date": h.date, "type": h.type, "festival": h.festival}
                    for h in out
                ],
            })
        except Exception as e:
            logger.exception("解析或保存假期数据失败")
            yield _sse_event({"type": "error", "message": f"保存假期数据失败：{e}"})

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )
