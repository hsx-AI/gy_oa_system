# -*- coding: utf-8 -*-
"""
考勤数据管理API路由 - 新版（使用SQLite）
"""
from fastapi import APIRouter, BackgroundTasks, File, UploadFile, HTTPException, Query, Form
from typing import Optional, List
from pydantic import BaseModel
from datetime import datetime, date as date_type, timedelta
import os
import tempfile
import logging
from io import BytesIO

from fastapi.responses import StreamingResponse

import uuid
from config import settings
from attendance_db import attendance_db
from database import db
from utils.excel_processor import ExcelProcessor
from utils.helpers import normalize_qj_tian_days
from routers.suggestions import (
    _is_female_employee,
    _is_march8_pm_interval,
    _suggestion_handled,
    _suggestion_under_review,
    get_attendance_exception_keys,
)
from routers.approvers import _get_user_info, _jb_match
from routers.db_manager import _get_admin1

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/attendance", tags=["考勤管理"])


def _attendance_report_httpx_timeout():
    """打卡服务器 GET 拉取报表：远端生成慢、本机卡顿时可能长时间无字节返回，单独放宽超时（手动「上传最新数据」与定时任务共用）。"""
    import httpx

    # 读超时 20 分钟：允许长时间挂起后仍返回；连接 3 分钟应对极慢网络/系统卡顿
    return httpx.Timeout(connect=180.0, read=1200.0, write=180.0, pool=180.0)


def _get_dakaman() -> Optional[str]:
    """从 webconfig 表读取 dakaman 字段（打卡数据上传权限用户名）。"""
    try:
        rows = db.execute_query("SELECT dakaman FROM webconfig WHERE id = %s LIMIT 1", ("1",))
        if rows and rows[0].get("dakaman") is not None:
            return (rows[0]["dakaman"] or "").strip() or None
    except Exception as e:
        logger.debug(f"读取 webconfig.dakaman 失败: {e}")
    return None




def _can_see_attendance_exceptions(current_user: str) -> tuple:
    """
    判断当前用户是否有权查看考勤异常。
    返回 (allowed: bool, lsys: str|None, is_dakaman: bool, include_buban: bool)。
    - 系统管理员(admin1)、打卡管理员(dakaman)：可看全部含部办；is_dakaman True。
    - 部长/副部长：可看全部含部办；is_dakaman False（无代处理列）。
    - 班组长/主任/副主任：仅本室；include_buban False。
    """
    current_user = (current_user or "").strip()
    if not current_user:
        return False, None, False, False
    dakaman = _get_dakaman()
    is_dk = bool(dakaman and current_user == dakaman)
    admin1 = _get_admin1()
    if admin1 and current_user == admin1:
        return True, None, True, True
    if is_dk:
        return True, None, True, True
    user = _get_user_info(current_user)
    if not user:
        return False, None, False, False
    jb = (user.get("jb") or "").strip()
    if _jb_match(jb, "部长") or _jb_match(jb, "副部长"):
        return True, None, False, True
    if _jb_match(jb, "组长") or _jb_match(jb, "主任") or _jb_match(jb, "副主任"):
        lsys = (user.get("lsys") or "").strip()
        return True, lsys if lsys else None, False, False
    return False, None, False, False


def _can_export_suggestion_attendance_report_all(current_user: str) -> bool:
    user = _get_user_info((current_user or "").strip())
    if not user:
        return False
    jb = (user.get("jb") or "").strip()
    lsys = (user.get("lsys") or "").strip()
    return lsys == "综合技术室" and (_jb_match(jb, "主任") or _jb_match(jb, "副主任"))


def _parse_date_only(val) -> Optional[date_type]:
    """Convert DB date/datetime/string values to a date for attendance exception filtering."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date_type):
        return val
    s = str(val).strip()
    if not s:
        return None
    if "." in s:
        s = s.split(".")[0]
    s = s[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def _date_in_range(day: date_type, start_val, end_val) -> bool:
    start_day = _parse_date_only(start_val)
    end_day = _parse_date_only(end_val) or start_day
    if not day or not start_day or not end_day:
        return False
    if end_day < start_day:
        start_day, end_day = end_day, start_day
    return start_day <= day <= end_day


def _filter_exception_keys_without_pending_process(exception_keys: List[tuple], year: int, month: int) -> List[tuple]:
    """
    考勤异常列表最终兜底过滤：同一天已有未驳回的请假/公出申请时，
    视为已在处理，不再展示为待处理考勤异常。
    """
    if not exception_keys:
        return []
    names = sorted({(n or "").strip() for n, _, _ in exception_keys if (n or "").strip()})
    if not names:
        return exception_keys

    import calendar

    _, last_day = calendar.monthrange(year, month)
    month_start = f"{year}-{month:02d}-01"
    month_end = f"{year}-{month:02d}-{last_day:02d} 23:59:59"
    pending_by_name = {}

    try:
        ph = ",".join(["%s"] * len(names))
        qj_rows = db.execute_query(
            f"""
            SELECT TRIM(xm) AS xm, timefrom, timeto
            FROM qj
            WHERE TRIM(xm) IN ({ph})
              AND COALESCE(qjzt, 0) != 22
              AND timefrom <= %s
              AND timeto >= %s
            """,
            tuple(names) + (month_end, month_start),
        ) or []
        for r in qj_rows:
            n = (r.get("xm") or "").strip()
            if n:
                pending_by_name.setdefault(n, []).append((r.get("timefrom"), r.get("timeto")))

        trip_rows = db.execute_query(
            f"""
            SELECT TRIM(gcr) AS xm,
                   COALESCE(yjcfsj, gcsj, wpsj, yjfhsj, sjfhtime) AS timefrom,
                   COALESCE(yjfhsj, sjfhtime, yjcfsj, gcsj, wpsj) AS timeto
            FROM gcsqb
            WHERE TRIM(gcr) IN ({ph})
              AND NOT (COALESCE(bldzt, 0) = 22 OR COALESCE(szrzt, 0) = 22)
              AND COALESCE(yjcfsj, gcsj, wpsj, yjfhsj, sjfhtime) <= %s
              AND COALESCE(yjfhsj, sjfhtime, yjcfsj, gcsj, wpsj) >= %s
            """,
            tuple(names) + (month_end, month_start),
        ) or []
        for r in trip_rows:
            n = (r.get("xm") or "").strip()
            if n:
                pending_by_name.setdefault(n, []).append((r.get("timefrom"), r.get("timeto")))

        kqyc_rows = db.execute_query(
            f"""
            SELECT TRIM(applicant) AS xm,
                   CONCAT(attendance_date, ' ', time_from) AS timefrom,
                   CONCAT(attendance_date, ' ', time_to) AS timeto
            FROM attendance_exception
            WHERE TRIM(applicant) IN ({ph})
              AND first_status != 2
              AND second_status != 2
              AND attendance_date >= %s
              AND attendance_date <= %s
            """,
            tuple(names) + (month_start[:10], month_end[:10]),
        ) or []
        for r in kqyc_rows:
            n = (r.get("xm") or "").strip()
            if n:
                pending_by_name.setdefault(n, []).append((r.get("timefrom"), r.get("timeto")))
    except Exception as e:
        logger.warning(f"考勤异常列表过滤审核中请假/公出失败，保留原异常数据: {e}")
        return exception_keys

    filtered = []
    for name, dept, date_str in exception_keys:
        day = _parse_date_only(date_str)
        n = (name or "").strip()
        if day and any(_date_in_range(day, s, e) for s, e in pending_by_name.get(n, [])):
            continue
        filtered.append((name, dept, date_str))
    return filtered


def _build_attendance_exceptions_data(year: int, month: int, filter_lsys: Optional[str], include_buban: bool = False) -> List[dict]:
    """
    构建指定年月的考勤异常列表原始数据（不做权限检查）。
    filter_lsys 不为空时，仅保留该科室(department)的数据。
    include_buban=True 时包含部办人员。
    """
    import calendar

    _, last_day = calendar.monthrange(year, month)
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{last_day:02d}"
    exception_keys = get_attendance_exception_keys(year, month, include_buban=include_buban)
    exception_keys = _filter_exception_keys_without_pending_process(exception_keys, year, month)
    if not exception_keys:
        return []
    if filter_lsys:
        exception_keys = [
            (n, d, dt) for n, d, dt in exception_keys
            if (d or "").strip() == filter_lsys
        ]
        if not exception_keys:
            return []
    all_records = attendance_db.get_all_records_by_date_range(start_date, end_date)
    records_by_key = {}
    for r in all_records:
        name = (r.get("employee_name") or "").strip()
        dept = (r.get("department") or "").strip()
        d = (r.get("attendance_date") or "")
        if hasattr(d, "strftime"):
            d = d.strftime("%Y-%m-%d")
        else:
            d = str(d)[:10]
        key = (name, dept, d)
        if key not in records_by_key:
            records_by_key[key] = r
    built: List[dict] = []
    for name, dept, date_str in exception_keys:
        key = (name, dept, date_str)
        if key in records_by_key:
            rec = dict(records_by_key[key])
            rec["full_day_absence"] = False
            built.append(rec)
        else:
            built.append({
                "id": None,
                "employee_id": "",
                "employee_name": name,
                "department": dept,
                "attendance_date": date_str,
                "time_1": None, "time_2": None, "time_3": None, "time_4": None,
                "time_5": None, "time_6": None, "time_7": None, "time_8": None,
                "time_9": None, "time_10": None,
                "full_day_absence": True,
            })
    built.sort(key=lambda x: (x.get("attendance_date") or "", x.get("employee_name") or ""))
    return built


# ==================== 数据模型 ====================

class AttendanceRecord(BaseModel):
    """考勤记录模型"""
    # 数据库中 id 目前为 VARCHAR(36) UUID，因此这里用 str 接收；前端不依赖该字段类型
    id: Optional[str] = None
    employee_id: str
    employee_name: str
    department: str
    attendance_date: str
    time_1: Optional[str] = None
    time_1_mark: Optional[int] = None
    time_2: Optional[str] = None
    time_2_mark: Optional[int] = None
    time_3: Optional[str] = None
    time_3_mark: Optional[int] = None
    time_4: Optional[str] = None
    time_4_mark: Optional[int] = None
    time_5: Optional[str] = None
    time_5_mark: Optional[int] = None
    time_6: Optional[str] = None
    time_6_mark: Optional[int] = None
    time_7: Optional[str] = None
    time_7_mark: Optional[int] = None
    time_8: Optional[str] = None
    time_8_mark: Optional[int] = None
    time_9: Optional[str] = None
    time_9_mark: Optional[int] = None
    time_10: Optional[str] = None
    time_10_mark: Optional[int] = None
    # 考勤异常接口专用：无打卡记录时为 True，前端显示「全天缺勤」
    full_day_absence: Optional[bool] = None


class AttendanceQueryResponse(BaseModel):
    """考勤查询响应"""
    success: bool
    message: Optional[str] = None
    total: int = 0
    data: List[AttendanceRecord] = []
    is_dakaman: bool = False


class UploadResponse(BaseModel):
    """上传响应"""
    success: bool
    message: str
    records_count: int = 0
    success_count: int = 0
    fail_count: int = 0


# ==================== API 路由 ====================


@router.get("/upload/config")
async def get_upload_config():
    """
    获取打卡/人事相关配置。返回 dakaman（打卡管理员）、admin2（人事管理员），前端用于权限展示。
    """
    dakaman = _get_dakaman()
    admin2 = ""
    try:
        wc = db.execute_query("SELECT admin2 FROM webconfig WHERE id = 1 LIMIT 1")
        if wc and wc[0].get("admin2") is not None:
            admin2 = (wc[0]["admin2"] or "").strip() or ""
    except Exception:
        pass
    fetch_url = (getattr(settings, "ATTENDANCE_REPORT_FETCH_URL", None) or "").strip()
    admin1 = _get_admin1()
    personnel_archive_url = (getattr(settings, "PERSONNEL_ARCHIVE_URL", None) or "").strip()
    from routers.attendance_scheduler_config import (
        get_manual_upload_default_cutoff,
        resolve_suggestion_cutoff_date,
    )
    cutoff_mode = get_manual_upload_default_cutoff()
    return {
        "success": True,
        "dakaman": dakaman or "",
        "admin2": admin2,
        "admin1": admin1 or "",
        "fetchReportUrl": fetch_url,
        "personnelArchiveUrl": personnel_archive_url,
        "suggestionCutoff": cutoff_mode,
        "suggestionCutoffLabel": "今日" if cutoff_mode == "today" else "前一日",
        "attendanceDataDate": resolve_suggestion_cutoff_date(cutoff_mode),
    }


_EXCLUDED_LSYS_FOR_ATTENDANCE = {"其他部门员工", "其他部门成员"}


def _yggl_employees_for_suggestions() -> List[tuple]:
    """在职、有姓名与隶属科室的员工，与打卡入库时 department=lsys 一致。
    排除"其他部门员工/成员"——不参与考勤统计。"""
    try:
        rows = db.execute_query(
            """
            SELECT TRIM(name) AS name, TRIM(lsys) AS lsys FROM yggl
            WHERE name IS NOT NULL AND TRIM(name) != ''
              AND lsys IS NOT NULL AND TRIM(lsys) != ''
              AND (COALESCE(zaizhi, 0) = 0)
            """
        )
    except Exception as e:
        logger.warning(f"[后台] 查询 yggl 员工列表失败: {e}")
        return []
    out = []
    for r in rows or []:
        n = (r.get("name") or "").strip()
        d = (r.get("lsys") or "").strip()
        if n and d and d not in _EXCLUDED_LSYS_FOR_ATTENDANCE:
            out.append((n, d))
    return out


def _generate_suggestions_bg(records: list, cutoff_date_str: str = None):
    """后台任务：上传打卡后，按涉及月份为全员（yggl 在职）重算智能建议，不阻塞上传响应。
    当月库中无打卡记录的人也会生成建议（如工作日全天缺勤），避免仅处理「本次文件里出现过的员工」。
    cutoff_date_str: 'YYYY-MM-DD'，当月仅生成截止到此日期的建议。
    优化：按月份批量查询考勤记录（1条SQL/月），缓存假期数据，批量写入建议。"""
    import time as _time
    from collections import defaultdict
    t0 = _time.time()
    try:
        from routers.suggestions import (generate_suggestions_for_month_with_records,
                                         load_holidays, _load_holiday_festival_map, _parse_record_date)
        attendance_db.ensure_suggestions_table()

        seen = set()
        for rec in records:
            name = (rec.get("employee_name") or "").strip()
            dept = (rec.get("department") or "").strip()
            ad = rec.get("attendance_date")
            if not name or not dept or not ad:
                continue
            y, m = None, None
            if isinstance(ad, datetime):
                y, m = ad.year, ad.month
            elif isinstance(ad, date_type):
                y, m = ad.year, ad.month
            elif isinstance(ad, str):
                parts = ad.replace("/", "-").split("-")
                if len(parts) >= 2:
                    try:
                        y, m = int(parts[0]), int(parts[1])
                    except (ValueError, IndexError):
                        continue
                else:
                    continue
            else:
                continue
            seen.add((name, dept, y, m))

        if not seen:
            return

        months = set()
        for (_, _, y, m) in seen:
            months.add((y, m))

        employees = _yggl_employees_for_suggestions()
        keys_to_process = set()
        for y, m in months:
            for name, dept in employees:
                keys_to_process.add((name, dept, y, m))

        if not keys_to_process:
            return

        # 避免单次 SQL 过长，分批删除旧建议
        key_list = list(keys_to_process)
        _chunk = 400
        for i in range(0, len(key_list), _chunk):
            attendance_db.delete_suggestions_batch(key_list[i : i + _chunk])

        holidays_cache: dict = {}
        holiday_festival_cache: dict = {}
        month_records: dict = {}
        for (y, m) in months:
            start_date = f"{y}-{m:02d}-01"
            if m == 12:
                end_date = f"{y}-12-31"
            else:
                last = (date_type(y, m + 1, 1) - timedelta(days=1))
                end_date = last.strftime("%Y-%m-%d")
            all_recs = attendance_db.get_all_records_by_date_range(start_date, end_date)
            grouped = defaultdict(list)
            for r in all_recs:
                key = ((r.get("employee_name") or "").strip(),
                       (r.get("department") or "").strip())
                grouped[key].append(r)
            month_records[(y, m)] = grouped

            year_str = str(y)
            if year_str not in holidays_cache:
                holidays_cache[year_str] = load_holidays(year_str)
            if year_str not in holiday_festival_cache:
                holiday_festival_cache[year_str] = _load_holiday_festival_map(y)

        for (name, dept, y, m) in keys_to_process:
            try:
                person_records = month_records.get((y, m), {}).get((name, dept), [])
                holidays = holidays_cache[str(y)]
                holiday_festival_map = holiday_festival_cache[str(y)]
                suggestions_list = generate_suggestions_for_month_with_records(
                    name, dept, y, m, person_records, holidays,
                    cutoff_date_str=cutoff_date_str,
                    holiday_festival_map=holiday_festival_map)
                attendance_db.insert_suggestions(name, dept, y, m, suggestions_list)
            except Exception as e:
                logger.warning(f"[后台] 生成智能建议失败 {(name, dept, y, m)}: {e}")

        elapsed = round(_time.time() - t0, 1)
        logger.info(f"[后台] 智能建议生成完成，共处理 {len(keys_to_process)} 个人月组合，耗时 {elapsed}s")
    except Exception as e:
        logger.warning(f"[后台] 上传后生成智能建议失败: {e}")


def _process_attendance_file_path(temp_file_path: str, filename: str):
    """
    处理考勤文件（Excel），返回 (success, message, records_count, success_count, fail_count, mapped_records)。
    不负责 log_upload、background_tasks、临时文件删除。
    """
    processor = ExcelProcessor(temp_file_path)
    success, merged_records, error_msg = processor.process_file(start_row=6)
    if not success:
        return False, error_msg, 0, 0, 0, []

    mapped_records = []
    skipped_gh = []
    for rec in merged_records:
        gh = (rec.get("employee_id") or "").strip()
        emp = attendance_db.get_employee_by_gh(gh) if gh else None
        if not emp:
            skipped_gh.append(gh or "(空)")
            continue
        rec["employee_name"] = emp.get("name") or ""
        rec["department"] = emp.get("lsys") or ""
        mapped_records.append(rec)
    if skipped_gh:
        unique_gh = sorted(set(skipped_gh))
        logger.warning(
            f"上传跳过未在 yggl 中匹配到的工号，共 {len(skipped_gh)} 条记录涉及 {len(unique_gh)} 个工号。"
            f"未匹配工号完整列表: {unique_gh}"
        )

    success_count, fail_count = attendance_db.batch_insert_records(mapped_records)
    parts = [f"文件处理完成！Excel 合并后 {len(merged_records)} 人天"]
    if skipped_gh:
        parts.append(f"跳过未匹配工号 {len(skipped_gh)} 条（涉及 {len(set(skipped_gh))} 个工号: {', '.join(sorted(set(skipped_gh))[:20])}）")
    parts.append(f"入库成功 {success_count} 条")
    if fail_count:
        parts.append(f"入库失败 {fail_count} 条（详见服务器日志）")
    msg = "，".join(parts)
    return True, msg, len(mapped_records), success_count, fail_count, mapped_records


@router.post("/upload", response_model=UploadResponse)
async def upload_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    uploader: Optional[str] = Form(None),
    attendance_data_date: Optional[str] = Form(None),
):
    """
    上传并处理考勤Excel文件。仅 webconfig 表中 dakaman 对应用户可上传。
    
    文件格式要求：
    - Excel格式（.xls 或 .xlsx）
    - 从第6行开始读取数据
    - 列结构：A~F 员工编号、姓名、部门1、部门2、考勤日期、考勤时间；H 列为进出标记（可选）
    
    处理逻辑：
    - 自动按员工编号和日期合并打卡记录
    - 同一人同一天的多次打卡合并为一行（最多 10 次）
    - 仅无进出标识且间隔 ≤5 秒的重复刷卡会去重；H 列有进/出的全部保留
    - H 列含「进/入」记为进、含「出」记为出；无关键字则按当日时间顺序与前一条交替（首条视为进）
    """
    dakaman = _get_dakaman()
    admin1 = _get_admin1()
    uploader_name = (uploader or "").strip()
    if not (admin1 and uploader_name == admin1) and not (dakaman and uploader_name == dakaman):
        raise HTTPException(
            status_code=403,
            detail="仅打卡管理员（webconfig.dakaman）或系统管理员（webconfig.admin1）可上传打卡数据"
        )

    cutoff = (attendance_data_date or "").strip() or None

    # 验证文件类型
    if not file.filename.endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="只支持 .xls 或 .xlsx 格式的Excel文件")
    
    # 创建临时文件保存上传的内容
    temp_file = None
    try:
        # 保存上传文件到临时目录
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name
        
        logger.info(f"开始处理文件: {file.filename}")
        success, message, records_count, success_count, fail_count, mapped_records = _process_attendance_file_path(temp_file_path, file.filename)
        if not success:
            attendance_db.log_upload(file.filename, 0, "失败", message)
            return UploadResponse(success=False, message=message, records_count=0)
        attendance_db.log_upload(file.filename, records_count, "成功", f"成功: {success_count}, 失败: {fail_count}")
        background_tasks.add_task(_generate_suggestions_bg, list(mapped_records), cutoff)
        return UploadResponse(success=True, message=message, records_count=records_count, success_count=success_count, fail_count=fail_count)
    
    except Exception as e:
        error_msg = f"处理失败: {str(e)}"
        logger.error(error_msg)
        attendance_db.log_upload(file.filename, 0, "异常", error_msg)
        raise HTTPException(status_code=500, detail=error_msg)
    
    finally:
        # 清理临时文件
        if temp_file and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except:
                pass


@router.post("/fetch-and-upload", response_model=UploadResponse)
async def fetch_and_upload(
    background_tasks: BackgroundTasks,
    uploader: Optional[str] = Form(None),
    attendance_data_date: Optional[str] = Form(None),
):
    """
    从打卡服务器 GET 拉取最新报表并导入。仅 dakaman 可操作。
    需在 config 或 .env 中配置 ATTENDANCE_REPORT_FETCH_URL；远端响应可能很慢，服务端 HTTP 读超时约 20 分钟。
    """
    dakaman = _get_dakaman()
    admin1 = _get_admin1()
    uploader_name = (uploader or "").strip()
    if not (admin1 and uploader_name == admin1) and not (dakaman and uploader_name == dakaman):
        raise HTTPException(status_code=403, detail="仅打卡管理员或系统管理员可执行拉取上传")
    cutoff = (attendance_data_date or "").strip() or None
    fetch_url = (getattr(settings, "ATTENDANCE_REPORT_FETCH_URL", None) or "").strip()
    if not fetch_url:
        raise HTTPException(status_code=400, detail="未配置打卡报表拉取地址（ATTENDANCE_REPORT_FETCH_URL）")

    import httpx
    temp_file_path = None
    try:
        logger.info("开始从打卡服务器拉取最新报表: %s", fetch_url[:80] + "..." if len(fetch_url) > 80 else fetch_url)
        async with httpx.AsyncClient(timeout=_attendance_report_httpx_timeout()) as client:
            resp = await client.get(fetch_url)
            resp.raise_for_status()
            content = resp.content
        if not content:
            raise HTTPException(status_code=502, detail="拉取返回为空")
        # 根据 Content-Disposition 或默认使用 .xlsx
        suffix = ".xlsx"
        cd = (resp.headers.get("content-disposition") or "").strip()
        if "filename=" in cd:
            import re
            m = re.search(r'filename\*?=(?:UTF-8\'\')?["\']?([^"\';]+)["\']?', cd, re.I) or re.search(r'filename=([^;\s]+)', cd, re.I)
            if m:
                fn = m.group(1).strip().lower()
                if fn.endswith(".xls") and not fn.endswith(".xlsx"):
                    suffix = ".xls"
                elif fn.endswith(".xlsx"):
                    suffix = ".xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            temp_file_path = tmp.name
        filename = "report" + suffix
        success, message, records_count, success_count, fail_count, mapped_records = _process_attendance_file_path(temp_file_path, filename)
        if not success:
            attendance_db.log_upload(filename, 0, "失败", message)
            return UploadResponse(success=False, message=message, records_count=0)
        attendance_db.log_upload(filename, records_count, "成功", f"成功: {success_count}, 失败: {fail_count}")
        background_tasks.add_task(_generate_suggestions_bg, list(mapped_records), cutoff)
        return UploadResponse(success=True, message=message, records_count=records_count, success_count=success_count, fail_count=fail_count)
    except httpx.HTTPStatusError as e:
        attendance_db.log_upload("report", 0, "失败", str(e.response.status_code))
        raise HTTPException(status_code=502, detail=f"拉取失败: {e.response.status_code}")
    except httpx.RequestError as e:
        attendance_db.log_upload("report", 0, "失败", str(e))
        raise HTTPException(status_code=502, detail=f"拉取请求失败: {str(e)}")
    finally:
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except Exception:
                pass


async def run_fetch_and_upload_report(suggestion_cutoff: Optional[str] = None):
    """供定时任务调用：拉取报表并导入；每条任务的 suggestion_cutoff 决定智能建议截止日。失败时最多重试 3 次。"""
    import asyncio
    import httpx
    fetch_url = (getattr(settings, "ATTENDANCE_REPORT_FETCH_URL", None) or "").strip()
    if not fetch_url:
        logger.warning("[定时] 拉取跳过: 未配置 ATTENDANCE_REPORT_FETCH_URL")
        return
    dakaman = _get_dakaman()
    admin1 = _get_admin1()
    if not dakaman and not admin1:
        logger.warning("[定时] 拉取跳过: webconfig 中未配置 dakaman 或 admin1，无法执行定时任务")
        return
    max_attempts = 4  # 1 次首次 + 最多重试 3 次
    retry_delay_seconds = 30  # 每次失败后间隔 30 秒再重试
    last_error = None
    for attempt in range(1, max_attempts + 1):
        temp_file_path = None
        try:
            logger.info("[定时] 拉取打卡报表 第 %d/%d 次: %s", attempt, max_attempts, fetch_url[:80] + "..." if len(fetch_url) > 80 else fetch_url)
            async with httpx.AsyncClient(timeout=_attendance_report_httpx_timeout()) as client:
                resp = await client.get(fetch_url)
                resp.raise_for_status()
                content = resp.content
            if not content:
                last_error = "拉取返回为空"
                logger.warning("[定时] 第 %d 次: %s", attempt, last_error)
                if attempt < max_attempts:
                    await asyncio.sleep(retry_delay_seconds)
                continue
            suffix = ".xlsx"
            cd = (resp.headers.get("content-disposition") or "").strip()
            if "filename=" in cd:
                import re
                m = re.search(r'filename\*?=(?:UTF-8\'\')?["\']?([^"\';]+)["\']?', cd, re.I) or re.search(r'filename=([^;\s]+)', cd, re.I)
                if m:
                    fn = m.group(1).strip().lower()
                    if fn.endswith(".xls") and not fn.endswith(".xlsx"):
                        suffix = ".xls"
                    elif fn.endswith(".xlsx"):
                        suffix = ".xlsx"
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(content)
                temp_file_path = tmp.name
            report_name = "report" + suffix
            success, message, records_count, success_count, fail_count, mapped_records = _process_attendance_file_path(temp_file_path, report_name)
            if not success:
                last_error = message
                attendance_db.log_upload(report_name, 0, "失败", message)
                logger.warning("[定时] 第 %d 次 处理失败: %s", attempt, message)
                if attempt < max_attempts:
                    await asyncio.sleep(retry_delay_seconds)
                continue
            attendance_db.log_upload(report_name, records_count, "成功", f"成功: {success_count}, 失败: {fail_count}")
            from routers.attendance_scheduler_config import resolve_suggestion_cutoff_date
            cutoff_str = resolve_suggestion_cutoff_date(suggestion_cutoff)
            records_copy = list(mapped_records)
            loop = asyncio.get_event_loop()
            loop.run_in_executor(
                None,
                lambda rec=records_copy, cut=cutoff_str: _generate_suggestions_bg(rec, cut),
            )
            logger.info("[定时] 拉取上传完成（第 %d 次）: %s", attempt, message)
            return
        except Exception as e:
            last_error = str(e)
            logger.exception("[定时] 第 %d 次 拉取上传异常: %s", attempt, e)
            attendance_db.log_upload("report", 0, "异常", f"第{attempt}次: {last_error}")
            if attempt < max_attempts:
                await asyncio.sleep(retry_delay_seconds)
        finally:
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                except Exception:
                    pass
    logger.error("[定时] 拉取打卡报表已重试 %d 次均失败，最后错误: %s", max_attempts, last_error)


@router.get("/exceptions", response_model=AttendanceQueryResponse)
async def get_attendance_exceptions(
    year: int = Query(..., description="年份"),
    month: int = Query(..., ge=1, le=12, description="月份 1-12"),
    current_user: Optional[str] = Query(None, description="当前登录用户姓名，用于权限校验"),
):
    """
    考勤异常列表。权限：打卡管理员、部长/副部长可看全部（含部办）；各科室班组长/主任/副主任仅可看本室。
    返回指定年月内「智能建议需请假/缺勤且未完成请假或公出」的异常日对应的打卡记录。
    """
    allowed, filter_lsys, is_dakaman, include_buban = _can_see_attendance_exceptions(current_user or "")
    if not allowed:
        raise HTTPException(
            status_code=403,
            detail="仅班组长/主任/副主任、部长/副部长或打卡管理员可查看考勤异常",
        )
    try:
        built = _build_attendance_exceptions_data(year, month, filter_lsys, include_buban=include_buban)
        if not built:
            msg = "本室无考勤异常" if filter_lsys else "无考勤异常"
            return AttendanceQueryResponse(success=True, message=msg, total=0, data=[], is_dakaman=is_dakaman)
        attendance_records = [AttendanceRecord(**rec) for rec in built]
        return AttendanceQueryResponse(
            success=True,
            message="查询成功",
            total=len(attendance_records),
            data=attendance_records,
            is_dakaman=is_dakaman,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"考勤异常查询失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"查询失败: {str(e)}")


@router.get("/exceptions/export")
async def export_attendance_exceptions(
    year: int = Query(..., description="年份"),
    month: int = Query(..., ge=1, le=12, description="月份 1-12"),
    current_user: str = Query(..., description="当前登录用户姓名，用于权限校验"),
):
    """
    导出指定月份的考勤异常列表为 Excel。
    权限同 /attendance/exceptions。
    """
    allowed, filter_lsys, is_dakaman, include_buban = _can_see_attendance_exceptions(current_user or "")
    if not allowed:
        raise HTTPException(
            status_code=403,
            detail="仅班组长/主任/副主任、部长/副部长或打卡管理员可导出考勤异常",
        )
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            raise HTTPException(status_code=500, detail="服务端未安装 openpyxl，无法生成 Excel")

        rows = _build_attendance_exceptions_data(year, month, filter_lsys, include_buban=include_buban)
        wb = Workbook()
        ws = wb.active
        ws.title = "考勤异常"

        headers = [
            "日期", "姓名", "所在单位",
            "考勤时间1", "考勤时间2", "考勤时间3", "考勤时间4",
            "考勤时间5", "考勤时间6", "考勤时间7", "考勤时间8",
            "是否全天缺勤",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for r in rows:
            date_str = r.get("attendance_date") or ""
            name = (r.get("employee_name") or "").strip()
            dept = (r.get("department") or "").strip()
            t1 = r.get("time_1") or ""
            t2 = r.get("time_2") or ""
            t3 = r.get("time_3") or ""
            t4 = r.get("time_4") or ""
            t5 = r.get("time_5") or ""
            t6 = r.get("time_6") or ""
            t7 = r.get("time_7") or ""
            t8 = r.get("time_8") or ""
            is_full = bool(r.get("full_day_absence")) or all(
                not (v or "").strip() for v in [t1, t2, t3, t4, t5, t6, t7, t8]
            )
            ws.append([
                date_str,
                name,
                dept,
                t1, t2, t3, t4, t5, t6, t7, t8,
                "是" if is_full else "",
            ])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename_ascii = f"attendance_exceptions_{year}{month:02d}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_ascii}"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出考勤异常失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


def _fmt_dt_text(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, date_type):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _fmt_time_text(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%H:%M:%S")
    return str(val).strip()


def _status_text(value, approved_value=4, pending_values=(0, 1, 3, 5), rejected_value=22) -> str:
    try:
        v = int(value) if value is not None else None
    except Exception:
        v = None
    if v == approved_value:
        return "已通过"
    if v == rejected_value:
        return "已驳回"
    if v in pending_values:
        return "审核中"
    return f"状态{value}" if value is not None else "未知"


def _kqyc_status_text(first_status, second_status, processed_to_trip=None) -> str:
    try:
        fs = int(first_status or 0)
        ss = int(second_status or 0)
        pt = int(processed_to_trip or 0)
    except Exception:
        fs, ss, pt = 0, 0, 0
    if fs == 2 or ss == 2:
        return "已驳回"
    if fs == 1 and ss == 1:
        return "已通过" + ("，已转市内公出" if pt == 1 else "")
    if fs == 1:
        return "一级已通过，二级待审批"
    return "一级待审批"


def _biztrip_type_text(raw) -> str:
    s = str(raw or "").strip()
    if not s:
        return "境内公出"
    return s


def _overlap_where(start_col: str, end_col: str) -> str:
    return f"{start_col} < %s AND {end_col} >= %s"


def _query_suggestion_process_rows(names: List[str], start_dt: str, end_dt: str):
    if not names:
        return {}, {}, {}, {}, {}, {}, {}
    ph = ",".join(["%s"] * len(names))

    def by_name(rows, field="xm"):
        out = {}
        for r in rows or []:
            n = (r.get(field) or "").strip()
            if n:
                out.setdefault(n, []).append(r)
        return out

    jiaban_approved = by_name(db.execute_query(
        f"SELECT xm, timefrom, timeto, spr, spr2, jiabanzt, hx FROM jiaban "
        f"WHERE xm IN ({ph}) AND jiabanzt = 4 AND {_overlap_where('timefrom', 'timeto')}",
        tuple(names) + (end_dt, start_dt),
    ) or [])
    jiaban_pending = by_name(db.execute_query(
        f"SELECT xm, timefrom, timeto, spr, spr2, jiabanzt, hx FROM jiaban "
        f"WHERE xm IN ({ph}) AND jiabanzt IN (0,1,3,5) AND {_overlap_where('timefrom', 'timeto')}",
        tuple(names) + (end_dt, start_dt),
    ) or [])
    qj_approved = by_name(db.execute_query(
        f"SELECT xm, timefrom, timeto, qjfs, spr, spr2, qjzt, sptime, sp2time, sctime FROM qj "
        f"WHERE xm IN ({ph}) AND qjzt = 4 AND {_overlap_where('timefrom', 'timeto')}",
        tuple(names) + (end_dt, start_dt),
    ) or [])
    qj_pending = by_name(db.execute_query(
        f"SELECT xm, timefrom, timeto, qjfs, spr, spr2, qjzt, sptime, sp2time, sctime FROM qj "
        f"WHERE xm IN ({ph}) AND qjzt IN (0,1,3) AND {_overlap_where('timefrom', 'timeto')}",
        tuple(names) + (end_dt, start_dt),
    ) or [])
    gcsqb_approved = by_name(db.execute_query(
        f"SELECT gcr AS xm, yjcfsj, yjfhsj, gcsj, sjfhtime, gclx, gcdd, szr, bld, szrzt, bldzt, szrpztime, bldpztime "
        f"FROM gcsqb WHERE gcr IN ({ph}) AND bldzt = 2 AND szrzt = 2 "
        f"AND COALESCE(yjcfsj, gcsj) < %s AND COALESCE(yjfhsj, sjfhtime, yjcfsj, gcsj) >= %s",
        tuple(names) + (end_dt, start_dt),
    ) or [])
    gcsqb_pending = by_name(db.execute_query(
        f"SELECT gcr AS xm, yjcfsj, yjfhsj, gcsj, sjfhtime, gclx, gcdd, szr, bld, szrzt, bldzt, szrpztime, bldpztime "
        f"FROM gcsqb WHERE gcr IN ({ph}) AND (bldzt != 2 OR szrzt != 2) AND bldzt != 22 AND szrzt != 22 "
        f"AND COALESCE(yjcfsj, gcsj) < %s AND COALESCE(yjfhsj, sjfhtime, yjcfsj, gcsj) >= %s",
        tuple(names) + (end_dt, start_dt),
    ) or [])
    kqyc_rows = by_name(db.execute_query(
        f"SELECT applicant AS xm, attendance_date, time_from, time_to, reason_type, description, "
        f"first_approver, second_approver, first_status, second_status, first_approve_time, second_approve_time, "
        f"processed_to_trip, processed_at, apply_time "
        f"FROM attendance_exception WHERE applicant IN ({ph}) "
        f"AND CONCAT(attendance_date, ' ', time_from) < %s AND CONCAT(attendance_date, ' ', time_to) >= %s",
        tuple(names) + (end_dt, start_dt),
    ) or [])
    for n, rows in kqyc_rows.items():
        for r in rows:
            if int(r.get("first_status") or 0) == 1 and int(r.get("second_status") or 0) == 1:
                continue
            if int(r.get("first_status") or 0) == 2 or int(r.get("second_status") or 0) == 2:
                continue
            qj_pending.setdefault(n, []).append({
                "xm": n,
                "timefrom": f"{r.get('attendance_date')} {r.get('time_from')}",
                "timeto": f"{r.get('attendance_date')} {r.get('time_to')}",
            })
    return jiaban_approved, jiaban_pending, qj_approved, qj_pending, gcsqb_approved, gcsqb_pending, kqyc_rows


def _row_intersects(row: dict, start_key: str, end_key: str, day_start: datetime, day_end: datetime) -> bool:
    s = _parse_datetime_for_excel(row.get(start_key))
    e = _parse_datetime_for_excel(row.get(end_key))
    if not s or not e:
        return False
    return s < day_end and e >= day_start


def _flow_for_day(name: str, date_str: str, qj_rows, gcsqb_rows, kqyc_rows) -> str:
    day_start = datetime.strptime(f"{date_str} 00:00:00", "%Y-%m-%d %H:%M:%S")
    day_end = day_start + timedelta(days=1)
    parts = []
    for r in qj_rows or []:
        if not _row_intersects(r, "timefrom", "timeto", day_start, day_end):
            continue
        qjzt = r.get("qjzt")
        st = _status_text(qjzt)
        approvers = []
        if r.get("spr"):
            approvers.append(f"一级{r.get('spr')}" + (f"({_fmt_dt_text(r.get('sptime'))})" if r.get("sptime") else ""))
        if r.get("spr2"):
            approvers.append(f"二级{r.get('spr2')}" + (f"({_fmt_dt_text(r.get('sp2time'))})" if r.get("sp2time") else ""))
        parts.append(
            f"请假[{r.get('qjfs') or ''}] {st}：{_fmt_dt_text(r.get('timefrom'))}~{_fmt_dt_text(r.get('timeto'))}"
            + (f"；审批：{'，'.join(approvers)}" if approvers else "")
        )
    for r in gcsqb_rows or []:
        start = r.get("yjcfsj") or r.get("gcsj")
        end = r.get("yjfhsj") or r.get("sjfhtime") or start
        if not _row_intersects({"s": start, "e": end}, "s", "e", day_start, day_end):
            continue
        szrzt = _status_text(r.get("szrzt"), approved_value=2, pending_values=(0, 1), rejected_value=22)
        bldzt = _status_text(r.get("bldzt"), approved_value=2, pending_values=(0, 1), rejected_value=22)
        parts.append(
            f"公出[{_biztrip_type_text(r.get('gclx'))}] {start}~{end}；地点：{r.get('gcdd') or ''}；"
            f"室主任{r.get('szr') or ''}({szrzt}{',' + _fmt_dt_text(r.get('szrpztime')) if r.get('szrpztime') else ''})，"
            f"部领导{r.get('bld') or ''}({bldzt}{',' + _fmt_dt_text(r.get('bldpztime')) if r.get('bldpztime') else ''})"
        )
    for r in kqyc_rows or []:
        kdate = str(r.get("attendance_date") or "")[:10]
        if kdate != date_str:
            continue
        st = _kqyc_status_text(r.get("first_status"), r.get("second_status"), r.get("processed_to_trip"))
        parts.append(
            f"打卡异常申请[{r.get('reason_type') or ''}] {st}：{kdate} {r.get('time_from')}~{r.get('time_to')}；"
            f"一级{r.get('first_approver') or ''}({_fmt_dt_text(r.get('first_approve_time')) or '待'})，"
            f"二级{r.get('second_approver') or ''}({_fmt_dt_text(r.get('second_approve_time')) or '待'})"
            + (f"；转公出时间：{_fmt_dt_text(r.get('processed_at'))}" if r.get("processed_at") else "")
        )
    return "\n".join([p for p in parts if p]) or "未查询到处理流程"


def _name_initial_sort_key(name: str) -> str:
    text = (name or "").strip()
    if not text:
        return ""
    try:
        from pypinyin import lazy_pinyin, Style
        return "".join(lazy_pinyin(text, style=Style.FIRST_LETTER)).upper()
    except Exception:
        return text


def _payload_first_time(payload: dict) -> str:
    suggestions = payload.get("suggestions") or []
    first_suggestion = min((_fmt_dt_text(s.get("start_time")) for s in suggestions if s.get("start_time")), default="")
    rec = payload.get("record") or {}
    first_punch = ""
    date_str = str(rec.get("attendance_date") or "")[:10]
    for i in range(1, 11):
        t = _fmt_time_text(rec.get(f"time_{i}"))
        if t:
            first_punch = f"{date_str} {t}" if date_str else t
            break
    return first_suggestion or first_punch


@router.get("/suggestion-attendance-report/export")
async def export_suggestion_attendance_report(
    start_date: str = Query(..., description="开始日期 YYYY-MM-DD"),
    end_date: str = Query(..., description="结束日期 YYYY-MM-DD"),
    current_user: str = Query(..., description="当前登录用户姓名，用于权限校验"),
):
    """
    导出日期段打卡与智能建议处理报表。
    每人每天一行：打卡记录 + 缺勤标记 + 是否完成处理 + 处理流程。
    权限同考勤异常管理。
    """
    allowed, filter_lsys, _, include_buban = _can_see_attendance_exceptions(current_user or "")
    can_export_all = _can_export_suggestion_attendance_report_all(current_user or "")
    if not allowed and not can_export_all:
        raise HTTPException(status_code=403, detail="仅班组长/主任/副主任、部长/副部长或打卡管理员可导出")
    if can_export_all:
        filter_lsys = None
        include_buban = True
    try:
        sd = datetime.strptime(start_date[:10], "%Y-%m-%d").date()
        ed = datetime.strptime(end_date[:10], "%Y-%m-%d").date()
    except Exception:
        raise HTTPException(status_code=400, detail="日期格式应为 YYYY-MM-DD")
    if sd > ed:
        raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")
    if (ed - sd).days > 370:
        raise HTTPException(status_code=400, detail="单次导出日期范围不能超过 370 天")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl，无法生成 Excel")

    try:
        start_dt = f"{sd} 00:00:00"
        end_dt = f"{ed + timedelta(days=1)} 00:00:00"
        params = [str(sd), str(ed)]
        lsys_cond = ""
        buban_cond = "" if include_buban else " AND TRIM(COALESCE(department,'')) != '部办'"
        if filter_lsys:
            lsys_cond = " AND TRIM(COALESCE(department,'')) = %s"
            params.append(filter_lsys)

        records = db.execute_query(
            f"""
            SELECT *
            FROM attendance_records
            WHERE attendance_date >= %s AND attendance_date <= %s
              AND TRIM(COALESCE(department,'')) NOT IN ('其他部门员工','其他部门成员')
              {buban_cond}
              {lsys_cond}
            ORDER BY attendance_date, department, employee_name
            """,
            tuple(params),
        ) or []

        sugg_params = [start_dt, end_dt]
        sugg_lsys_cond = ""
        sugg_buban_cond = "" if include_buban else " AND TRIM(COALESCE(s.department,'')) != '部办'"
        if filter_lsys:
            sugg_lsys_cond = " AND TRIM(COALESCE(s.department,'')) = %s"
            sugg_params.append(filter_lsys)
        suggestion_rows = db.execute_query(
            f"""
            SELECT s.employee_name, s.department, DATE(s.start_time) AS date, s.day_type, s.message,
                   s.start_time, s.end_time, s.status, y.xbie
            FROM attendance_suggestions s
            LEFT JOIN yggl y ON y.name COLLATE utf8mb4_unicode_ci = s.employee_name COLLATE utf8mb4_unicode_ci
            WHERE s.start_time >= %s AND s.start_time < %s
              AND TRIM(COALESCE(s.department,'')) NOT IN ('其他部门员工','其他部门成员')
              {sugg_buban_cond}
              {sugg_lsys_cond}
            ORDER BY s.start_time, s.department, s.employee_name
            """,
            tuple(sugg_params),
        ) or []

        keys = {}
        for r in records:
            date_str = str(r.get("attendance_date") or "")[:10]
            name = (r.get("employee_name") or "").strip()
            dept = (r.get("department") or "").strip()
            if date_str and name:
                keys[(name, dept, date_str)] = {"record": r, "suggestions": []}
        for s in suggestion_rows:
            st = s.get("status") if s.get("status") is not None else 0
            if int(st or 0) != 1:
                continue
            date_str = str(s.get("date") or "")[:10]
            name = (s.get("employee_name") or "").strip()
            dept = (s.get("department") or "").strip()
            if not date_str or not name:
                continue
            if ("女" in (s.get("xbie") or "") or _is_female_employee(name)) and _is_march8_pm_interval(date_str, s.get("start_time"), s.get("end_time")):
                continue
            keys.setdefault((name, dept, date_str), {"record": None, "suggestions": []})["suggestions"].append(s)

        names = sorted({k[0] for k in keys})
        jiaban_ok, jiaban_pending, qj_ok, qj_pending, gcsqb_ok, gcsqb_pending, kqyc_rows = _query_suggestion_process_rows(names, start_dt, end_dt)

        wb = Workbook()
        ws = wb.active
        ws.title = "打卡与智能建议"
        headers = ["日期", "姓名", "科室"] + [f"打卡{i}" for i in range(1, 11)] + [
            "缺勤标记", "是否完成处理", "处理流程", "智能建议"
        ]
        ws.append(headers)

        fill_header = PatternFill("solid", fgColor="D9EAF7")
        fill_absent = PatternFill("solid", fgColor="F8CBAD")
        fill_normal = PatternFill("solid", fgColor="C6E0B4")
        fill_pending = PatternFill("solid", fgColor="FFE699")
        fill_missing = PatternFill("solid", fgColor="F4B084")
        thin = Side(style="thin", color="7F7F7F")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = fill_header
            c.border = border

        sorted_items = sorted(
            keys.items(),
            key=lambda x: (_name_initial_sort_key(x[0][0]), x[0][0], x[0][2], _payload_first_time(x[1])),
        )
        for (name, dept, date_str), payload in sorted_items:
            rec = payload.get("record") or {}
            suggestions_for_day = payload.get("suggestions") or []
            qj_all = (qj_ok.get(name, []) or []) + (qj_pending.get(name, []) or [])
            gcsqb_all = (gcsqb_ok.get(name, []) or []) + (gcsqb_pending.get(name, []) or [])
            flow = _flow_for_day(name, date_str, qj_all, gcsqb_all, kqyc_rows.get(name, []))
            punch_values = []
            for i in range(1, 11):
                t = _fmt_time_text(rec.get(f"time_{i}"))
                mark = rec.get(f"time_{i}_mark")
                if t and mark is not None:
                    mark_text = "进" if str(mark) == "0" else "出" if str(mark) == "1" else ""
                    if mark_text:
                        t = f"{t}({mark_text})"
                punch_values.append(t)

            absent = bool(suggestions_for_day)
            suggestion_text = "\n".join([
                f"{_fmt_dt_text(s.get('start_time'))}~{_fmt_dt_text(s.get('end_time'))} {s.get('message') or ''}"
                for s in suggestions_for_day
            ])
            if not absent:
                done_text = "已完成"
            else:
                suggestion_states = []
                for s in suggestions_for_day:
                    st = s.get("status") if s.get("status") is not None else 0
                    handled = _suggestion_handled(
                        s.get("start_time"), s.get("end_time"), int(st or 0),
                        jiaban_ok.get(name, []), qj_ok.get(name, []), gcsqb_ok.get(name, []),
                    )
                    under_review = (not handled) and _suggestion_under_review(
                        s.get("start_time"), s.get("end_time"), int(st or 0),
                        jiaban_pending.get(name, []), qj_pending.get(name, []), gcsqb_pending.get(name, []),
                    )
                    if handled:
                        suggestion_states.append("已完成")
                    elif under_review:
                        suggestion_states.append("审核中")
                    else:
                        suggestion_states.append("未处理")
                if not suggestion_states or "未处理" in suggestion_states:
                    done_text = "未处理"
                elif "审核中" in suggestion_states:
                    done_text = "审核中"
                else:
                    done_text = "已完成"

            ws.append([
                date_str,
                name,
                dept,
                *punch_values,
                "是" if absent else "否",
                done_text,
                flow if absent or flow != "未查询到处理流程" else "",
                suggestion_text,
            ])
            row_idx = ws.max_row
            absent_cell = ws.cell(row=row_idx, column=14)
            done_cell = ws.cell(row=row_idx, column=15)
            absent_cell.fill = fill_absent if absent else fill_normal
            if done_text == "已完成":
                done_cell.fill = fill_normal
            elif done_text == "审核中":
                done_cell.fill = fill_pending
            else:
                done_cell.fill = fill_missing

        for row in ws.iter_rows():
            for c in row:
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = [13, 12, 20] + [14] * 10 + [11, 13, 48, 54]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename_ascii = f"attendance_suggestion_report_{sd}_{ed}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_ascii}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出打卡与智能建议报表失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


def _dict_get_ci(row: Optional[dict], *keys: str):
    """从查询结果行按列名读取（大小写不敏感，兼容不同驱动返回的键名）。"""
    if not row:
        return None
    lower_map = {str(k).lower(): v for k, v in row.items()}
    for k in keys:
        if k.lower() in lower_map:
            return lower_map[k.lower()]
    return None


def _yggl_real_name_sql(alias: str, fallback_expr: str) -> str:
    """导出姓名：优先 yggl.真实姓名，否则回退到 name 或业务表姓名。"""
    return (
        f"COALESCE(NULLIF(TRIM({alias}.`真实姓名`), ''), "
        f"NULLIF(TRIM({alias}.name), ''), {fallback_expr})"
    )


def _leave_handler_category_from_gclx(raw) -> str:
    """
    异常处理表「请假类别」中公出一行的展示文案：与 gcsqb.gclx 一致。
    标准值：市内公出 / 境内公出 / 境外公出；gclx 为空时默认「境内公出」。
    """
    if raw is None:
        return "境内公出"
    s = str(raw).strip()
    if not s:
        return "境内公出"
    if s in ("市内公出", "境内公出", "境外公出"):
        return s
    # 兼容简写或历史脏数据
    if "境外" in s:
        return "境外公出"
    if "市内" in s:
        return "市内公出"
    if "境内" in s:
        return "境内公出"
    return s


def _parse_datetime_for_excel(val) -> Optional[datetime]:
    """将 DB 返回的 timefrom/timeto 转为 datetime，用于 Excel 的 DATE TIME 列。"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    if "." in s:
        s = s.split(".")[0]
    s = s[:19]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


@router.get("/leave-handler-export")
async def export_leave_handler_table(
    year: int = Query(..., description="年份"),
    month: int = Query(..., ge=1, le=12, description="月份 1-12"),
    current_user: str = Query(..., description="当前登录用户姓名，用于权限校验"),
):
    """
    导出异常处理表：按月全员请假信息 + 公出信息，XLS（Excel）格式。
    列顺序：A 员工代码(string) B 姓名(string，优先 yggl.真实姓名) C 部门(string，固定「智能制造工艺部」)
    D 请假/公出开始时间(DATE TIME) E 实际请假/公出结束时间(DATE TIME) F 请假类别(string)。
    请假走 qj.qjfs；公出走 gcsqb.gclx（市内公出/境内公出/境外公出），gclx 为空时默认「境内公出」。
    数据来源：qj 表（已通过 qjzt=4 + 审核中 qjzt IN(0,1,3)）+ gcsqb 表（已通过+审核中，排除驳回 bldzt/szrzt=22）。
    """
    import calendar
    allowed, _, _, _ = _can_see_attendance_exceptions(current_user or "")
    if not allowed:
        raise HTTPException(
            status_code=403,
            detail="仅班组长/主任/副主任、部长/副部长或打卡管理员可导出异常处理表",
        )
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            raise HTTPException(status_code=500, detail="服务端未安装 openpyxl，无法生成 Excel")

        month_str = f"{year}-{month:02d}"
        _, last_day = calendar.monthrange(year, month)
        start_date = f"{year}-{month:02d}-01"
        end_date = f"{year}-{month:02d}-{last_day:02d}"

        # 全员请假：已通过(qjzt=4) + 审核中(qjzt IN 0,1,3)，排除驳回；按月份筛选
        xm_expr = _yggl_real_name_sql("yggl", "qj.xm")
        sql = f"""
            SELECT {xm_expr} AS xm, qj.timefrom AS timefrom, qj.timeto AS timeto, qj.qjfs AS qjfs,
                   TRIM(yggl.gh) AS gh
            FROM qj
            LEFT JOIN yggl ON qj.xm = yggl.name
            WHERE qj.qjzt IN (0, 1, 3, 4)
              AND (qj.timefrom LIKE %s OR SUBSTRING(qj.timefrom, 1, 7) = %s)
              AND qj.xm NOT IN (SELECT name FROM yggl WHERE TRIM(lsys) IN ('其他部门员工','其他部门成员'))
            ORDER BY qj.timefrom ASC
        """
        rows = db.execute_query(sql, (f"{month_str}%", month_str)) or []

        # 公出：已通过 + 审核中，排除驳回(bldzt=22 或 szrzt=22)
        try:
            gcsqb_sql = f"""
                SELECT {_yggl_real_name_sql('y', 'g.gcr')} AS xm, g.yjcfsj AS timefrom, g.yjfhsj AS timeto,
                       NULLIF(TRIM(COALESCE(g.gclx, '')), '') AS gclx, TRIM(y.gh) AS gh
                FROM gcsqb g
                LEFT JOIN yggl y ON g.gcr = y.name
                WHERE g.bldzt != 22 AND g.szrzt != 22
                  AND (g.yjcfsj IS NOT NULL OR g.yjfhsj IS NOT NULL)
                  AND (
                    (g.yjcfsj IS NOT NULL AND DATE(g.yjcfsj) >= %s AND DATE(g.yjcfsj) <= %s)
                    OR (g.yjfhsj IS NOT NULL AND DATE(g.yjfhsj) >= %s AND DATE(g.yjfhsj) <= %s)
                  )
                  AND g.gcr NOT IN (SELECT name FROM yggl WHERE TRIM(lsys) IN ('其他部门员工','其他部门成员'))
                ORDER BY g.yjcfsj ASC
            """
            gcsqb_rows = db.execute_query(gcsqb_sql, (start_date, end_date, start_date, end_date)) or []
            for r in gcsqb_rows:
                category = _leave_handler_category_from_gclx(_dict_get_ci(r, "gclx"))
                gh_v = _dict_get_ci(r, "gh")
                xm_v = _dict_get_ci(r, "xm")
                rows.append({
                    "gh": "" if gh_v is None else str(gh_v).strip(),
                    "xm": "" if xm_v is None else str(xm_v).strip(),
                    "timefrom": _dict_get_ci(r, "timefrom"),
                    "timeto": _dict_get_ci(r, "timeto"),
                    "qjfs": category,
                })
        except Exception as e:
            logger.warning("导出异常处理表时查询公出失败（可能无 gcsqb/yjcfsj 列）: %s", e)

        # 按开始时间、姓名排序，便于与请假一起统一输出
        rows.sort(key=lambda x: (_parse_datetime_for_excel(x.get("timefrom")) or datetime.min, (x.get("xm") or "")))

        # 3 月导出时，若 3 月 8 日为工作日（非周六日且非系统节假日），为全体女性增加半天请假（13:00-17:00，请假类别「三八节」）
        if month == 3:
            march8 = datetime(year, 3, 8)
            weekday = march8.weekday()  # 0=周一, 5=周六, 6=周日
            is_weekend = weekday >= 5
            is_holiday = False
            try:
                from utils.holiday_loader import load_holidays_dict
                holidays = load_holidays_dict(str(year))
                march8_str = march8.strftime("%Y-%m-%d")
                if march8_str in holidays:
                    t = (holidays[march8_str] or "").strip()
                    if "假" in t or "休" in t:
                        is_holiday = True
            except Exception:
                pass
            if not is_weekend and not is_holiday:
                female_rows = db.execute_query(
                    "SELECT TRIM(gh) AS gh, "
                    "COALESCE(NULLIF(TRIM(`真实姓名`), ''), name) AS xm FROM yggl "
                    "WHERE (xbie LIKE %s OR xbie = %s) AND name IS NOT NULL AND TRIM(name) != '' "
                    "AND RIGHT(TRIM(name), 1) != '1' AND (lsys IS NULL OR RIGHT(TRIM(lsys), 1) != '1') "
                    "AND (TRIM(lsys) != %s OR lsys IS NULL) "
                    "AND TRIM(COALESCE(lsys,'')) NOT IN ('其他部门员工','其他部门成员') "
                    "AND (COALESCE(zaizhi, 0) = 0)",
                    ("%女%", "女", "部办"),
                ) or []
                march8_start = datetime(year, 3, 8, 13, 0, 0)
                march8_end = datetime(year, 3, 8, 17, 0, 0)
                for r in female_rows:
                    rows.append({
                        "gh": (r.get("gh") or "").strip(),
                        "xm": (r.get("xm") or "").strip(),
                        "timefrom": march8_start,
                        "timeto": march8_end,
                        "qjfs": "三八节",
                    })
                rows.sort(key=lambda x: (_parse_datetime_for_excel(x.get("timefrom")) or datetime.min, (x.get("xm") or "")))

        wb = Workbook()
        ws = wb.active
        ws.title = "异常处理表"

        headers = ["员工代码", "姓名", "部门", "请假开始时间", "实际请假结束时间", "请假类别"]
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            c = ws.cell(row=1, column=col)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        dept_fixed = "智能制造工艺部"
        for r in rows:
            gh = (r.get("gh") or "").strip()
            xm = (r.get("xm") or "").strip()
            timefrom_val = r.get("timefrom")
            timeto_val = r.get("timeto")
            qjfs = (r.get("qjfs") or "").strip()

            dt_start = _parse_datetime_for_excel(timefrom_val)
            dt_end = _parse_datetime_for_excel(timeto_val)

            row_data = [
                gh,
                xm,
                dept_fixed,
                dt_start if dt_start is not None else (timefrom_val if timefrom_val is not None else ""),
                dt_end if dt_end is not None else (timeto_val if timeto_val is not None else ""),
                qjfs,
            ]
            ws.append(row_data)

        # 设置 D、E 列为日期时间格式（对已写入的 datetime 对象生效）
        for row_idx in range(2, ws.max_row + 1):
            for col_letter in ("D", "E"):
                cell = ws[f"{col_letter}{row_idx}"]
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm:ss"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename_ascii = f"leave_handler_{year}{month:02d}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_ascii}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出异常处理表失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/query", response_model=AttendanceQueryResponse)
async def query_attendance(
    name: Optional[str] = Query(None, description="员工姓名"),
    dept: Optional[str] = Query(None, description="部门"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)")
):
    """
    查询考勤记录
    
    参数：
    - name: 员工姓名（可选）
    - dept: 部门（可选）
    - start_date: 开始日期（可选）
    - end_date: 结束日期（可选）
    
    如果不提供日期范围，则查询所有记录
    """
    
    try:
        records = []
        
        if start_date and end_date:
            # 按日期范围查询
            records = attendance_db.query_by_date_range(start_date, end_date, name, dept)
        elif name and dept:
            # 按姓名和部门查询
            records = attendance_db.query_by_name_and_dept(name, dept)
        else:
            return AttendanceQueryResponse(
                success=False,
                message="请提供查询条件：(name + dept) 或 (start_date + end_date)",
                total=0
            )
        
        # 转换为响应模型
        attendance_records = [AttendanceRecord(**record) for record in records]
        
        return AttendanceQueryResponse(
            success=True,
            message="查询成功",
            total=len(attendance_records),
            data=attendance_records
        )
    
    except Exception as e:
        logger.error(f"查询失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"查询失败: {str(e)}")


@router.get("/dates", response_model=dict)
async def get_attendance_dates(
    name: str = Query(..., description="员工姓名"),
    dept: str = Query(..., description="部门")
):
    """
    获取某个员工的所有打卡日期
    
    用于前端判断哪些日期有打卡记录
    """
    
    try:
        dates = attendance_db.get_all_attendance_dates(name, dept)
        
        return {
            "success": True,
            "name": name,
            "dept": dept,
            "dates": dates,
            "total": len(dates)
        }
    
    except Exception as e:
        logger.error(f"查询日期失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"查询失败: {str(e)}")


@router.delete("/clear")
async def clear_all_data(confirm: str = Query(..., description="确认码，输入'CONFIRM'以确认删除")):
    """
    清空所有考勤数据（危险操作）
    
    需要确认码：CONFIRM
    """
    
    if confirm != "CONFIRM":
        raise HTTPException(status_code=400, detail="确认码不正确")
    
    try:
        conn = attendance_db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("DELETE FROM attendance_records")
        deleted_count = cursor.rowcount
        
        conn.commit()
        conn.close()
        
        logger.warning(f"已清空所有考勤数据，共删除 {deleted_count} 条记录")
        
        return {
            "success": True,
            "message": f"已清空所有数据，共删除 {deleted_count} 条记录"
        }
    
    except Exception as e:
        logger.error(f"清空数据失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"操作失败: {str(e)}")


# ==================== 打卡管理员代处理考勤异常 ====================

class DakamanProcessRequest(BaseModel):
    current_user: str
    employee_name: str
    department: str
    attendance_date: str  # YYYY-MM-DD
    process_type: str  # leave | business_trip
    leave_type: str = "事假"  # qjfs: 换休/带薪年休假/事假/病假 etc.
    trip_scope: str = "境内公出"  # gclx: 市内公出/境内公出/境外公出
    reason: str = "打卡管理员代处理"


@router.post("/dakaman-process")
async def dakaman_process_exception(req: DakamanProcessRequest):
    """
    打卡管理员(dakaman)代替员工处理考勤异常。
    根据 attendance_suggestions 中该员工当日 status=1 的建议时间段，
    自动创建已审批通过的请假/公出记录。
    """
    current_user = (req.current_user or "").strip()
    dakaman = _get_dakaman()
    admin1 = _get_admin1()
    is_allowed = (dakaman and current_user == dakaman) or (admin1 and current_user == admin1)
    if not is_allowed:
        raise HTTPException(status_code=403, detail="仅打卡管理员或系统管理员可执行此操作")

    emp_name = (req.employee_name or "").strip()
    dept = (req.department or "").strip()
    att_date = (req.attendance_date or "").strip()
    if not emp_name or not dept or not att_date:
        raise HTTPException(status_code=400, detail="参数不完整")

    try:
        date_parts = att_date.split("-")
        year = int(date_parts[0])
        month = int(date_parts[1])
    except (ValueError, IndexError):
        raise HTTPException(status_code=400, detail="日期格式错误")

    suggestions = attendance_db.get_suggestions(emp_name, dept, year, month)
    day_suggestions = [
        s for s in suggestions
        if str(s.get("date", ""))[:10] == att_date and s.get("status") == 1
    ]
    if not day_suggestions:
        raise HTTPException(status_code=404, detail="该员工当日无需处理的缺勤异常")

    emp_info = _get_user_info(emp_name)
    lsys = (emp_info.get("lsys") or dept) if emp_info else dept
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    created_ids = []

    for s in day_suggestions:
        start_time = s.get("start_time")
        end_time = s.get("end_time")
        if not start_time or not end_time:
            continue
        st_str = str(start_time)[:19].replace("T", " ")
        et_str = str(end_time)[:19].replace("T", " ")
        if len(st_str) == 10:
            st_str += " 00:00:00"
        if len(et_str) == 10:
            et_str += " 23:59:59"

        new_id = uuid.uuid4().hex

        if req.process_type == "leave":
            from utils.helpers import normalize_datetime_for_db
            st_norm = normalize_datetime_for_db(st_str)
            et_norm = normalize_datetime_for_db(et_str)

            try:
                st_dt = datetime.strptime(st_norm[:19], "%Y-%m-%d %H:%M:%S")
                et_dt = datetime.strptime(et_norm[:19], "%Y-%m-%d %H:%M:%S")
                diff_hours = (et_dt - st_dt).total_seconds() / 3600
                raw_days = (diff_hours / 8) if diff_hours > 0 else 0.5
                dur_days = normalize_qj_tian_days(raw_days)
            except Exception:
                dur_days = normalize_qj_tian_days(0.5)
            xiaoshi = str(round(dur_days * 8, 2))

            sql = """
                INSERT INTO qj (id, bz, xm, qjfs, bc, gx, jy, smcl, smclwj, timefrom, timeto, timefromdate,
                    tian, xiaoshi, qjtime, qjzt, spr, `2j`, spr2, content, lsys, hxpxh, hxwc, hxps)
                VALUES (%s, %s, %s, %s, '白班', '', %s, '打卡管理员代处理', '', %s, %s, %s,
                    %s, %s, %s, 4, %s, 0, '', %s, %s, 0, 0, 0)
            """
            params = (
                new_id, dept, emp_name, req.leave_type,
                req.reason or "打卡管理员代处理",
                st_norm, et_norm, st_norm[:10],
                str(dur_days), xiaoshi, now,
                current_user, req.reason or "打卡管理员代处理", lsys,
            )
            db.execute_insert(sql, params)
            created_ids.append(new_id)

        elif req.process_type == "business_trip":
            gclx = (req.trip_scope or "").strip()
            if gclx not in ("市内公出", "境内公出", "境外公出"):
                gclx = "境内公出"
            sql = """
                INSERT INTO gcsqb (id, gclx, wpdw, gcr, gzh, gcdw, lxdh, wpsj,
                    yjfhsj, yjcfsj, xmmc, tzdbh, bcgczrs, gcdd, qkje, gcrw,
                    szr, bld, gcsj, sjfhtime, bldzt, szrzt, fhdj_status)
                VALUES (%s, %s, '', %s, '', %s, '', %s,
                    %s, %s, '', '', '1', '', 0, %s,
                    %s, %s, %s, %s, 2, 2, 1)
            """
            params = (
                new_id, gclx, emp_name, dept, now,
                et_str, st_str,
                req.reason or "打卡管理员代处理",
                current_user, current_user, st_str, et_str,
            )
            db.execute_insert(sql, params)
            created_ids.append(new_id)
        else:
            raise HTTPException(status_code=400, detail="process_type 必须为 leave 或 business_trip")

    return {
        "success": True,
        "message": f"已处理 {len(created_ids)} 条异常记录",
        "ids": created_ids,
    }
