# -*- coding: utf-8 -*-
"""
论文保密审批台账 API
- 全员可填写、查看、编辑和导出
- 导出格式按《智能制造工艺部论文保密审批台账》模板生成
"""
from datetime import datetime
from io import BytesIO
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Form, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from database import db


router = APIRouter(prefix="/confidentiality-ledger", tags=["论文保密审批台账"])

FIELDS = [
    "applicant",
    "paper_title",
    "apply_time",
    "material_form",
    "publish_channel",
    "is_confidential",
    "military_research",
]

HEADERS = ["序号", "申请人", "论文名称", "申请时间", "资料形式", "发布途径", "是否涉密", "是否涉及军工及科研生产"]
WIDTHS = [13, 14, 56, 14, 14, 18, 12, 26]

_table_ready = False


def _ensure_table():
    global _table_ready
    if _table_ready:
        return
    sql = """
    CREATE TABLE IF NOT EXISTS confidentiality_ledger (
        id INT AUTO_INCREMENT PRIMARY KEY,
        applicant VARCHAR(100) DEFAULT '' COMMENT '申请人',
        paper_title VARCHAR(1000) DEFAULT '' COMMENT '论文名称',
        apply_time VARCHAR(50) DEFAULT '' COMMENT '申请时间',
        material_form VARCHAR(100) DEFAULT '' COMMENT '资料形式',
        publish_channel VARCHAR(200) DEFAULT '' COMMENT '发布途径',
        is_confidential VARCHAR(50) DEFAULT '' COMMENT '是否涉密',
        military_research VARCHAR(50) DEFAULT '' COMMENT '是否涉及军工及科研生产',
        created_by VARCHAR(100) DEFAULT '' COMMENT '创建人',
        updated_by VARCHAR(100) DEFAULT '' COMMENT '最后修改人',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
        INDEX idx_apply_time (apply_time),
        INDEX idx_applicant (applicant),
        INDEX idx_updated_at (updated_at)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='论文保密审批台账';
    """
    db.execute_update(sql)
    _table_ready = True


def _fmt_dt(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def _normalize_date_for_excel(value: str) -> str:
    s = (value or "").strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            d = datetime.strptime(s, fmt)
            return f"{d.year}.{d.month}.{d.day}"
        except ValueError:
            pass
    return s


def _clean_row(row: dict) -> dict:
    result = {k: row.get(k) or "" for k in ["id", *FIELDS, "created_by", "updated_by"]}
    result["created_at"] = _fmt_dt(row.get("created_at"))
    result["updated_at"] = _fmt_dt(row.get("updated_at"))
    return result


def _build_where(keyword: str = ""):
    where = ["1=1"]
    params = []
    if keyword and keyword.strip():
        kw = f"%{keyword.strip()}%"
        where.append("(applicant LIKE %s OR paper_title LIKE %s OR publish_channel LIKE %s)")
        params.extend([kw, kw, kw])
    return " AND ".join(where), params


@router.get("/records")
async def list_records(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    keyword: str = Query(""),
):
    """分页获取台账记录。"""
    _ensure_table()
    where_clause, params = _build_where(keyword)
    total = db.execute_scalar(f"SELECT COUNT(*) AS cnt FROM confidentiality_ledger WHERE {where_clause}", tuple(params)) or 0
    offset = (page - 1) * page_size
    rows = db.execute_query(
        f"""
        SELECT id, applicant, paper_title, apply_time, material_form, publish_channel,
               is_confidential, military_research, created_by, updated_by, created_at, updated_at
        FROM confidentiality_ledger
        WHERE {where_clause}
        ORDER BY id ASC
        LIMIT %s OFFSET %s
        """,
        tuple(params) + (page_size, offset),
    )
    return {"success": True, "data": [_clean_row(r) for r in rows], "total": total, "page": page, "page_size": page_size}


@router.post("/records")
async def create_record(
    applicant: str = Form(""),
    paper_title: str = Form(""),
    apply_time: str = Form(""),
    material_form: str = Form(""),
    publish_channel: str = Form(""),
    is_confidential: str = Form(""),
    military_research: str = Form(""),
    current_user: str = Form(""),
):
    """新增一条台账记录。"""
    _ensure_table()
    if not any((applicant, paper_title, apply_time, material_form, publish_channel, is_confidential, military_research)):
        raise HTTPException(status_code=400, detail="请至少填写一项内容")
    new_id = db.execute_insert(
        """
        INSERT INTO confidentiality_ledger
            (applicant, paper_title, apply_time, material_form, publish_channel, is_confidential, military_research, created_by, updated_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            applicant.strip(),
            paper_title.strip(),
            apply_time.strip(),
            material_form.strip(),
            publish_channel.strip(),
            is_confidential.strip(),
            military_research.strip(),
            current_user.strip(),
            current_user.strip(),
        ),
    )
    if new_id is None:
        raise HTTPException(status_code=500, detail="新增失败")
    return {"success": True, "id": new_id}


@router.put("/records/{record_id}")
async def update_record(
    record_id: int,
    applicant: str = Form(""),
    paper_title: str = Form(""),
    apply_time: str = Form(""),
    material_form: str = Form(""),
    publish_channel: str = Form(""),
    is_confidential: str = Form(""),
    military_research: str = Form(""),
    current_user: str = Form(""),
):
    """更新台账记录。"""
    _ensure_table()
    affected = db.execute_update(
        """
        UPDATE confidentiality_ledger
        SET applicant = %s, paper_title = %s, apply_time = %s, material_form = %s,
            publish_channel = %s, is_confidential = %s, military_research = %s, updated_by = %s
        WHERE id = %s
        """,
        (
            applicant.strip(),
            paper_title.strip(),
            apply_time.strip(),
            material_form.strip(),
            publish_channel.strip(),
            is_confidential.strip(),
            military_research.strip(),
            current_user.strip(),
            record_id,
        ),
    )
    if affected <= 0:
        raise HTTPException(status_code=404, detail="记录不存在或未发生变化")
    return {"success": True}


@router.delete("/records/{record_id}")
async def delete_record(record_id: int):
    """删除台账记录。"""
    _ensure_table()
    affected = db.execute_update("DELETE FROM confidentiality_ledger WHERE id = %s", (record_id,))
    if affected <= 0:
        raise HTTPException(status_code=404, detail="记录不存在")
    return {"success": True}


def _build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "保密审批台账"

    ws.merge_cells("A1:H1")
    ws["A1"] = "智能制造工艺部论文保密审批台账"
    ws["A1"].font = Font(name="宋体", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(2, idx, header)
        cell.font = Font(name="宋体", size=11, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(rows, 3):
        values = [
            row_idx - 2,
            row.get("applicant") or "",
            row.get("paper_title") or "",
            _normalize_date_for_excel(row.get("apply_time") or ""),
            row.get("material_form") or "",
            row.get("publish_channel") or "",
            row.get("is_confidential") or "",
            row.get("military_research") or "",
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.font = Font(name="宋体", size=11)
            cell.alignment = Alignment(horizontal="center" if col_idx != 3 else "left", vertical="center", wrap_text=True)
            cell.border = border

    for idx, width in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[2].height = 24
    for row_idx in range(3, max(3, len(rows) + 3)):
        ws.row_dimensions[row_idx].height = 40

    return wb


@router.get("/export")
async def export_records(keyword: Optional[str] = Query("")):
    """导出论文保密审批台账 Excel。"""
    _ensure_table()
    where_clause, params = _build_where(keyword or "")
    rows = db.execute_query(
        f"""
        SELECT applicant, paper_title, apply_time, material_form, publish_channel, is_confidential, military_research
        FROM confidentiality_ledger
        WHERE {where_clause}
        ORDER BY id ASC
        """,
        tuple(params),
    )
    wb = _build_workbook(rows)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = "智能制造工艺部论文保密审批台账.xlsx"
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
