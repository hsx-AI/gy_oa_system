# -*- coding: utf-8 -*-
"""
低值易耗报销申请 API
- 员工提交申请，必须上传已购买实物照片和发票
- 二级审批人：主管领导（副经理）
- 三级审批人：经理
- 三级通过后流转到综合技术室主任/副主任完成报销闭环
- 支持导出与模板一致的低值易耗报销台账 Excel
"""
from datetime import datetime
from pathlib import Path
import io
import logging
import os
import re
import uuid
import zipfile

from fastapi import APIRouter, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse

from config import settings
from database import db

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/low-value-reimbursement", tags=["低值易耗报销"])

_BASE = Path(__file__).resolve().parent.parent
UPLOAD_DIR = _BASE / settings.UPLOAD_DIR / "low_value_reimbursement"
EXPORT_DIR = _BASE / "temp_docs"

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}
INVOICE_EXTENSIONS = IMAGE_EXTENSIONS | {".pdf", ".ofd"}

STATUS_PENDING_SECOND = 0
STATUS_PENDING_THIRD = 1
STATUS_PENDING_COMPLETE = 2
STATUS_COMPLETED = 3
STATUS_REJECTED = 22


def _ensure_dirs():
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def _init_table():
    sql = """
    CREATE TABLE IF NOT EXISTS low_value_reimbursement (
        id INT AUTO_INCREMENT PRIMARY KEY,
        material_name VARCHAR(200) NOT NULL COMMENT '物资名称',
        specification VARCHAR(200) DEFAULT '' COMMENT '规格',
        unit_price DECIMAL(12,2) NOT NULL DEFAULT 0.00 COMMENT '单价',
        quantity DECIMAL(12,2) NOT NULL DEFAULT 0.00 COMMENT '数量',
        total_price DECIMAL(12,2) NOT NULL DEFAULT 0.00 COMMENT '总价',
        supplier VARCHAR(200) DEFAULT '' COMMENT '供应商名称',
        work_no VARCHAR(100) DEFAULT '' COMMENT '工作号/科研号',
        part_no VARCHAR(100) DEFAULT '' COMMENT '部套号',
        usage_detail TEXT COMMENT '用途（详细说明）',
        photo_attachment VARCHAR(500) NOT NULL COMMENT '实物照片存储文件名',
        photo_original VARCHAR(500) NOT NULL COMMENT '实物照片原始文件名',
        invoice_attachment VARCHAR(500) NOT NULL COMMENT '发票存储文件名',
        invoice_original VARCHAR(500) NOT NULL COMMENT '发票原始文件名',
        applicant VARCHAR(50) NOT NULL COMMENT '申请人',
        department VARCHAR(100) DEFAULT '' COMMENT '申请人科室',
        approver2 VARCHAR(50) NOT NULL COMMENT '二级审批人',
        approver3 VARCHAR(50) NOT NULL COMMENT '三级审批人',
        completer VARCHAR(50) DEFAULT '' COMMENT '报销完成人',
        status TINYINT NOT NULL DEFAULT 0 COMMENT '0待二级 1待三级 2待报销完成 3已完成 22已驳回',
        reject_reason VARCHAR(500) DEFAULT '' COMMENT '驳回原因',
        apply_time DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '申请日期',
        approve2_time DATETIME NULL COMMENT '二级审批时间',
        approve3_time DATETIME NULL COMMENT '三级审批时间',
        complete_time DATETIME NULL COMMENT '报销完成时间',
        remark VARCHAR(500) DEFAULT '' COMMENT '备注',
        INDEX idx_applicant (applicant),
        INDEX idx_approver2 (approver2),
        INDEX idx_approver3 (approver3),
        INDEX idx_status (status),
        INDEX idx_apply_time (apply_time)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='低值易耗报销申请表';
    """
    db.execute_update(sql)
    budget_sql = """
    CREATE TABLE IF NOT EXISTS low_value_budget (
        id INT AUTO_INCREMENT PRIMARY KEY,
        budget_year INT NOT NULL COMMENT '年度',
        total_amount DECIMAL(14,2) NOT NULL DEFAULT 0.00 COMMENT '年度低值易耗总额度',
        remark VARCHAR(500) DEFAULT '' COMMENT '备注',
        updated_by VARCHAR(50) DEFAULT '' COMMENT '最后修改人',
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
        UNIQUE KEY uk_budget_year (budget_year)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='低值易耗年度额度配置';
    """
    db.execute_update(budget_sql)


_table_ready = False


def _ensure_table():
    global _table_ready
    if not _table_ready:
        _init_table()
        _table_ready = True


def _to_money(value) -> float:
    try:
        return round(float(value or 0), 2)
    except Exception:
        return 0.0


def _budget_summary(year: int) -> dict:
    """按年度汇总：总额度、已完成消耗、审核中金额、结余。"""
    budget_row = db.execute_query(
        "SELECT total_amount, remark, updated_by, updated_at FROM low_value_budget WHERE budget_year = %s LIMIT 1",
        (year,),
    )
    total_amount = _to_money(budget_row[0].get("total_amount")) if budget_row else 0.0
    remark = (budget_row[0].get("remark") or "") if budget_row else ""
    updated_by = (budget_row[0].get("updated_by") or "") if budget_row else ""
    updated_at = _fmt_dt(budget_row[0].get("updated_at")) if budget_row else ""

    completed_amount = _to_money(db.execute_scalar(
        """
        SELECT COALESCE(SUM(total_price), 0)
        FROM low_value_reimbursement
        WHERE status = %s AND complete_time IS NOT NULL AND YEAR(complete_time) = %s
        """,
        (STATUS_COMPLETED, year),
    ))
    pending_amount = _to_money(db.execute_scalar(
        """
        SELECT COALESCE(SUM(total_price), 0)
        FROM low_value_reimbursement
        WHERE status IN (%s, %s, %s) AND YEAR(apply_time) = %s
        """,
        (STATUS_PENDING_SECOND, STATUS_PENDING_THIRD, STATUS_PENDING_COMPLETE, year),
    ))
    remaining_amount = round(total_amount - completed_amount, 2)
    projected_remaining = round(total_amount - completed_amount - pending_amount, 2)
    return {
        "year": year,
        "total_amount": total_amount,
        "completed_amount": completed_amount,
        "pending_amount": pending_amount,
        "remaining_amount": remaining_amount,
        "projected_remaining": projected_remaining,
        "configured": bool(budget_row),
        "remark": remark,
        "updated_by": updated_by,
        "updated_at": updated_at,
    }


def _fmt_dt(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def _fmt_date(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def _safe_zip_name(value: str) -> str:
    value = (value or "").strip() or "invoice.pdf"
    return re.sub(r'[\\/:*?"<>|\r\n]+', "_", value)


def _to_float(value):
    try:
        return float(str(value).replace(",", "").strip())
    except Exception:
        return None


def _extract_invoice_number(text: str) -> str:
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    label_pattern = re.compile(r"(发票号码|发票号|数电票号码)")
    same_line_pattern = re.compile(r"(?:发票号码|发票号|数电票号码)[:：]?\s*([0-9]{8,30})")
    pure_number_pattern = re.compile(r"^[0-9]{8,30}$")

    for idx, line in enumerate(lines):
        match = same_line_pattern.search(line)
        if match:
            return match.group(1)
        if label_pattern.search(line):
            for candidate in lines[idx + 1:idx + 45]:
                compact = re.sub(r"\s+", "", candidate)
                if pure_number_pattern.fullmatch(compact):
                    return compact

    normalized = re.sub(r"[ \t\r\n]+", "", text or "")
    for pattern in (r"发票号码[:：]?([0-9]{8,30})", r"数电票号码[:：]?([0-9]{8,30})"):
        match = re.search(pattern, normalized)
        if match:
            return match.group(1)
    return ""


def _normalize_supplier(value: str) -> str:
    return re.sub(r"\s+", "", value or "")


def _extract_item_numbers(lines: list[str], item_start: int | None) -> tuple[float | None, float | None]:
    """Return (quantity, raw_unit_price) from the item-detail area only."""
    if item_start is None:
        return None, None

    item_numbers = []
    stop_prefixes = ("订单", "收款人", "复核人", "下载次数", "备注")
    for line in lines[item_start + 1:]:
        if line.startswith(stop_prefixes):
            break
        if re.fullmatch(r"-?\d+(?:\.\d+)?", line):
            item_numbers.append(float(line))

    # 数电发票抽出的商品列通常是：金额、税额、不含税单价、数量。
    # 如果数量为空，最后一个数字会变成“不含税单价”，不能误当数量。
    if len(item_numbers) >= 4:
        return item_numbers[-1], item_numbers[-2]
    if len(item_numbers) >= 3:
        return None, item_numbers[-1]
    return None, None


def _extract_invoice_fields(text: str, filename: str = "") -> dict:
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    company_lines = [
        line for line in lines
        if ("公司" in line or "商行" in line or "经营部" in line or "个体工商户" in line)
        and "电子发票" not in line
    ]
    supplier = company_lines[1] if len(company_lines) >= 2 else (company_lines[0] if company_lines else "")

    amount_values = []
    for m in re.finditer(r"[¥￥]\s*([0-9]+(?:\.[0-9]+)?)", text or ""):
        val = _to_float(m.group(1))
        if val is not None:
            amount_values.append(val)
    invoice_total = max(amount_values) if amount_values else None

    date_match = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text or "")
    invoice_date = ""
    if date_match:
        invoice_date = f"{date_match.group(1)}-{int(date_match.group(2)):02d}-{int(date_match.group(3)):02d}"

    item_name = ""
    item_start = None
    for idx, line in enumerate(lines):
        if line.startswith("*"):
            item_start = idx
            item_name = "".join(lines[idx:idx + 4])
            break

    # 数电发票文本常按列抽出：金额、税额、不含税单价、数量。
    # 仅在商品明细区域取数，避免数量为空时抓到票面其他数字。
    quantity = None
    raw_unit_price = None
    quantity, raw_unit_price = _extract_item_numbers(lines, item_start)

    file_match = re.search(
        r"(?P<amount>[0-9]+(?:\.[0-9]+)?)元-(?P<supplier>.+?)-(?P<date>\d{4}\.\d{1,2}\.\d{1,2})",
        filename or "",
    )
    if file_match:
        if invoice_total is None:
            invoice_total = _to_float(file_match.group("amount"))
        if not supplier:
            supplier = file_match.group("supplier").strip()
        if not invoice_date:
            invoice_date = file_match.group("date").replace(".", "-")

    quantity_defaulted = False
    if not quantity or quantity <= 0:
        quantity = 1.0
        quantity_defaulted = True
    unit_price = round(float(invoice_total) / float(quantity), 2) if invoice_total is not None and quantity else raw_unit_price

    return {
        "supplier": supplier,
        "invoice_number": _extract_invoice_number(text),
        "quantity": quantity,
        "quantity_defaulted": quantity_defaulted,
        "unit_price": unit_price,
        "invoice_total": invoice_total,
        "raw_unit_price": raw_unit_price,
        "invoice_date": invoice_date,
        "material_name": item_name,
        "text": text or "",
    }


def _extract_pdf_text(content: bytes) -> str:
    try:
        import fitz
    except Exception:
        raise HTTPException(status_code=500, detail="服务端未安装 PyMuPDF，无法解析 PDF 发票")
    try:
        doc = fitz.open(stream=content, filetype="pdf")
        return "\n".join(page.get_text("text") for page in doc)
    except Exception as e:
        logger.warning("发票 PDF 解析失败: %s", e)
        raise HTTPException(status_code=400, detail="发票 PDF 解析失败")


def _pdf_first_page_png(pdf_path: Path) -> Path | None:
    try:
        import fitz
    except Exception:
        return None
    try:
        doc = fitz.open(str(pdf_path))
        if len(doc) <= 0:
            return None
        page = doc[0]
        pix = page.get_pixmap(matrix=fitz.Matrix(1.2, 1.2), alpha=False)
        out = EXPORT_DIR / f"{uuid.uuid4().hex}.png"
        pix.save(str(out))
        return out
    except Exception as e:
        logger.warning("发票 PDF 首页渲染失败: %s", e)
        return None


def _image_source_for_excel(filename: str) -> Path | None:
    if not filename:
        return None
    path = UPLOAD_DIR / filename
    if not path.exists():
        return None
    ext = path.suffix.lower()
    if ext in IMAGE_EXTENSIONS:
        return path
    if ext == ".pdf":
        return _pdf_first_page_png(path)
    return None


def _add_image_to_sheet(ws, image_path: Path | None, cell_ref: str, max_width: int = 130, max_height: int = 82) -> bool:
    if not image_path or not image_path.exists():
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage
    except Exception:
        return False
    try:
        img = XLImage(str(image_path))
        scale = min(max_width / float(img.width or max_width), max_height / float(img.height or max_height), 1)
        img.width = int((img.width or max_width) * scale)
        img.height = int((img.height or max_height) * scale)
        ws.add_image(img, cell_ref)
        return True
    except Exception as e:
        logger.warning("Excel 嵌入图片失败: %s", e)
        return False


def _build_records_where(keyword: str, status: str, date_from: str = "", date_to: str = ""):
    parts = ["1=1"]
    params: list = []
    if keyword.strip():
        kw = f"%{keyword.strip()}%"
        parts.append(
            "(material_name LIKE %s OR supplier LIKE %s OR work_no LIKE %s OR part_no LIKE %s OR applicant LIKE %s OR usage_detail LIKE %s)"
        )
        params.extend([kw, kw, kw, kw, kw, kw])
    status_map = {
        "pending2": STATUS_PENDING_SECOND,
        "pending3": STATUS_PENDING_THIRD,
        "pending-complete": STATUS_PENDING_COMPLETE,
        "completed": STATUS_COMPLETED,
        "rejected": STATUS_REJECTED,
    }
    if status in status_map:
        parts.append("status = %s")
        params.append(status_map[status])
    elif status == "active":
        parts.append("status != %s")
        params.append(STATUS_REJECTED)
    if (date_from or "").strip():
        parts.append("DATE(apply_time) >= %s")
        params.append(date_from.strip())
    if (date_to or "").strip():
        parts.append("DATE(apply_time) <= %s")
        params.append(date_to.strip())
    return " AND ".join(parts), params


def _status_text(status) -> str:
    try:
        st = int(status)
    except Exception:
        st = -1
    return {
        STATUS_PENDING_SECOND: "待二级审批",
        STATUS_PENDING_THIRD: "待三级审批",
        STATUS_PENDING_COMPLETE: "待报销完成",
        STATUS_COMPLETED: "已完成",
        STATUS_REJECTED: "已驳回",
    }.get(st, "未知")


def _attach_display_fields(row: dict) -> None:
    row["status_text"] = _status_text(row.get("status"))
    for key in ("apply_time", "approve2_time", "approve3_time", "complete_time"):
        if row.get(key):
            row[key] = _fmt_dt(row.get(key))
    for key in ("unit_price", "quantity", "total_price"):
        if row.get(key) is not None:
            row[key] = float(row.get(key) or 0)


def _is_reimbursement_handler(name: str) -> bool:
    if not (name or "").strip():
        return False
    rows = db.execute_query(
        "SELECT name, jb, lsys FROM yggl WHERE name = %s AND COALESCE(zaizhi, 0) = 0 LIMIT 1",
        (name.strip(),),
    )
    if not rows:
        return False
    jb = (rows[0].get("jb") or "").strip()
    lsys = (rows[0].get("lsys") or "").strip()
    return lsys == "综合技术室" and ("主任" in jb)


def _can_view_ledger(name: str) -> bool:
    if not (name or "").strip():
        return False
    rows = db.execute_query(
        "SELECT name, jb, lsys FROM yggl WHERE name = %s AND COALESCE(zaizhi, 0) = 0 LIMIT 1",
        (name.strip(),),
    )
    if not rows:
        return False
    jb = (rows[0].get("jb") or "").strip()
    lsys = (rows[0].get("lsys") or "").strip()
    if lsys == "综合技术室" and "主任" in jb:
        return True
    return jb == "经理" or jb.startswith("经理") or jb == "副经理" or jb.startswith("副经理")


def _require_ledger_permission(current_user: str) -> None:
    if not _can_view_ledger(current_user):
        raise HTTPException(status_code=403, detail="仅综合技术室主任/副主任、经理、副经理可查看报销台账")


def _save_upload(file: UploadFile, allowed_exts: set[str], label: str) -> tuple[str, str]:
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail=f"请上传{label}")
    original = file.filename
    ext = os.path.splitext(original)[1].lower()
    if ext not in allowed_exts:
        allowed = "、".join(sorted(allowed_exts))
        raise HTTPException(status_code=400, detail=f"{label}格式不支持，请上传：{allowed}")
    stored = f"{uuid.uuid4().hex}{ext}"
    return stored, original


async def _write_upload(file: UploadFile, stored_name: str, label: str) -> None:
    try:
        content = await file.read()
        if not content:
            raise ValueError("empty file")
        with open(UPLOAD_DIR / stored_name, "wb") as f:
            f.write(content)
    except Exception as e:
        logger.error("%s保存失败: %s", label, e)
        raise HTTPException(status_code=500, detail=f"{label}保存失败，请重试")


def _list_people(where_sql: str):
    rows = db.execute_query(
        f"""
        SELECT name, jb, lsys
        FROM yggl
        WHERE COALESCE(zaizhi, 0) = 0 AND {where_sql}
        ORDER BY name
        """
    )
    return [
        {
            "name": r.get("name") or "",
            "jb": r.get("jb") or "",
            "lsys": r.get("lsys") or "",
            "label": f"{r.get('name') or ''}（{r.get('jb') or ''}）",
        }
        for r in rows
        if r.get("name")
    ]


@router.get("/approvers")
def get_reimbursement_approvers():
    """获取二级/三级审批人以及报销完成人候选。"""
    _ensure_table()
    second = _list_people("(jb = '副经理' OR jb LIKE '副经理%')")
    third = _list_people("(jb = '经理' OR jb LIKE '经理%')")
    completers = _list_people("lsys = '综合技术室' AND (jb LIKE '%主任%')")
    return {"success": True, "data": {"second": second, "third": third, "completers": completers}}


@router.post("/apply")
async def submit_reimbursement(
    material_name: str = Form(...),
    specification: str = Form(""),
    unit_price: float = Form(...),
    quantity: float = Form(...),
    supplier: str = Form(...),
    work_no: str = Form(""),
    part_no: str = Form(""),
    usage_detail: str = Form(...),
    applicant: str = Form(...),
    department: str = Form(""),
    approver2: str = Form(...),
    approver3: str = Form(...),
    remark: str = Form(""),
    photo: UploadFile = File(...),
    invoice: UploadFile = File(...),
):
    _ensure_table()
    _ensure_dirs()

    if not material_name.strip():
        raise HTTPException(status_code=400, detail="物资名称不能为空")
    if unit_price <= 0:
        raise HTTPException(status_code=400, detail="单价必须大于0")
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="数量必须大于0")
    if not supplier.strip():
        raise HTTPException(status_code=400, detail="供应商名称不能为空")
    if not usage_detail.strip():
        raise HTTPException(status_code=400, detail="用途（详细说明）不能为空")
    if not applicant.strip():
        raise HTTPException(status_code=400, detail="申请人不能为空")
    if not approver2.strip():
        raise HTTPException(status_code=400, detail="请选择二级审批人")
    if not approver3.strip():
        raise HTTPException(status_code=400, detail="请选择三级审批人")

    photo_stored, photo_original = _save_upload(photo, IMAGE_EXTENSIONS, "已购买的实物照片")
    invoice_stored, invoice_original = _save_upload(invoice, INVOICE_EXTENSIONS, "发票")
    await _write_upload(photo, photo_stored, "已购买的实物照片")
    await _write_upload(invoice, invoice_stored, "发票")

    total_price = round(float(unit_price) * float(quantity), 2)
    new_id = db.execute_insert(
        """
        INSERT INTO low_value_reimbursement
            (material_name, specification, unit_price, quantity, total_price,
             supplier, work_no, part_no, usage_detail,
             photo_attachment, photo_original, invoice_attachment, invoice_original,
             applicant, department, approver2, approver3, status, apply_time, remark)
        VALUES
            (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, NOW(), %s)
        """,
        (
            material_name.strip(),
            specification.strip(),
            round(float(unit_price), 2),
            round(float(quantity), 2),
            total_price,
            supplier.strip(),
            work_no.strip(),
            part_no.strip(),
            usage_detail.strip(),
            photo_stored,
            photo_original,
            invoice_stored,
            invoice_original,
            applicant.strip(),
            department.strip(),
            approver2.strip(),
            approver3.strip(),
            remark.strip(),
        ),
    )
    if new_id is None:
        raise HTTPException(status_code=500, detail="申请提交失败，请重试")
    return {"success": True, "message": "低值易耗报销申请已提交", "id": new_id}


@router.post("/invoice/parse")
async def parse_invoice(invoice: UploadFile = File(...)):
    """解析上传的发票 PDF，返回可回填字段。"""
    if not invoice or not invoice.filename:
        raise HTTPException(status_code=400, detail="请上传发票 PDF")
    ext = os.path.splitext(invoice.filename)[1].lower()
    content = await invoice.read()
    if ext != ".pdf":
        data = _extract_invoice_fields("", invoice.filename)
        data.pop("text", None)
        return {
            "success": True,
            "data": data,
            "message": "当前仅支持 PDF 发票自动解析",
        }
    text = _extract_pdf_text(content)
    data = _extract_invoice_fields(text, invoice.filename)
    data.pop("text", None)
    return {"success": True, "data": data}


@router.get("/pending")
def get_pending_reimbursements(approver: str = Query(...)):
    _ensure_table()
    name = approver.strip()
    params: list = [name, name]
    where = "(status = 0 AND approver2 = %s) OR (status = 1 AND approver3 = %s)"
    if _is_reimbursement_handler(name):
        where = f"({where}) OR status = 2"
    rows = db.execute_query(
        f"""
        SELECT *
        FROM low_value_reimbursement
        WHERE {where}
        ORDER BY apply_time DESC
        """,
        tuple(params),
    )
    for row in rows:
        _attach_display_fields(row)
    return {"success": True, "data": rows, "total": len(rows)}


def _do_action(row: dict, name: str, action: str, reject_reason: str = "") -> str:
    """对单条申请执行审批动作，成功返回提示语，失败抛出 HTTPException。"""
    rid = row.get("id")
    status = int(row.get("status") or 0)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if action == "reject":
        if status in (STATUS_COMPLETED, STATUS_REJECTED):
            raise HTTPException(status_code=400, detail="该申请已结束，无法驳回")
        if status == STATUS_PENDING_SECOND and row.get("approver2") != name:
            raise HTTPException(status_code=403, detail="您不是该申请的二级审批人")
        if status == STATUS_PENDING_THIRD and row.get("approver3") != name:
            raise HTTPException(status_code=403, detail="您不是该申请的三级审批人")
        if status == STATUS_PENDING_COMPLETE and not _is_reimbursement_handler(name):
            raise HTTPException(status_code=403, detail="仅综合技术室主任/副主任可处理该环节")
        db.execute_update(
            "UPDATE low_value_reimbursement SET status = 22, reject_reason = %s WHERE id = %s",
            ((reject_reason or "").strip()[:500], rid),
        )
        return "已驳回"

    if action == "approve":
        if status == STATUS_PENDING_SECOND:
            if row.get("approver2") != name:
                raise HTTPException(status_code=403, detail="您不是该申请的二级审批人")
            db.execute_update(
                "UPDATE low_value_reimbursement SET status = 1, approve2_time = %s WHERE id = %s AND status = 0",
                (now, rid),
            )
            return "二级审批已通过"
        if status == STATUS_PENDING_THIRD:
            if row.get("approver3") != name:
                raise HTTPException(status_code=403, detail="您不是该申请的三级审批人")
            db.execute_update(
                "UPDATE low_value_reimbursement SET status = 2, approve3_time = %s WHERE id = %s AND status = 1",
                (now, rid),
            )
            return "三级审批已通过，已流转至报销完成环节"
        raise HTTPException(status_code=400, detail="当前状态无法审批")

    if action == "complete":
        if status != STATUS_PENDING_COMPLETE:
            raise HTTPException(status_code=400, detail="仅三级审批通过后的申请可标记完成")
        if not _is_reimbursement_handler(name):
            raise HTTPException(status_code=403, detail="仅综合技术室主任/副主任可标记报销完成")
        db.execute_update(
            "UPDATE low_value_reimbursement SET status = 3, completer = %s, complete_time = %s WHERE id = %s AND status = 2",
            (name, now, rid),
        )
        return "已完成报销闭环"

    raise HTTPException(status_code=400, detail="无效操作")


@router.post("/action")
def reimbursement_action(
    id: int = Form(...),
    operator: str = Form(...),
    action: str = Form(...),
    reject_reason: str = Form(""),
):
    _ensure_table()
    name = operator.strip()
    if action not in ("approve", "reject", "complete"):
        raise HTTPException(status_code=400, detail="无效操作")
    rows = db.execute_query("SELECT * FROM low_value_reimbursement WHERE id = %s LIMIT 1", (id,))
    if not rows:
        raise HTTPException(status_code=404, detail="申请记录不存在")
    message = _do_action(rows[0], name, action, reject_reason)
    return {"success": True, "message": message}


@router.post("/action-batch")
def reimbursement_action_batch(
    ids: str = Form(...),
    operator: str = Form(...),
    action: str = Form(...),
    reject_reason: str = Form(""),
):
    """批量审批：ids 为逗号/空格分隔的申请编号。
    action=approve 时，对处于“待报销完成”状态的申请自动按“完成报销”处理，
    使领导可一键推进所选申请。"""
    _ensure_table()
    name = operator.strip()
    if action not in ("approve", "reject", "complete"):
        raise HTTPException(status_code=400, detail="无效操作")
    id_list = [int(x) for x in re.split(r"[,\s]+", ids.strip()) if x.strip().isdigit()]
    if not id_list:
        raise HTTPException(status_code=400, detail="请选择要处理的申请")

    processed = 0
    failed = []
    for rid in id_list:
        try:
            rows = db.execute_query("SELECT * FROM low_value_reimbursement WHERE id = %s LIMIT 1", (rid,))
            if not rows:
                failed.append({"id": rid, "reason": "记录不存在"})
                continue
            row = rows[0]
            act = action
            if action == "approve" and int(row.get("status") or 0) == STATUS_PENDING_COMPLETE:
                act = "complete"
            _do_action(row, name, act, reject_reason)
            processed += 1
        except HTTPException as e:
            failed.append({"id": rid, "reason": str(e.detail)})
        except Exception as e:  # noqa: BLE001
            logger.warning("批量审批失败 id=%s: %s", rid, e)
            failed.append({"id": rid, "reason": "处理失败"})

    return {
        "success": True,
        "processed": processed,
        "failed": failed,
        "total": len(id_list),
    }


@router.post("/invoice/check")
def check_reimbursement_invoices(
    operator: str = Form(""),
):
    """智能校验近一年内未驳回发票：重复发票号、同供应商同开票日期拆分风险。"""
    _ensure_table()
    _ensure_dirs()
    name = operator.strip()
    _require_ledger_permission(name)

    checked = []
    skipped = []
    rows = db.execute_query(
        """
        SELECT *
        FROM low_value_reimbursement
        WHERE status != %s
          AND apply_time >= DATE_SUB(NOW(), INTERVAL 1 YEAR)
        ORDER BY apply_time DESC
        """,
        (STATUS_REJECTED,),
    )
    for row in rows:
        rid = int(row.get("id") or 0)
        status = int(row.get("status") or 0)
        if status not in (STATUS_PENDING_SECOND, STATUS_PENDING_THIRD, STATUS_PENDING_COMPLETE, STATUS_COMPLETED):
            skipped.append({"id": rid, "reason": "当前状态不在校验范围"})
            continue

        stored = row.get("invoice_attachment") or ""
        file_path = UPLOAD_DIR / stored
        if Path(stored).suffix.lower() != ".pdf":
            skipped.append({"id": rid, "reason": "仅支持 PDF 发票校验"})
            continue
        if not file_path.exists():
            skipped.append({"id": rid, "reason": "发票文件不存在"})
            continue
        try:
            text = _extract_pdf_text(file_path.read_bytes())
            fields = _extract_invoice_fields(text, row.get("invoice_original") or stored)
        except HTTPException as e:
            skipped.append({"id": rid, "reason": str(e.detail)})
            continue
        checked.append({
            "id": rid,
            "material_name": row.get("material_name") or "",
            "applicant": row.get("applicant") or "",
            "supplier": fields.get("supplier") or row.get("supplier") or "",
            "invoice_number": fields.get("invoice_number") or "",
            "invoice_date": fields.get("invoice_date") or "",
            "invoice_attachment": stored,
            "invoice_original": row.get("invoice_original") or "",
            "total_price": _to_money(row.get("total_price")),
            "apply_time": _fmt_dt(row.get("apply_time")),
            "status": status,
            "status_text": _status_text(status),
        })

    duplicate_groups = []
    by_number: dict[str, list[dict]] = {}
    for item in checked:
        number = item.get("invoice_number") or ""
        if number:
            by_number.setdefault(number, []).append(item)
    for number, items in by_number.items():
        if len(items) > 1:
            duplicate_groups.append({"invoice_number": number, "items": items})

    split_groups = []
    by_supplier_date: dict[tuple[str, str], list[dict]] = {}
    for item in checked:
        supplier_key = _normalize_supplier(item.get("supplier") or "")
        invoice_date = item.get("invoice_date") or ""
        if supplier_key and invoice_date:
            by_supplier_date.setdefault((supplier_key, invoice_date), []).append(item)
    for (supplier_key, invoice_date), items in by_supplier_date.items():
        if len(items) > 1:
            supplier = items[0].get("supplier") or supplier_key
            split_groups.append({"supplier": supplier, "invoice_date": invoice_date, "items": items})

    risk_reasons_by_id: dict[int, list[str]] = {}
    for group in duplicate_groups:
        for item in group["items"]:
            risk_reasons_by_id.setdefault(item["id"], []).append("发票号码重复")
    for group in split_groups:
        for item in group["items"]:
            risk_reasons_by_id.setdefault(item["id"], []).append("疑似拆分报销")
    for item in checked:
        reasons = risk_reasons_by_id.get(item["id"], [])
        item["check_passed"] = not reasons
        item["risk_reasons"] = reasons

    return {
        "success": True,
        "data": {
            "checked": checked,
            "skipped": skipped,
            "duplicate_invoices": duplicate_groups,
            "split_risks": split_groups,
            "summary": {
                "checked_count": len(checked),
                "passed_count": sum(1 for item in checked if item["check_passed"]),
                "skipped_count": len(skipped),
                "duplicate_count": len(duplicate_groups),
                "split_risk_count": len(split_groups),
                "scope": "近一年未驳回申请",
            },
        },
    }


def _remove_attachment_file(filename: str) -> None:
    if not filename:
        return
    path = UPLOAD_DIR / filename
    try:
        if path.exists():
            path.unlink()
    except Exception as e:
        logger.warning("删除附件失败 %s: %s", filename, e)


@router.post("/delete")
def delete_rejected_reimbursement(
    id: int = Form(...),
    operator: str = Form(...),
):
    """删除已驳回的申请。申请人可删自己的；台账查看权限者可删任意已驳回记录。"""
    _ensure_table()
    name = operator.strip()
    if not name:
        raise HTTPException(status_code=400, detail="请提供操作人")
    rows = db.execute_query("SELECT * FROM low_value_reimbursement WHERE id = %s LIMIT 1", (id,))
    if not rows:
        raise HTTPException(status_code=404, detail="申请记录不存在")
    row = rows[0]
    if int(row.get("status") or 0) != STATUS_REJECTED:
        raise HTTPException(status_code=400, detail="仅已驳回的申请可删除")
    applicant = (row.get("applicant") or "").strip()
    if applicant != name and not _can_view_ledger(name):
        raise HTTPException(status_code=403, detail="无权删除该申请")
    _remove_attachment_file(row.get("photo_attachment") or "")
    _remove_attachment_file(row.get("invoice_attachment") or "")
    db.execute_update(
        "DELETE FROM low_value_reimbursement WHERE id = %s AND status = %s",
        (id, STATUS_REJECTED),
    )
    return {"success": True, "message": "已删除"}


@router.get("/budget/summary")
def get_budget_summary(
    year: int = Query(0, ge=0),
    current_user: str = Query(""),
):
    """获取指定年度低值易耗额度统计（结余 / 已完成 / 审核中）。"""
    _ensure_table()
    _require_ledger_permission(current_user)
    target_year = year or datetime.now().year
    return {"success": True, "data": _budget_summary(target_year)}


@router.get("/budget/list")
def list_budget_years(current_user: str = Query("")):
    """列出已配置的年度额度。"""
    _ensure_table()
    _require_ledger_permission(current_user)
    rows = db.execute_query(
        """
        SELECT budget_year, total_amount, remark, updated_by, updated_at
        FROM low_value_budget
        ORDER BY budget_year DESC
        """
    )
    data = []
    for row in rows:
        data.append({
            "year": int(row.get("budget_year") or 0),
            "total_amount": _to_money(row.get("total_amount")),
            "remark": row.get("remark") or "",
            "updated_by": row.get("updated_by") or "",
            "updated_at": _fmt_dt(row.get("updated_at")),
        })
    return {"success": True, "data": data}


@router.post("/budget")
def upsert_budget(
    budget_year: int = Form(...),
    total_amount: float = Form(...),
    remark: str = Form(""),
    operator: str = Form(""),
):
    """新增或更新某年度低值易耗总额度。"""
    _ensure_table()
    _require_ledger_permission(operator)
    year = int(budget_year)
    if year < 2000 or year > 2100:
        raise HTTPException(status_code=400, detail="请输入有效年度")
    amount = _to_money(total_amount)
    if amount < 0:
        raise HTTPException(status_code=400, detail="额度不能为负数")
    name = (operator or "").strip()
    existing = db.execute_query(
        "SELECT id FROM low_value_budget WHERE budget_year = %s LIMIT 1",
        (year,),
    )
    if existing:
        db.execute_update(
            """
            UPDATE low_value_budget
            SET total_amount = %s, remark = %s, updated_by = %s
            WHERE budget_year = %s
            """,
            (amount, (remark or "").strip()[:500], name, year),
        )
    else:
        db.execute_insert(
            """
            INSERT INTO low_value_budget (budget_year, total_amount, remark, updated_by)
            VALUES (%s, %s, %s, %s)
            """,
            (year, amount, (remark or "").strip()[:500], name),
        )
    return {"success": True, "message": "年度额度已保存", "data": _budget_summary(year)}


@router.get("/records")
def get_reimbursement_records(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: str = Query(""),
    status: str = Query(""),
    date_from: str = Query(""),
    date_to: str = Query(""),
    current_user: str = Query(""),
):
    _ensure_table()
    _require_ledger_permission(current_user)
    where, params = _build_records_where(keyword, status, date_from, date_to)
    total = db.execute_scalar(f"SELECT COUNT(*) AS cnt FROM low_value_reimbursement WHERE {where}", tuple(params)) or 0
    offset = (page - 1) * page_size
    rows = db.execute_query(
        f"""
        SELECT *
        FROM low_value_reimbursement
        WHERE {where}
        ORDER BY apply_time DESC
        LIMIT %s OFFSET %s
        """,
        tuple(params) + (page_size, offset),
    )
    for row in rows:
        _attach_display_fields(row)
    return {"success": True, "data": rows, "total": total, "page": page, "page_size": page_size}


@router.get("/my-applications")
def get_my_reimbursements(name: str = Query(...)):
    _ensure_table()
    rows = db.execute_query(
        "SELECT * FROM low_value_reimbursement WHERE applicant = %s ORDER BY apply_time DESC",
        (name.strip(),),
    )
    for row in rows:
        _attach_display_fields(row)
    return {"success": True, "data": rows}


_CONTENT_TYPE_MAP = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".gif": "image/gif",
    ".bmp": "image/bmp",
    ".webp": "image/webp",
    ".pdf": "application/pdf",
    ".ofd": "application/octet-stream",
}


@router.get("/attachment")
def download_reimbursement_attachment(
    kind: str = Query(..., pattern="^(photo|invoice)$"),
    filename: str = Query(...),
    disposition: str = Query("attachment", pattern="^(attachment|inline)$"),
):
    _ensure_dirs()
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="附件不存在")
    if kind == "photo":
        rows = db.execute_query(
            "SELECT photo_original AS original_name FROM low_value_reimbursement WHERE photo_attachment = %s LIMIT 1",
            (filename,),
        )
    else:
        rows = db.execute_query(
            "SELECT invoice_original AS original_name FROM low_value_reimbursement WHERE invoice_attachment = %s LIMIT 1",
            (filename,),
        )
    original_name = rows[0]["original_name"] if rows else filename
    ext = Path(filename).suffix.lower()

    if disposition == "inline":
        # 浏览器内预览：使用真实 content-type 并以 inline 方式展示图片/PDF
        media_type = _CONTENT_TYPE_MAP.get(ext, "application/octet-stream")
        headers = {"Content-Disposition": "inline"}
        return FileResponse(path=str(file_path), media_type=media_type, headers=headers)

    return FileResponse(path=str(file_path), filename=original_name, media_type="application/octet-stream")


@router.get("/export")
def export_reimbursement_ledger(
    keyword: str = Query(""),
    status: str = Query(""),
    date_from: str = Query(""),
    date_to: str = Query(""),
    current_user: str = Query(""),
):
    _ensure_table()
    _require_ledger_permission(current_user)
    _ensure_dirs()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl，无法生成 Excel")

    where, params = _build_records_where(keyword, status, date_from, date_to)
    # 导出台账不包含已驳回记录
    where = f"({where}) AND status != %s"
    params = list(params) + [STATUS_REJECTED]
    rows = db.execute_query(
        f"""
        SELECT *
        FROM low_value_reimbursement
        WHERE {where}
        ORDER BY apply_time DESC
        """,
        tuple(params),
    )

    headers = [
        "物资名称", "规格", "单价", "数量", "总价", "供应商名称", "工作号/科研号", "部套号",
        "用途（详细说明）", "上传已购买的实物照片", "上传发票", "申请人", "二级审批人", "三级审批人", "申请日期", "进度",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "低值易耗报销台账"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1, value="低值易耗报销台账")
    title.font = Font(name="宋体", size=16, bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(name="宋体", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
        cell.border = border

    for r_idx, row in enumerate(rows, 3):
        ws.row_dimensions[r_idx].height = 70
        values = [
            row.get("material_name") or "",
            row.get("specification") or "",
            float(row.get("unit_price") or 0),
            float(row.get("quantity") or 0),
            float(row.get("total_price") or 0),
            row.get("supplier") or "",
            row.get("work_no") or "",
            row.get("part_no") or "",
            row.get("usage_detail") or "",
            row.get("photo_original") or "",
            row.get("invoice_original") or "",
            row.get("applicant") or "",
            row.get("approver2") or "",
            row.get("approver3") or "",
            _fmt_date(row.get("apply_time")),
            _status_text(row.get("status")),
        ]
        for c_idx, value in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name="宋体", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        photo_cell = f"J{r_idx}"
        invoice_cell = f"K{r_idx}"
        if _add_image_to_sheet(ws, _image_source_for_excel(row.get("photo_attachment") or ""), photo_cell):
            ws[photo_cell].value = ""
        if _add_image_to_sheet(ws, _image_source_for_excel(row.get("invoice_attachment") or ""), invoice_cell):
            ws[invoice_cell].value = ""

    widths = [15.33, 12, 10, 10, 10, 14, 20, 15.5, 24, 24, 16, 12, 13.5, 15.8, 14, 14]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    filename = f"low_value_reimbursement_{uuid.uuid4().hex}.xlsx"
    path = EXPORT_DIR / filename
    wb.save(path)
    return FileResponse(
        path=str(path),
        filename=f"低值易耗报销台账_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/invoices/export-zip")
def export_invoice_zip(
    keyword: str = Query(""),
    status: str = Query(""),
    date_from: str = Query(""),
    date_to: str = Query(""),
    current_user: str = Query(""),
):
    _ensure_table()
    _require_ledger_permission(current_user)
    _ensure_dirs()
    where, params = _build_records_where(keyword, status, date_from, date_to)
    rows = db.execute_query(
        f"""
        SELECT id, applicant, invoice_attachment, invoice_original, apply_time
        FROM low_value_reimbursement
        WHERE {where}
        ORDER BY apply_time DESC
        """,
        tuple(params),
    )
    zip_path = EXPORT_DIR / f"low_value_invoices_{uuid.uuid4().hex}.zip"
    added = 0
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for row in rows:
            stored = row.get("invoice_attachment") or ""
            if Path(stored).suffix.lower() != ".pdf":
                continue
            file_path = UPLOAD_DIR / stored
            if not file_path.exists():
                continue
            apply_date = _fmt_date(row.get("apply_time")) or "no-date"
            original = _safe_zip_name(row.get("invoice_original") or stored)
            arcname = _safe_zip_name(f"{apply_date}_{row.get('id')}_{row.get('applicant') or ''}_{original}")
            zf.write(file_path, arcname)
            added += 1
        if added == 0:
            zf.writestr("README.txt", "当前筛选条件下没有可导出的 PDF 发票。")
    return FileResponse(
        path=str(zip_path),
        filename=f"低值易耗发票PDF_{datetime.now().strftime('%Y%m%d%H%M%S')}.zip",
        media_type="application/zip",
    )
