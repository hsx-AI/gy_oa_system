# -*- coding: utf-8 -*-
"""
排班管理 API
"""
import calendar
import json
import logging
import re
from datetime import datetime, date, timedelta
from io import BytesIO
from typing import Optional, List, Set, Tuple
from urllib.parse import quote

from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel, Field
from database import db, db_demo
from utils.holiday_loader import load_holidays_for_year

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/shift", tags=["排班管理"])

SHIFT_DEPARTMENT_ORDER = [
    "水轮机工艺室",
    "水发工艺室",
    "汽发工艺室",
    "焊接工艺室",
    "综合技术室",
    "智能制造技术室",
    "工具技术室",
    "非标技术室",
    "数控编程室",
]

SHIFT_EMAIL_DEPARTMENTS_CONFIG_COLUMN = "shift_schedule_email_departments"


def _ensure_tables():
    """确保排班相关表存在。

    注意：database.execute_query 在表不存在时内部吞掉异常并返回 []，不会向外抛，
    因此不能用「先 SELECT 再 CREATE」的 try/except 判断；应直接 CREATE TABLE IF NOT EXISTS。
    """
    db.execute_update("""
        CREATE TABLE IF NOT EXISTS shift_config (
          id INT AUTO_INCREMENT PRIMARY KEY,
          department VARCHAR(100) NOT NULL,
          workday_day INT NOT NULL DEFAULT 2,
          workday_night INT NOT NULL DEFAULT 2,
          weekend_day INT NOT NULL DEFAULT 2,
          weekend_night INT NOT NULL DEFAULT 2,
          email_recipients TEXT NULL,
          updated_by VARCHAR(50) NULL,
          updated_at DATETIME NULL,
          UNIQUE KEY uk_dept (department)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    # 极旧库在增加 workday_day 前已建过 shift_config 时补列
    col_rows = db.execute_query(
        "SELECT 1 FROM information_schema.COLUMNS "
        "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'shift_config' AND COLUMN_NAME = 'workday_day' "
        "LIMIT 1"
    )
    if not col_rows:
        db.execute_update(
            "ALTER TABLE shift_config ADD COLUMN workday_day INT NOT NULL DEFAULT 2 AFTER department"
        )
    recipient_col = db.execute_query(
        "SELECT 1 FROM information_schema.COLUMNS "
        "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'shift_config' AND COLUMN_NAME = 'email_recipients' "
        "LIMIT 1"
    )
    if not recipient_col:
        db.execute_update(
            "ALTER TABLE shift_config ADD COLUMN email_recipients TEXT NULL "
            "COMMENT '排班邮件收件人JSON [{name,email}]' AFTER weekend_night"
        )
    send_wd_col = db.execute_query(
        "SELECT 1 FROM information_schema.COLUMNS "
        "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'shift_config' AND COLUMN_NAME = 'email_send_weekday' "
        "LIMIT 1"
    )
    if not send_wd_col:
        db.execute_update(
            "ALTER TABLE shift_config ADD COLUMN email_send_weekday INT NOT NULL DEFAULT 4 "
            "COMMENT '排班邮件自动发送星期几(0=周一…6=周日)，固定17:00发送' AFTER email_recipients"
        )
    include_send_col = db.execute_query(
        "SELECT 1 FROM information_schema.COLUMNS "
        "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'shift_config' AND COLUMN_NAME = 'email_include_send_day' "
        "LIMIT 1"
    )
    if not include_send_col:
        db.execute_update(
            "ALTER TABLE shift_config ADD COLUMN email_include_send_day TINYINT(1) NOT NULL DEFAULT 0 "
            "COMMENT '排班邮件区间是否含发送当天(0=否,次日始;1=是,发送日至下周同日前一天)' AFTER email_send_weekday"
        )
    start_offset_col = db.execute_query(
        "SELECT 1 FROM information_schema.COLUMNS "
        "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'shift_config' AND COLUMN_NAME = 'email_start_offset_days' "
        "LIMIT 1"
    )
    if not start_offset_col:
        db.execute_update(
            "ALTER TABLE shift_config ADD COLUMN email_start_offset_days INT NOT NULL DEFAULT 1 "
            "COMMENT '排班邮件区间开始日相对发送日偏移天数(0=当天,1=次日,3=三天后)' AFTER email_include_send_day"
        )
        db.execute_update(
            "UPDATE shift_config SET email_start_offset_days = CASE WHEN email_include_send_day = 1 THEN 0 ELSE 1 END"
        )
    holiday_days_col = db.execute_query(
        "SELECT 1 FROM information_schema.COLUMNS "
        "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'shift_config' AND COLUMN_NAME = 'holiday_email_days_before' "
        "LIMIT 1"
    )
    if not holiday_days_col:
        db.execute_update(
            "ALTER TABLE shift_config ADD COLUMN holiday_email_days_before INT NOT NULL DEFAULT -1 "
            "COMMENT '节假日值班表邮件：假期首日前几天17:00自动发送(-1=关闭,0=首日当天,1=前1天…)' AFTER email_start_offset_days"
        )
    db.execute_update("""
        CREATE TABLE IF NOT EXISTS shift_schedule (
          id INT AUTO_INCREMENT PRIMARY KEY,
          department VARCHAR(100) NOT NULL,
          employee_name VARCHAR(50) NOT NULL,
          shift_date DATE NOT NULL,
          shift_type VARCHAR(10) NOT NULL DEFAULT '',
          year INT NOT NULL,
          month INT NOT NULL,
          updated_by VARCHAR(50) NULL,
          updated_at DATETIME NULL,
          UNIQUE KEY uk_emp_date (employee_name, shift_date),
          INDEX idx_dept_ym (department, year, month)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    # shift_location 字段：值班位置（准备组/服务组）
    loc_col = db.execute_query(
        "SELECT 1 FROM information_schema.COLUMNS "
        "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'shift_schedule' AND COLUMN_NAME = 'shift_location' "
        "LIMIT 1"
    )
    if not loc_col:
        try:
            db.execute_update(
                "ALTER TABLE shift_schedule ADD COLUMN shift_location VARCHAR(20) NOT NULL DEFAULT '' "
                "COMMENT '值班位置：准备组/服务组' AFTER shift_type"
            )
        except Exception:
            pass

    db.execute_update("""
        CREATE TABLE IF NOT EXISTS shift_day_plan (
          id INT AUTO_INCREMENT PRIMARY KEY,
          department VARCHAR(100) NOT NULL,
          plan_date DATE NOT NULL,
          content TEXT NULL,
          updated_by VARCHAR(50) NULL,
          updated_at DATETIME NULL,
          UNIQUE KEY uk_dept_plan_date (department, plan_date),
          INDEX idx_dept_date (department, plan_date)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    db.execute_update("""
        CREATE TABLE IF NOT EXISTS shift_day_lock (
          id INT AUTO_INCREMENT PRIMARY KEY,
          department VARCHAR(100) NOT NULL,
          lock_date DATE NOT NULL,
          is_open TINYINT(1) NOT NULL DEFAULT 0,
          opened_by VARCHAR(50) NULL,
          opened_at DATETIME NULL,
          UNIQUE KEY uk_dept_lock_date (department, lock_date)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)
    db.execute_update("""
        CREATE TABLE IF NOT EXISTS shift_day_noduty (
          id INT AUTO_INCREMENT PRIMARY KEY,
          department VARCHAR(100) NOT NULL,
          noduty_date DATE NOT NULL,
          updated_by VARCHAR(50) NULL,
          updated_at DATETIME NULL,
          UNIQUE KEY uk_dept_noduty_date (department, noduty_date),
          INDEX idx_dept_noduty_date (department, noduty_date)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """)


def _load_no_duty_dates(department: str, date_lo: str, date_hi: str) -> Set[str]:
    """读取科室在区间内标记为「当日不设置值班」的日期集合。"""
    try:
        rows = db.execute_query(
            "SELECT noduty_date FROM shift_day_noduty "
            "WHERE department = %s AND noduty_date >= %s AND noduty_date <= %s",
            (department, date_lo, date_hi),
        )
    except Exception as e:
        logger.warning("读取 shift_day_noduty 失败: %s", e)
        return set()
    result = set()
    for r in rows or []:
        nd = r.get("noduty_date")
        if hasattr(nd, "strftime"):
            result.add(nd.strftime("%Y-%m-%d"))
        else:
            ds = str(nd)[:10] if nd else ""
            if ds:
                result.add(ds)
    return result


def _get_dept_employees(department: str) -> List[str]:
    """获取该科室在职员工列表（排除名字末尾为1的测试账号）"""
    rows = db.execute_query(
        "SELECT name FROM yggl WHERE lsys = %s AND COALESCE(zaizhi,0) = 0 "
        "AND RIGHT(TRIM(name),1) != '1' ORDER BY gh",
        (department,),
    )
    return [(r.get("name") or "").strip() for r in rows if (r.get("name") or "").strip()]


def _get_shift_departments() -> List[str]:
    rows = db.execute_query(
        "SELECT DISTINCT lsys FROM yggl WHERE lsys IS NOT NULL AND lsys != '' "
        "AND RIGHT(TRIM(lsys),1) != '1' AND TRIM(lsys) != '部办' "
        "AND TRIM(lsys) NOT IN ('其他部门员工','其他部门成员') "
        "AND COALESCE(zaizhi,0) = 0 ORDER BY lsys"
    )
    depts = [(r.get("lsys") or "").strip() for r in rows if (r.get("lsys") or "").strip()]

    def _dept_sort_key(name: str):
        try:
            return (SHIFT_DEPARTMENT_ORDER.index(name), "")
        except ValueError:
            return (len(SHIFT_DEPARTMENT_ORDER), name)

    return sorted(depts, key=_dept_sort_key)


def _ensure_shift_email_feature_config_column() -> None:
    """确保 webconfig 中有排班邮件功能启用科室配置列。"""
    exists = db.execute_query(
        "SELECT 1 FROM information_schema.COLUMNS "
        "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'webconfig' AND COLUMN_NAME = %s LIMIT 1",
        (SHIFT_EMAIL_DEPARTMENTS_CONFIG_COLUMN,),
    )
    if exists:
        return
    db.execute_update(
        f"ALTER TABLE webconfig ADD COLUMN {SHIFT_EMAIL_DEPARTMENTS_CONFIG_COLUMN} MEDIUMTEXT NULL "
        "COMMENT '启用排班邮件功能的科室JSON数组'"
    )


def _load_shift_email_feature_config(all_departments: Optional[List[str]] = None) -> Tuple[Set[str], bool]:
    """返回启用排班邮件功能的科室集合；未配置时默认全部启用。"""
    departments = all_departments if all_departments is not None else _get_shift_departments()
    valid = set(departments)
    _ensure_shift_email_feature_config_column()
    try:
        rows = db.execute_query(
            f"SELECT {SHIFT_EMAIL_DEPARTMENTS_CONFIG_COLUMN} FROM webconfig WHERE id = %s LIMIT 1",
            ("1",),
        )
        raw = rows[0].get(SHIFT_EMAIL_DEPARTMENTS_CONFIG_COLUMN) if rows else None
        if raw is None or raw == "":
            return set(departments), False
        if isinstance(raw, (bytes, bytearray)):
            raw = raw.decode("utf-8", errors="ignore")
        data = json.loads(raw) if isinstance(raw, str) else raw
        if not isinstance(data, list):
            return set(departments), False
        enabled = {(str(item) or "").strip() for item in data}
        enabled = {name for name in enabled if name in valid}
        return enabled, True
    except Exception as e:
        logger.warning("读取排班邮件功能科室配置失败: %s", e)
        return set(departments), False


def _is_shift_email_feature_enabled(department: str) -> bool:
    dept = (department or "").strip()
    if not dept:
        return False
    enabled, _configured = _load_shift_email_feature_config()
    return dept in enabled


def _get_shift_dept_leader_recipients(department: str) -> List[dict]:
    """本科室主任/副主任/班组长且已配置企业邮箱（与周排班邮件发送对象一致）。"""
    from routers.approvers import _jb_match

    rows = db.execute_query(
        "SELECT name, jb, enterprise_email FROM yggl "
        "WHERE lsys = %s AND name IS NOT NULL AND TRIM(name) != '' "
        "AND COALESCE(zaizhi,0) = 0",
        (department,),
    )
    leaders = []
    for r in rows or []:
        jb = (r.get("jb") or "").strip()
        if not (_jb_match(jb, "主任") or _jb_match(jb, "副主任") or _jb_match(jb, "组长")):
            continue
        email = (r.get("enterprise_email") or "").strip()
        if not email:
            continue
        leaders.append({
            "name": (r.get("name") or "").strip(),
            "email": email,
            "jb": jb,
        })
    return leaders


def _is_shift_company_leader_jb(jb: str) -> bool:
    """公司级领导：经理 / 副经理 / 经理助理。"""
    from routers.approvers import _jb_match

    j = (jb or "").strip()
    if not j:
        return False
    if _jb_match(j, "经理"):
        return True
    if j == "副经理" or j.startswith("副经理"):
        return True
    return False


def _get_shift_company_leader_recipients() -> List[dict]:
    """全公司领导抄送：yggl.jb 为经理/副经理/经理助理且已配置企业邮箱。"""
    rows = db.execute_query(
        "SELECT name, jb, enterprise_email FROM yggl "
        "WHERE name IS NOT NULL AND TRIM(name) != '' "
        "AND COALESCE(zaizhi,0) = 0",
        (),
    )
    leaders = []
    seen = set()
    for r in rows or []:
        jb = (r.get("jb") or "").strip()
        if not _is_shift_company_leader_jb(jb):
            continue
        email = (r.get("enterprise_email") or "").strip()
        if not email:
            continue
        key = email.lower()
        if key in seen:
            continue
        seen.add(key)
        leaders.append({
            "name": (r.get("name") or "").strip(),
            "email": email,
            "jb": jb,
        })
    leaders.sort(key=lambda x: (x.get("jb") or "", x.get("name") or ""))
    return leaders


def _get_shift_email_feature_config_items() -> dict:
    departments = _get_shift_departments()
    enabled, configured = _load_shift_email_feature_config(departments)
    _ensure_tables()
    rows = db.execute_query(
        "SELECT department, email_recipients, email_send_weekday, email_include_send_day, email_start_offset_days, holiday_email_days_before FROM shift_config"
    )
    by_dept = {(r.get("department") or "").strip(): r for r in (rows or [])}
    items = []
    for dept in departments:
        row = by_dept.get(dept) or {}
        items.append({
            "department": dept,
            "enabled": dept in enabled,
            "email_send_weekday": _normalize_email_send_weekday(row.get("email_send_weekday")),
            "email_include_send_day": _normalize_email_include_send_day(row.get("email_include_send_day")),
            "email_start_offset_days": _normalize_email_start_offset_days(
                row.get("email_start_offset_days"),
                row.get("email_include_send_day"),
            ),
            "holiday_email_days_before": _normalize_holiday_email_days_before(
                row.get("holiday_email_days_before") if row else DEFAULT_HOLIDAY_EMAIL_DAYS_BEFORE
            ),
            "email_recipients": _parse_shift_email_recipients(row.get("email_recipients")),
            "leader_recipients": _get_shift_dept_leader_recipients(dept),
        })
    return {
        "departments": departments,
        "enabledDepartments": [dept for dept in departments if dept in enabled],
        "configured": configured,
        "company_leader_recipients": _get_shift_company_leader_recipients(),
        "items": items,
    }


def _upsert_shift_email_settings(
    department: str,
    email_send_weekday: int,
    email_recipients,
    updated_by: str,
    email_include_send_day: bool = False,
    email_start_offset_days=None,
    holiday_email_days_before=None,
) -> None:
    """仅更新科室排班邮件发送时间与收件人（系统管理员配置）。"""
    _ensure_tables()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    recipients_json = json.dumps(_normalize_shift_email_recipients(email_recipients), ensure_ascii=False)
    send_wd = _normalize_email_send_weekday(email_send_weekday)
    start_offset = _normalize_email_start_offset_days(email_start_offset_days, email_include_send_day)
    include_send = 1 if start_offset == 0 else 0
    holiday_days = _normalize_holiday_email_days_before(holiday_email_days_before)
    existing = db.execute_query(
        "SELECT id FROM shift_config WHERE department = %s LIMIT 1", (department,)
    )
    if existing:
        db.execute_update(
            "UPDATE shift_config SET email_recipients=%s, email_send_weekday=%s, "
            "email_include_send_day=%s, email_start_offset_days=%s, holiday_email_days_before=%s, "
            "updated_by=%s, updated_at=%s WHERE department=%s",
            (recipients_json, send_wd, include_send, start_offset, holiday_days, updated_by, now, department),
        )
    else:
        db.execute_update(
            "INSERT INTO shift_config (department, workday_day, workday_night, weekend_day, weekend_night, "
            "email_recipients, email_send_weekday, email_include_send_day, email_start_offset_days, holiday_email_days_before, updated_by, updated_at) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (department, 2, 2, 2, 2, recipients_json, send_wd, include_send, start_offset, holiday_days, updated_by, now),
        )


def _save_shift_email_feature_config(
    enabled_departments: List[str],
    department_email_settings: Optional[List] = None,
    updated_by: str = "",
) -> dict:
    departments = _get_shift_departments()
    valid = set(departments)
    normalized = []
    seen = set()
    for item in enabled_departments or []:
        dept = (item or "").strip()
        if not dept or dept in seen or dept not in valid:
            continue
        seen.add(dept)
        normalized.append(dept)
    _ensure_shift_email_feature_config_column()
    db.execute_update(
        f"UPDATE webconfig SET {SHIFT_EMAIL_DEPARTMENTS_CONFIG_COLUMN} = %s WHERE id = %s",
        (json.dumps(normalized, ensure_ascii=False), "1"),
    )
    if department_email_settings is not None:
        by_dept = {}
        for item in department_email_settings:
            if isinstance(item, dict):
                dept = (item.get("department") or "").strip()
                send_wd = item.get("email_send_weekday", DEFAULT_EMAIL_SEND_WEEKDAY)
                include_send = item.get("email_include_send_day", False)
                start_offset = item.get("email_start_offset_days", None)
                holiday_days = item.get("holiday_email_days_before", DEFAULT_HOLIDAY_EMAIL_DAYS_BEFORE)
                recipients = item.get("email_recipients") or []
            else:
                dept = (getattr(item, "department", None) or "").strip()
                send_wd = getattr(item, "email_send_weekday", DEFAULT_EMAIL_SEND_WEEKDAY)
                include_send = getattr(item, "email_include_send_day", False)
                start_offset = getattr(item, "email_start_offset_days", None)
                holiday_days = getattr(item, "holiday_email_days_before", DEFAULT_HOLIDAY_EMAIL_DAYS_BEFORE)
                recipients = getattr(item, "email_recipients", None) or []
            if not dept or dept not in valid:
                continue
            by_dept[dept] = (send_wd, include_send, start_offset, holiday_days, recipients)
        for dept in departments:
            if dept in by_dept:
                send_wd, include_send, start_offset, holiday_days, recipients = by_dept[dept]
            else:
                send_wd, include_send, start_offset, holiday_days, recipients = (
                    DEFAULT_EMAIL_SEND_WEEKDAY, False, DEFAULT_EMAIL_START_OFFSET_DAYS,
                    DEFAULT_HOLIDAY_EMAIL_DAYS_BEFORE, [],
                )
            _upsert_shift_email_settings(
                dept,
                send_wd,
                recipients,
                updated_by,
                email_include_send_day=include_send,
                email_start_offset_days=start_offset,
                holiday_email_days_before=holiday_days,
            )
    return _get_shift_email_feature_config_items()


def _get_dept_people(department: str) -> List[dict]:
    """获取科室在职人员及职务、联系方式，用于节假日值班表导出。"""
    rows = db.execute_query(
        "SELECT name, gh, jb, sfzh FROM yggl WHERE lsys = %s AND COALESCE(zaizhi,0) = 0 "
        "AND RIGHT(TRIM(name),1) != '1' ORDER BY gh",
        (department,),
    )
    people = []
    sfzh_map = {}
    for r in rows or []:
        name = (r.get("name") or "").strip()
        if not name:
            continue
        sfzh = (r.get("sfzh") or "").strip().replace(" ", "")
        item = {
            "name": name,
            "gh": (r.get("gh") or "").strip(),
            "jb": (r.get("jb") or "").strip(),
            "sfzh": sfzh,
            "mobile": "",
            "telephone": "",
        }
        people.append(item)
        if sfzh:
            sfzh_map[sfzh] = item

    if sfzh_map:
        ids = list(sfzh_map.keys())
        for i in range(0, len(ids), 100):
            batch = ids[i:i + 100]
            ph = ",".join(["%s"] * len(batch))
            try:
                phone_rows = db_demo.execute_query(
                    f"SELECT id_card, mobile, telephone FROM employee_info WHERE id_card IN ({ph})",
                    tuple(batch),
                )
                for pr in phone_rows or []:
                    p = sfzh_map.get((pr.get("id_card") or "").strip())
                    if p:
                        p["mobile"] = (pr.get("mobile") or "").strip()
                        p["telephone"] = (pr.get("telephone") or "").strip()
            except Exception as e:
                logger.warning("查询员工联系方式失败: %s", e)
    return people


def _get_dept_business_trips(department: str, start: date, end: date) -> dict:
    """查询科室已审批通过且与日期区间重叠的公出，返回 {姓名: {日期: 项目名}}。"""
    business_trips = {}
    try:
        trip_rows = db.execute_query(
            "SELECT g.gcr, g.xmmc, "
            "  COALESCE(g.gcsj, g.yjcfsj) AS trip_start, "
            "  COALESCE(g.sjfhtime, g.yjfhsj) AS trip_end "
            "FROM gcsqb g "
            "JOIN yggl y ON g.gcr = y.name AND COALESCE(y.zaizhi, 0) = 0 "
            "WHERE y.lsys = %s "
            "  AND g.bldzt = 2 AND g.szrzt = 2 "
            "  AND COALESCE(g.gcsj, g.yjcfsj) IS NOT NULL "
            "  AND COALESCE(g.sjfhtime, g.yjfhsj) IS NOT NULL "
            "  AND DATE(COALESCE(g.gcsj, g.yjcfsj)) <= %s "
            "  AND DATE(COALESCE(g.sjfhtime, g.yjfhsj)) >= %s",
            (department, end.strftime("%Y-%m-%d"), start.strftime("%Y-%m-%d")),
        )
        for tr in trip_rows or []:
            name = (tr.get("gcr") or "").strip()
            xmmc = (tr.get("xmmc") or "").strip()
            ts = tr.get("trip_start")
            te = tr.get("trip_end")
            if not name or not ts or not te:
                continue
            if isinstance(ts, datetime):
                ts = ts.date()
            elif isinstance(ts, str):
                ts = datetime.strptime(ts[:10], "%Y-%m-%d").date()
            elif not isinstance(ts, date):
                continue
            if isinstance(te, datetime):
                te = te.date()
            elif isinstance(te, str):
                te = datetime.strptime(te[:10], "%Y-%m-%d").date()
            elif not isinstance(te, date):
                continue
            d_cur = max(ts, start)
            d_end = min(te, end)
            while d_cur <= d_end:
                ds = d_cur.strftime("%Y-%m-%d")
                business_trips.setdefault(name, {})[ds] = xmmc or "公出"
                d_cur += timedelta(days=1)
    except Exception as e:
        logger.warning("查询公出数据失败: %s", e)
    return business_trips


def _month_dates(year: int, month: int) -> List[date]:
    """返回指定月份的所有日期列表"""
    d = date(year, month, 1)
    dates = []
    while d.month == month:
        dates.append(d)
        d += timedelta(days=1)
    return dates


def _parse_iso_date(s: str) -> Optional[date]:
    if not s:
        return None
    s = str(s).strip()[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def _daterange(start: date, end: date) -> List[date]:
    out = []
    d = start
    while d <= end:
        out.append(d)
        d += timedelta(days=1)
    return out


DEFAULT_EMAIL_SEND_WEEKDAY = 4  # 周五；与历史「周六至下周五」排班邮件周期一致


DEFAULT_EMAIL_START_OFFSET_DAYS = 1


def _normalize_email_send_weekday(value) -> int:
    """0=周一 … 6=周日"""
    try:
        wd = int(value)
    except (TypeError, ValueError):
        wd = DEFAULT_EMAIL_SEND_WEEKDAY
    if wd < 0 or wd > 6:
        return DEFAULT_EMAIL_SEND_WEEKDAY
    return wd


def _normalize_email_include_send_day(value) -> bool:
    if value in (True, 1, "1", "true", "True", "yes", "是"):
        return True
    return False


def _normalize_email_start_offset_days(value, include_send_day=None) -> int:
    try:
        offset = int(value)
    except (TypeError, ValueError):
        if include_send_day is not None:
            return 0 if _normalize_email_include_send_day(include_send_day) else 1
        return DEFAULT_EMAIL_START_OFFSET_DAYS
    if offset < 0:
        return 0
    if offset > 6:
        return 6
    return offset


DEFAULT_HOLIDAY_EMAIL_DAYS_BEFORE = -1  # -1=关闭


def _normalize_holiday_email_days_before(value) -> int:
    """节假日值班表邮件提前天数：-1=关闭，0=假期首日当天，1..15=节前N天。"""
    try:
        days = int(value)
    except (TypeError, ValueError):
        return DEFAULT_HOLIDAY_EMAIL_DAYS_BEFORE
    if days < 0:
        return -1
    if days > 15:
        return 15
    return days


def _get_shift_holiday_email_days_before(department: str) -> int:
    _ensure_tables()
    if not department:
        return DEFAULT_HOLIDAY_EMAIL_DAYS_BEFORE
    rows = db.execute_query(
        "SELECT holiday_email_days_before FROM shift_config WHERE department = %s LIMIT 1",
        (department,),
    )
    if rows:
        return _normalize_holiday_email_days_before(rows[0].get("holiday_email_days_before"))
    return DEFAULT_HOLIDAY_EMAIL_DAYS_BEFORE


def _week_range_for_send_day(
    anchor: date,
    send_weekday: int,
    include_send_day: bool = False,
    start_offset_days=None,
) -> tuple[date, date]:
    """
    按「发送日」划分的 7 天排班区间（含首尾共 7 天）。
    include_send_day=False：发送日次日 至 下周同一发送日（例：周五发 → 周六至下周五）。
    include_send_day=True：发送日当天 至 下周同一发送日前一天（例：周五发 → 周五至下周四）。
    """
    send_weekday = _normalize_email_send_weekday(send_weekday)
    offset = _normalize_email_start_offset_days(start_offset_days, include_send_day)
    start_weekday = (send_weekday + offset) % 7
    start = anchor - timedelta(days=(anchor.weekday() - start_weekday) % 7)
    return start, start + timedelta(days=6)


def _get_shift_email_send_weekday(department: str) -> int:
    _ensure_tables()
    if not department:
        return DEFAULT_EMAIL_SEND_WEEKDAY
    rows = db.execute_query(
        "SELECT email_send_weekday FROM shift_config WHERE department = %s LIMIT 1",
        (department,),
    )
    if rows:
        return _normalize_email_send_weekday(rows[0].get("email_send_weekday"))
    return DEFAULT_EMAIL_SEND_WEEKDAY


def _get_shift_email_include_send_day(department: str) -> bool:
    _ensure_tables()
    if not department:
        return False
    rows = db.execute_query(
        "SELECT email_include_send_day FROM shift_config WHERE department = %s LIMIT 1",
        (department,),
    )
    if rows:
        return _normalize_email_include_send_day(rows[0].get("email_include_send_day"))
    return False


def _get_shift_email_start_offset_days(department: str) -> int:
    _ensure_tables()
    if not department:
        return DEFAULT_EMAIL_START_OFFSET_DAYS
    rows = db.execute_query(
        "SELECT email_start_offset_days, email_include_send_day FROM shift_config WHERE department = %s LIMIT 1",
        (department,),
    )
    if rows:
        return _normalize_email_start_offset_days(
            rows[0].get("email_start_offset_days"),
            rows[0].get("email_include_send_day"),
        )
    return DEFAULT_EMAIL_START_OFFSET_DAYS


def _shift_schedule_target_week_start(department: str, now: Optional[datetime] = None) -> date:
    """发送当日 17:00 邮件对应的排班周起始日。"""
    current = now or datetime.now()
    start_offset_days = _get_shift_email_start_offset_days(department)
    include_send_day = start_offset_days == 0
    anchor = current.date() + timedelta(days=start_offset_days)
    week_start, _ = _week_range_for_send_day(
        anchor,
        _get_shift_email_send_weekday(department),
        include_send_day,
        start_offset_days,
    )
    return week_start


def _shift_coverage_check_range(
    today: date,
    send_weekday: int = DEFAULT_EMAIL_SEND_WEEKDAY,
    include_send_day: bool = False,
    start_offset_days=None,
) -> Tuple[date, date]:
    """日常排班缺口检测区间：与本科室邮件排班周期一致（含首尾，共 7 天）。"""
    offset = _normalize_email_start_offset_days(start_offset_days, include_send_day)
    return _week_range_for_send_day(today + timedelta(days=offset), send_weekday, include_send_day, offset)


def _prev_month_same_day(d: date) -> date:
    """目标日在上月同一天（上月无该日则取上月最后一天）"""
    if d.month == 1:
        y, m = d.year - 1, 12
    else:
        y, m = d.year, d.month - 1
    last = calendar.monthrange(y, m)[1]
    day = min(d.day, last)
    return date(y, m, day)


def _week_saturday_range(anchor: date) -> tuple[date, date]:
    """返回 anchor 所在排班周：本周六到下周五（发送日=周五时的邮件周期）。"""
    return _week_range_for_send_day(anchor, DEFAULT_EMAIL_SEND_WEEKDAY)


def _holiday_map_for_years(years: Set[int]) -> dict:
    holiday_by_date = {}
    for y in years:
        for r in load_holidays_for_year(str(y)):
            nk = _normalize_date_key(r.get("date"))
            if nk:
                holiday_by_date[nk] = {
                    "type": (r.get("type") or "").strip(),
                    "festival": (r.get("festival") or "").strip() if r.get("festival") is not None else "",
                }
    return holiday_by_date


def _normalize_date_key(val) -> str:
    """将 holiday 表 date 字段规范为 YYYY-MM-DD，便于与排班日期对齐"""
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip().replace("/", "-").split()[0]
    if not s:
        return ""
    parts = [p for p in s.split("-") if p != ""]
    if len(parts) >= 3:
        try:
            y, mo, da = int(parts[0]), int(parts[1]), int(parts[2])
            return date(y, mo, da).strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            pass
    return s[:10]


def _holiday_header_mark(holiday_type: str, festival: str) -> str:
    """表头短标记：调休上班 / 节假日"""
    ht = (holiday_type or "").strip()
    fest = (festival or "").strip()
    if "班" in ht:
        return "班"
    if fest:
        return fest[:4] if len(fest) > 4 else fest
    if "假" in ht or "休" in ht:
        return "休"
    return ht[:3] if ht else ""


def _is_workday(d: date, holidays: dict) -> bool:
    """判断是否工作日（考虑假期调休）"""
    ds = d.strftime("%Y-%m-%d")
    if ds in holidays:
        ht = holidays[ds]
        if "班" in ht:
            return True
        if "假" in ht or "休" in ht:
            return False
    return d.weekday() < 5


def _is_manager_of_dept(name: str, department: str) -> bool:
    """判断用户是否为指定科室的管理人员（组长/主任/副主任）"""
    if not name or not department:
        return False
    rows = db.execute_query(
        "SELECT jb, lsys FROM yggl WHERE TRIM(name) = %s AND COALESCE(zaizhi,0) = 0 LIMIT 1",
        (name.strip(),),
    )
    if not rows:
        return False
    jb = (rows[0].get("jb") or "").strip()
    dept = (rows[0].get("lsys") or "").strip()
    return ("组长" in jb or "主任" in jb) and dept == department


def _get_user_shift_manager_dept(name: str) -> Optional[str]:
    """返回排班管理人员所在科室；普通员工返回 None。"""
    if not name:
        return None
    rows = db.execute_query(
        "SELECT jb, lsys FROM yggl WHERE TRIM(name) = %s AND COALESCE(zaizhi,0) = 0 LIMIT 1",
        (name.strip(),),
    )
    if not rows:
        return None
    jb = (rows[0].get("jb") or "").strip()
    dept = (rows[0].get("lsys") or "").strip()
    if dept and ("组长" in jb or "主任" in jb):
        return dept
    return None


def _is_dept_member(name: str, department: str) -> bool:
    """是否为该科室在职员工（含管理人员）"""
    if not name or not department:
        return False
    rows = db.execute_query(
        "SELECT 1 FROM yggl WHERE TRIM(name) = %s AND TRIM(lsys) = %s AND COALESCE(zaizhi,0) = 0 LIMIT 1",
        (name.strip(), department.strip()),
    )
    return bool(rows)


# ==================== 日期开放 / 锁定 ====================

class SetDayLocksRequest(BaseModel):
    department: str
    dates: List[str]
    is_open: bool
    current_user: str = ""


@router.post("/day-locks")
def set_day_locks(req: SetDayLocksRequest):
    """管理人员设置日期开放/锁定"""
    _ensure_tables()
    if not _is_manager_of_dept(req.current_user, req.department):
        raise HTTPException(status_code=403, detail="仅本科室管理人员可操作排班开放权限")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cnt = 0
    for date_str in req.dates:
        d = _parse_iso_date(date_str)
        if not d:
            continue
        ds = d.strftime("%Y-%m-%d")
        if req.is_open:
            db.execute_update(
                "INSERT INTO shift_day_lock (department, lock_date, is_open, opened_by, opened_at) "
                "VALUES (%s, %s, 1, %s, %s) "
                "ON DUPLICATE KEY UPDATE is_open = 1, opened_by = %s, opened_at = %s",
                (req.department, ds, req.current_user, now, req.current_user, now),
            )
        else:
            db.execute_update(
                "UPDATE shift_day_lock SET is_open = 0, opened_by = %s, opened_at = %s "
                "WHERE department = %s AND lock_date = %s",
                (req.current_user, now, req.department, ds),
            )
        cnt += 1
    action = "解锁" if req.is_open else "锁定"
    return {"success": True, "message": f"已{action} {cnt} 天排班权限"}


class SetDayNoDutyRequest(BaseModel):
    department: str
    dates: List[str]
    no_duty: bool = True
    current_user: str = ""


@router.post("/day-noduty")
def set_day_noduty(req: SetDayNoDutyRequest):
    """管理人员标记/取消「当日不设置值班」（该日不参与缺排检测与排班邮件拦截）。"""
    _ensure_tables()
    if not _is_manager_of_dept(req.current_user, req.department):
        raise HTTPException(status_code=403, detail="仅本科室管理人员可设置当日不值班")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cnt = 0
    for date_str in req.dates:
        d = _parse_iso_date(date_str)
        if not d:
            continue
        ds = d.strftime("%Y-%m-%d")
        if req.no_duty:
            db.execute_update(
                "INSERT INTO shift_day_noduty (department, noduty_date, updated_by, updated_at) "
                "VALUES (%s, %s, %s, %s) "
                "ON DUPLICATE KEY UPDATE updated_by = %s, updated_at = %s",
                (req.department, ds, req.current_user, now, req.current_user, now),
            )
        else:
            db.execute_update(
                "DELETE FROM shift_day_noduty WHERE department = %s AND noduty_date = %s",
                (req.department, ds),
            )
        cnt += 1
    action = "标记不设置值班" if req.no_duty else "取消不设置值班"
    return {"success": True, "message": f"已{action} {cnt} 天"}


# ==================== 配置 ====================

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
SHIFT_EMAIL_RECIPIENT_UNITS = {
    "水电分厂",
    "汽发分厂",
    "线圈分厂",
    "冲剪分厂",
    "冷作分厂",
    "成品分厂",
    "大电机研究所",
    "金工分厂",
    "其他",
}


class ShiftEmailRecipient(BaseModel):
    name: str = ""
    email: str = ""
    unit: str = "其他"


class ShiftConfigRequest(BaseModel):
    department: str
    workday_day: int = 2
    workday_night: int = 2
    weekend_day: int = 2
    weekend_night: int = 2
    email_recipients: List[ShiftEmailRecipient] = Field(default_factory=list)
    email_send_weekday: int = DEFAULT_EMAIL_SEND_WEEKDAY
    current_user: str = ""


def _normalize_shift_email_recipients(recipients) -> List[dict]:
    normalized = []
    seen = set()
    for item in recipients or []:
        raw_name = item.get("name", "") if isinstance(item, dict) else getattr(item, "name", "")
        raw_email = item.get("email", "") if isinstance(item, dict) else getattr(item, "email", "")
        raw_unit = item.get("unit", "") if isinstance(item, dict) else getattr(item, "unit", "")
        name = (raw_name or "").strip()
        email = (raw_email or "").strip()
        unit = (raw_unit or "").strip() or "其他"
        if not name and not email:
            continue
        if not name or not email:
            raise HTTPException(status_code=400, detail="排班表收件人姓名和邮箱均不能为空")
        if not EMAIL_RE.match(email):
            raise HTTPException(status_code=400, detail=f"排班表收件人邮箱格式不正确：{email}")
        if unit not in SHIFT_EMAIL_RECIPIENT_UNITS:
            unit = "其他"
        key = email.lower()
        if key in seen:
            continue
        seen.add(key)
        normalized.append({"name": name, "email": email, "unit": unit})
    return normalized


def _parse_shift_email_recipients(raw) -> List[dict]:
    if not raw:
        return []
    if isinstance(raw, (bytes, bytearray)):
        raw = raw.decode("utf-8", errors="ignore")
    try:
        data = json.loads(raw) if isinstance(raw, str) else raw
    except (TypeError, json.JSONDecodeError):
        return []
    try:
        return _normalize_shift_email_recipients(data)
    except HTTPException:
        return []


@router.get("/config")
def get_shift_config(department: str = Query(...)):
    """获取科室排班配置"""
    _ensure_tables()
    rows = db.execute_query(
        "SELECT workday_day, workday_night, weekend_day, weekend_night, email_recipients, "
        "email_send_weekday, email_include_send_day, email_start_offset_days "
        "FROM shift_config WHERE department = %s LIMIT 1",
        (department,),
    )
    if rows:
        data = dict(rows[0])
        data["email_recipients"] = _parse_shift_email_recipients(data.get("email_recipients"))
        data["email_send_weekday"] = _normalize_email_send_weekday(data.get("email_send_weekday"))
        data["email_include_send_day"] = _normalize_email_include_send_day(data.get("email_include_send_day"))
        data["email_start_offset_days"] = _normalize_email_start_offset_days(
            data.get("email_start_offset_days"),
            data.get("email_include_send_day"),
        )
        data["email_feature_enabled"] = _is_shift_email_feature_enabled(department)
        return {"success": True, "data": data}
    return {
        "success": True,
        "data": {
            "workday_day": 2,
            "workday_night": 2,
            "weekend_day": 2,
            "weekend_night": 2,
            "email_recipients": [],
            "email_send_weekday": DEFAULT_EMAIL_SEND_WEEKDAY,
            "email_include_send_day": False,
            "email_start_offset_days": DEFAULT_EMAIL_START_OFFSET_DAYS,
            "email_feature_enabled": _is_shift_email_feature_enabled(department),
        },
    }


@router.get("/coverage-gap")
def get_shift_coverage_gap(current_user: str = Query(..., description="当前登录人姓名")):
    """首页待办：检查当前用户所在科室在「本周六—下周五」是否存在日常排班人数缺口。

    只检测普通上班日与公休日；命中系统节假日配置的日期不纳入检测。
    """
    _ensure_tables()
    dept = _get_user_shift_manager_dept(current_user)
    if not dept:
        return {"success": True, "hasPending": False, "department": "", "issues": [], "totalIssues": 0}

    today = date.today()
    send_wd = _get_shift_email_send_weekday(dept)
    start_offset = _get_shift_email_start_offset_days(dept)
    include_send = start_offset == 0
    start_day, end_day = _shift_coverage_check_range(today, send_wd, include_send, start_offset)
    dates = _daterange(start_day, end_day)
    years_set = {d.year for d in dates}
    holiday_by_date = _holiday_map_for_years(years_set)

    config_rows = db.execute_query(
        "SELECT workday_day, workday_night, weekend_day, weekend_night "
        "FROM shift_config WHERE department = %s LIMIT 1",
        (dept,),
    )
    cfg = config_rows[0] if config_rows else {
        "workday_day": 2,
        "workday_night": 2,
        "weekend_day": 2,
        "weekend_night": 2,
    }

    employees = _get_dept_employees(dept)
    ds_lo = start_day.strftime("%Y-%m-%d")
    ds_hi = end_day.strftime("%Y-%m-%d")
    no_duty_dates = _load_no_duty_dates(dept, ds_lo, ds_hi)
    counts = {}
    if employees:
        ph = ",".join(["%s"] * len(employees))
        rows = db.execute_query(
            f"SELECT shift_date, shift_type, COUNT(*) AS cnt FROM shift_schedule "
            f"WHERE department = %s AND shift_date >= %s AND shift_date <= %s "
            f"AND employee_name IN ({ph}) AND shift_type IN ('白班', '夜班', '白+夜') "
            f"GROUP BY shift_date, shift_type",
            (dept, ds_lo, ds_hi) + tuple(employees),
        )
        for r in rows or []:
            sd = r.get("shift_date")
            ds = sd.strftime("%Y-%m-%d") if hasattr(sd, "strftime") else (str(sd)[:10] if sd else "")
            st = (r.get("shift_type") or "").strip()
            cnt = int(r.get("cnt") or 0)
            if not ds:
                continue
            day_counts = counts.setdefault(ds, {"day": 0, "night": 0})
            if st == "白班":
                day_counts["day"] += cnt
            elif st == "夜班":
                day_counts["night"] += cnt
            elif st == "白+夜":
                day_counts["day"] += cnt
                day_counts["night"] += cnt

    issues = []
    for d in dates:
        ds = d.strftime("%Y-%m-%d")
        if ds in holiday_by_date or ds in no_duty_dates:
            continue
        is_workday = d.weekday() < 5
        req_day = max(0, int(cfg.get("workday_day" if is_workday else "weekend_day") or 0))
        req_night = max(0, int(cfg.get("workday_night" if is_workday else "weekend_night") or 0))
        actual = counts.get(ds, {"day": 0, "night": 0})
        day_missing = max(0, req_day - int(actual.get("day") or 0))
        night_missing = max(0, req_night - int(actual.get("night") or 0))
        if day_missing or night_missing:
            issues.append({
                "date": ds,
                "weekday": ["一", "二", "三", "四", "五", "六", "日"][d.weekday()],
                "dayType": "普通上班日" if is_workday else "公休日",
                "holidayMark": "",
                "requiredDay": req_day,
                "requiredNight": req_night,
                "actualDay": int(actual.get("day") or 0),
                "actualNight": int(actual.get("night") or 0),
                "missingDay": day_missing,
                "missingNight": night_missing,
            })

    sample = issues[:5]
    sample_parts = []
    for i in sample:
        missing_parts = []
        if i["missingDay"]:
            missing_parts.append(f"白班{i['missingDay']}人")
        if i["missingNight"]:
            missing_parts.append(f"夜班{i['missingNight']}人")
        sample_parts.append(f"{i['date'][5:]}缺{'、'.join(missing_parts)}")
    sample_text = "；".join(sample_parts)
    if len(issues) > len(sample):
        sample_text += f"；另有{len(issues) - len(sample)}天"

    return {
        "success": True,
        "hasPending": bool(issues),
        "department": dept,
        "startDate": start_day.strftime("%Y-%m-%d"),
        "endDate": end_day.strftime("%Y-%m-%d"),
        "totalIssues": len(issues),
        "summary": sample_text,
        "issues": issues,
    }


@router.post("/config")
def save_shift_config(req: ShiftConfigRequest):
    """保存科室排班人数规则与排班表收件人；自动发送时间仍由系统管理员配置。"""
    _ensure_tables()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    email_recipients = _normalize_shift_email_recipients(req.email_recipients)
    email_recipients_json = json.dumps(email_recipients, ensure_ascii=False)
    existing = db.execute_query(
        "SELECT id FROM shift_config WHERE department = %s LIMIT 1", (req.department,)
    )
    if existing:
        db.execute_update(
            "UPDATE shift_config SET workday_day=%s, workday_night=%s, weekend_day=%s, weekend_night=%s, "
            "email_recipients=%s, updated_by=%s, updated_at=%s WHERE department=%s",
            (req.workday_day, req.workday_night, req.weekend_day, req.weekend_night,
             email_recipients_json, req.current_user, now, req.department),
        )
    else:
        db.execute_update(
            "INSERT INTO shift_config (department, workday_day, workday_night, weekend_day, weekend_night, "
            "email_recipients, email_send_weekday, updated_by, updated_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (req.department, req.workday_day, req.workday_night, req.weekend_day, req.weekend_night,
             email_recipients_json, DEFAULT_EMAIL_SEND_WEEKDAY, req.current_user, now),
        )
    return {"success": True, "message": "配置已保存"}


# ==================== 排班数据 ====================

@router.get("/schedule")
def get_schedule(
    department: str = Query(...),
    start_date: Optional[str] = Query(None, description="起始日 YYYY-MM-DD，与 end_date 成对使用"),
    end_date: Optional[str] = Query(None, description="结束日 YYYY-MM-DD，含当日"),
    year: Optional[int] = Query(None, description="与 month 成对，整月模式（兼容旧接口）"),
    month: Optional[int] = Query(None, ge=1, le=12),
):
    """获取科室排班：按日期区间（推荐）或整月。"""
    _ensure_tables()
    employees = _get_dept_employees(department)
    if not employees:
        return {"success": True, "employees": [], "schedule": {}, "dates": [], "dayPlans": {}}

    d0 = _parse_iso_date(start_date) if start_date else None
    d1 = _parse_iso_date(end_date) if end_date else None
    if d0 and d1:
        if d1 < d0:
            raise HTTPException(status_code=400, detail="结束日期不能早于起始日期")
        if (d1 - d0).days > 62:
            raise HTTPException(status_code=400, detail="查询区间最多 63 天")
        dates = _daterange(d0, d1)
        years_set = {d.year for d in dates}
    elif year is not None and month is not None:
        dates = _month_dates(year, month)
        years_set = {year}
    else:
        raise HTTPException(status_code=400, detail="请提供 start_date 与 end_date，或 year 与 month")

    holiday_by_date = _holiday_map_for_years(years_set)
    holidays = {k: v["type"] for k, v in holiday_by_date.items()}
    date_info = []
    for d in dates:
        ds = d.strftime("%Y-%m-%d")
        wd = _is_workday(d, holidays)
        h = holiday_by_date.get(ds)
        ht = h["type"] if h else ""
        fest = h["festival"] if h else ""
        mark = _holiday_header_mark(ht, fest) if h else ""
        date_info.append({
            "date": ds,
            "weekday": d.weekday(),
            "isWorkday": wd,
            "label": ["一", "二", "三", "四", "五", "六", "日"][d.weekday()],
            "holidayType": ht,
            "holidayFestival": fest,
            "holidayMark": mark,
        })

    ph = ",".join(["%s"] * len(employees))
    ds_lo = dates[0].strftime("%Y-%m-%d")
    ds_hi = dates[-1].strftime("%Y-%m-%d")
    rows = db.execute_query(
        f"SELECT employee_name, shift_date, shift_type, shift_location FROM shift_schedule "
        f"WHERE department = %s AND shift_date >= %s AND shift_date <= %s AND employee_name IN ({ph})",
        (department, ds_lo, ds_hi) + tuple(employees),
    )
    schedule = {}
    locations = {}
    for r in rows:
        name = (r.get("employee_name") or "").strip()
        sd = r.get("shift_date")
        if hasattr(sd, "strftime"):
            sd = sd.strftime("%Y-%m-%d")
        else:
            sd = str(sd)[:10] if sd else ""
        st = (r.get("shift_type") or "").strip()
        sl = (r.get("shift_location") or "").strip()
        if name and sd:
            schedule.setdefault(name, {})[sd] = st
            if sl:
                locations.setdefault(name, {})[sd] = sl

    day_plans = {}
    try:
        plan_rows = db.execute_query(
            "SELECT plan_date, content FROM shift_day_plan "
            "WHERE department = %s AND plan_date >= %s AND plan_date <= %s",
            (department, ds_lo, ds_hi),
        )
        for pr in plan_rows or []:
            pd = pr.get("plan_date")
            if hasattr(pd, "strftime"):
                pds = pd.strftime("%Y-%m-%d")
            else:
                pds = str(pd)[:10] if pd else ""
            if pds:
                day_plans[pds] = (pr.get("content") or "").strip()
    except Exception as e:
        logger.warning("读取 shift_day_plan 失败: %s", e)

    open_dates = []
    try:
        lock_rows = db.execute_query(
            "SELECT lock_date FROM shift_day_lock "
            "WHERE department = %s AND lock_date >= %s AND lock_date <= %s AND is_open = 1",
            (department, ds_lo, ds_hi),
        )
        for lr in lock_rows or []:
            ld = lr.get("lock_date")
            if hasattr(ld, "strftime"):
                open_dates.append(ld.strftime("%Y-%m-%d"))
            else:
                lds = str(ld)[:10] if ld else ""
                if lds:
                    open_dates.append(lds)
    except Exception as e:
        logger.warning("读取 shift_day_lock 失败: %s", e)

    no_duty_dates = sorted(_load_no_duty_dates(department, ds_lo, ds_hi))
    business_trips = _get_dept_business_trips(department, dates[0], dates[-1])

    return {
        "success": True,
        "employees": employees,
        "schedule": schedule,
        "locations": locations,
        "dates": date_info,
        "dayPlans": day_plans,
        "openDates": open_dates,
        "noDutyDates": no_duty_dates,
        "businessTrips": business_trips,
    }


class SaveScheduleRequest(BaseModel):
    department: str
    year: int = 0
    month: int = 0
    schedule: dict  # { "张三": { "2026-03-01": "白班", ... }, ... }
    locations: dict = {}  # { "张三": { "2026-03-01": "准备组", ... }, ... }
    current_user: str = ""


@router.post("/schedule")
def save_schedule(req: SaveScheduleRequest):
    """保存排班数据（每条记录的 year/month 按 shift_date 解析，支持跨月区间）"""
    _ensure_tables()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    is_mgr = _is_manager_of_dept(req.current_user, req.department)
    today_d = date.today()
    editable_from_d = today_d - timedelta(days=1)
    caller = (req.current_user or "").strip()
    if not is_mgr and not _is_dept_member(req.current_user, req.department):
        raise HTTPException(status_code=403, detail="仅本科室成员可保存排班")

    open_set = None
    if not is_mgr:
        all_dates = sorted({str(ds)[:10] for _, dm in req.schedule.items() for ds in dm.keys()})
        if all_dates:
            rows = db.execute_query(
                "SELECT lock_date FROM shift_day_lock "
                "WHERE department = %s AND is_open = 1 AND lock_date >= %s AND lock_date <= %s",
                (req.department, all_dates[0], all_dates[-1]),
            )
            open_set = set()
            for r in rows:
                ld = r.get("lock_date")
                open_set.add(ld.strftime("%Y-%m-%d") if hasattr(ld, "strftime") else str(ld)[:10])
        else:
            open_set = set()

    loc_map = req.locations or {}
    params_list = []
    for emp_name, day_map in req.schedule.items():
        en = (emp_name or "").strip()
        if not is_mgr:
            if not caller or en != caller:
                continue
        for date_str, shift_type in day_map.items():
            d = _parse_iso_date(str(date_str))
            if not d:
                continue
            ds = d.strftime("%Y-%m-%d")
            if open_set is not None and ds not in open_set:
                continue
            if d < editable_from_d:
                continue
            y, m = d.year, d.month
            loc = (loc_map.get(en, {}).get(ds) or "").strip()
            params_list.append((
                req.department, en, ds, shift_type, loc,
                y, m, req.current_user, now,
                shift_type, loc, req.current_user, now,
            ))
    if not params_list:
        return {"success": True, "message": "无排班数据"}
    sql = (
        "INSERT INTO shift_schedule (department, employee_name, shift_date, shift_type, shift_location, year, month, updated_by, updated_at) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) "
        "ON DUPLICATE KEY UPDATE shift_type = %s, shift_location = %s, updated_by = %s, updated_at = %s"
    )
    n = db.execute_many(sql, params_list)
    if n < 0:
        raise HTTPException(status_code=500, detail="保存排班失败，请稍后重试")
    return {"success": True, "message": f"已保存 {len(params_list)} 条排班记录"}


MAX_DAY_PLAN_LEN = 2000


class SaveDayPlansRequest(BaseModel):
    department: str
    current_user: str = ""
    plans: dict  # {"2026-03-01": "值班工作计划…"}


@router.post("/day-plans")
def save_day_plans(req: SaveDayPlansRequest):
    """按科室、按日保存值班工作计划（协同编辑，与排班表头下计划行对应）"""
    _ensure_tables()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    is_mgr = _is_manager_of_dept(req.current_user, req.department)
    editable_from_d = date.today() - timedelta(days=1)
    open_set = None
    if not is_mgr:
        if not _is_dept_member(req.current_user, req.department):
            raise HTTPException(status_code=403, detail="仅本科室成员可保存工作计划")
        parsed = []
        for date_str in (req.plans or {}).keys():
            d = _parse_iso_date(str(date_str))
            if d:
                parsed.append(d.strftime("%Y-%m-%d"))
        parsed.sort()
        open_set = set()
        if parsed:
            rows = db.execute_query(
                "SELECT lock_date FROM shift_day_lock "
                "WHERE department = %s AND is_open = 1 AND lock_date >= %s AND lock_date <= %s",
                (req.department, parsed[0], parsed[-1]),
            )
            for r in rows or []:
                ld = r.get("lock_date")
                open_set.add(ld.strftime("%Y-%m-%d") if hasattr(ld, "strftime") else str(ld)[:10])

    sql_ins = (
        "INSERT INTO shift_day_plan (department, plan_date, content, updated_by, updated_at) "
        "VALUES (%s, %s, %s, %s, %s) ON DUPLICATE KEY UPDATE "
        "content=%s, updated_by=%s, updated_at=%s"
    )
    for date_str, raw in (req.plans or {}).items():
        d = _parse_iso_date(str(date_str))
        if not d:
            continue
        if d < editable_from_d:
            continue
        ds = d.strftime("%Y-%m-%d")
        if open_set is not None and ds not in open_set:
            continue
        text = (str(raw) if raw is not None else "").strip()
        if len(text) > MAX_DAY_PLAN_LEN:
            text = text[:MAX_DAY_PLAN_LEN]
        if not text:
            db.execute_update(
                "DELETE FROM shift_day_plan WHERE department=%s AND plan_date=%s",
                (req.department, ds),
            )
        else:
            db.execute_update(
                sql_ins,
                (req.department, ds, text, req.current_user, now, text, req.current_user, now),
            )
    return {"success": True, "message": "工作计划已保存"}


# ==================== 自动排班 ====================

class AutoScheduleRequest(BaseModel):
    department: str
    year: Optional[int] = None
    month: Optional[int] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    current_user: str = ""
    workday_day: Optional[int] = None
    workday_night: Optional[int] = None
    weekend_day: Optional[int] = None
    weekend_night: Optional[int] = None


@router.post("/auto-schedule")
def auto_schedule(req: AutoScheduleRequest):
    """
    自动排班：严格按配置人数安排，配置几人就排几人，其余留空。
    工作日：workday_day 人白班，workday_night 人夜班。
    周末：weekend_day 人白班，weekend_night 人夜班。
    今天之前的日期不写入、不覆盖；轮转仍按整段日期推进，使「今天起」的排班与连续排满时一致。
    """
    _ensure_tables()
    employees = _get_dept_employees(req.department)
    if not employees:
        return {"success": False, "message": "该科室无在职员工"}
    cfg_rows = db.execute_query(
        "SELECT workday_day, workday_night, weekend_day, weekend_night FROM shift_config WHERE department = %s LIMIT 1",
        (req.department,),
    )
    cfg = cfg_rows[0] if cfg_rows else {}

    def _pick(req_val, cfg_key, fallback=2):
        return max(0, int(req_val) if req_val is not None else int(cfg.get(cfg_key) or fallback))

    workday_day = _pick(req.workday_day, "workday_day")
    workday_night = _pick(req.workday_night, "workday_night")
    weekend_day = _pick(req.weekend_day, "weekend_day")
    weekend_night = _pick(req.weekend_night, "weekend_night")

    d0 = _parse_iso_date(req.start_date) if req.start_date else None
    d1 = _parse_iso_date(req.end_date) if req.end_date else None
    if d0 and d1:
        if d1 < d0:
            raise HTTPException(status_code=400, detail="结束日期不能早于起始日期")
        if (d1 - d0).days > 62:
            raise HTTPException(status_code=400, detail="自动排班区间最多 63 天")
        dates = _daterange(d0, d1)
        years_set = {d.year for d in dates}
    elif req.year is not None and req.month is not None:
        dates = _month_dates(req.year, req.month)
        years_set = {req.year}
    else:
        raise HTTPException(status_code=400, detail="请提供 start_date 与 end_date，或 year 与 month")

    holidays = {}
    for y in years_set:
        for r in load_holidays_for_year(str(y)):
            nk = _normalize_date_key(r.get("date"))
            if nk:
                holidays[nk] = (r.get("type") or "").strip()
    n = len(employees)
    today_d = date.today()

    schedule = {emp: {} for emp in employees}
    wd_day_idx = 0
    wd_night_idx = 0
    we_day_idx = 0
    we_night_idx = 0

    for d in dates:
        ds = d.strftime("%Y-%m-%d")
        is_wd = _is_workday(d, holidays)
        if is_wd:
            d_count = min(workday_day, n)
            n_count = min(workday_night, n)
        else:
            d_count = min(weekend_day, n)
            n_count = min(weekend_night, n)

        day_set = set()
        night_set = set()
        day_idx = wd_day_idx if is_wd else we_day_idx
        night_idx = wd_night_idx if is_wd else we_night_idx

        if d >= today_d:
            for i in range(d_count):
                emp = employees[(day_idx + i) % n]
                schedule[emp][ds] = "白班"
                day_set.add(emp)

            for i in range(n_count):
                emp = employees[(night_idx + i) % n]
                if emp not in day_set:
                    schedule[emp][ds] = "夜班"
                    night_set.add(emp)
                else:
                    for j in range(1, n):
                        candidate = employees[(night_idx + i + j) % n]
                        if candidate not in day_set and candidate not in night_set:
                            schedule[candidate][ds] = "夜班"
                            night_set.add(candidate)
                            break

        if is_wd:
            wd_day_idx = (wd_day_idx + d_count) % n if d_count else wd_day_idx
            wd_night_idx = (wd_night_idx + n_count) % n if n_count else wd_night_idx
        else:
            we_day_idx = (we_day_idx + d_count) % n if d_count else we_day_idx
            we_night_idx = (we_night_idx + n_count) % n if n_count else we_night_idx

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    params_list = []
    for emp_name, day_map in schedule.items():
        for date_str, shift_type in day_map.items():
            if not shift_type:
                continue
            d = _parse_iso_date(date_str)
            if not d:
                continue
            params_list.append((
                req.department, emp_name, date_str, shift_type,
                d.year, d.month, req.current_user, now,
                shift_type, req.current_user, now,
            ))
    ds_first = dates[0].strftime("%Y-%m-%d")
    ds_last = dates[-1].strftime("%Y-%m-%d")
    today_str = today_d.strftime("%Y-%m-%d")
    del_from = ds_first if ds_first > today_str else today_str
    if del_from <= ds_last:
        db.execute_update(
            "DELETE FROM shift_schedule WHERE department = %s AND shift_date >= %s AND shift_date <= %s",
            (req.department, del_from, ds_last),
        )
    if params_list:
        sql = (
            "INSERT INTO shift_schedule (department, employee_name, shift_date, shift_type, year, month, updated_by, updated_at) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s) "
            "ON DUPLICATE KEY UPDATE shift_type = %s, updated_by = %s, updated_at = %s"
        )
        rc = db.execute_many(sql, params_list)
        if rc < 0:
            raise HTTPException(status_code=500, detail="自动排班写入失败，请稍后重试")
    sched_days = sum(1 for d in dates if d >= today_d)
    return {
        "success": True,
        "message": f"已自动排班：{len(employees)} 人，本段共 {len(dates)} 天（其中 {sched_days} 天为今天及之后已写入）",
        "schedule": schedule,
    }


# ==================== 复制上月排班 ====================

class CopyScheduleRequest(BaseModel):
    department: str
    year: Optional[int] = None
    month: Optional[int] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    current_user: str = ""


@router.post("/copy-last-month")
def copy_last_month(req: CopyScheduleRequest):
    """复制上月对应日期的排班：按日期区间（当前屏）或整月。"""
    _ensure_tables()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    d0 = _parse_iso_date(req.start_date) if req.start_date else None
    d1 = _parse_iso_date(req.end_date) if req.end_date else None

    if d0 and d1:
        if d1 < d0:
            raise HTTPException(status_code=400, detail="结束日期不能早于起始日期")
        dst_dates = _daterange(d0, d1)
        src_dates_list = [_prev_month_same_day(d) for d in dst_dates]
        src_set = sorted({s.strftime("%Y-%m-%d") for s in src_dates_list})
        ph_s = ",".join(["%s"] * len(src_set))
        rows = db.execute_query(
            f"SELECT employee_name, shift_date, shift_type FROM shift_schedule "
            f"WHERE department = %s AND shift_date IN ({ph_s})",
            (req.department,) + tuple(src_set),
        )
        if not rows:
            return {"success": False, "message": "对应上月日期无排班记录"}
        src_map = {}
        for r in rows:
            name = (r.get("employee_name") or "").strip()
            sd = r.get("shift_date")
            if hasattr(sd, "strftime"):
                sd = sd.strftime("%Y-%m-%d")
            else:
                sd = str(sd)[:10]
            src_map.setdefault(name, {})[sd] = (r.get("shift_type") or "").strip()

        params_list = []
        for name, day_map in src_map.items():
            for dst_d, src_d in zip(dst_dates, src_dates_list):
                src_ds = src_d.strftime("%Y-%m-%d")
                st = day_map.get(src_ds, "")
                if not st:
                    continue
                dst_ds = dst_d.strftime("%Y-%m-%d")
                params_list.append((
                    req.department, name, dst_ds, st,
                    dst_d.year, dst_d.month, req.current_user, now,
                    st, req.current_user, now,
                ))
    elif req.year is not None and req.month is not None:
        if req.month == 1:
            src_year, src_month = req.year - 1, 12
        else:
            src_year, src_month = req.year, req.month - 1

        rows = db.execute_query(
            "SELECT employee_name, shift_date, shift_type FROM shift_schedule "
            "WHERE department = %s AND year = %s AND month = %s",
            (req.department, src_year, src_month),
        )
        if not rows:
            return {"success": False, "message": "上月无排班记录"}

        src_dates = _month_dates(src_year, src_month)
        dst_dates = _month_dates(req.year, req.month)
        day_count = min(len(src_dates), len(dst_dates))

        src_map = {}
        for r in rows:
            name = (r.get("employee_name") or "").strip()
            sd = r.get("shift_date")
            if hasattr(sd, "strftime"):
                sd = sd.strftime("%Y-%m-%d")
            else:
                sd = str(sd)[:10]
            src_map.setdefault(name, {})[sd] = (r.get("shift_type") or "").strip()

        params_list = []
        for name, day_map in src_map.items():
            for i, src_d in enumerate(src_dates[:day_count]):
                src_ds = src_d.strftime("%Y-%m-%d")
                st = day_map.get(src_ds, "")
                if not st:
                    continue
                dst_d = dst_dates[i]
                dst_ds = dst_d.strftime("%Y-%m-%d")
                params_list.append((
                    req.department, name, dst_ds, st,
                    dst_d.year, dst_d.month, req.current_user, now,
                    st, req.current_user, now,
                ))
    else:
        raise HTTPException(status_code=400, detail="请提供 start_date 与 end_date，或 year 与 month")

    if not params_list:
        return {"success": False, "message": "上月排班记录为空"}
    sql = (
        "INSERT INTO shift_schedule (department, employee_name, shift_date, shift_type, year, month, updated_by, updated_at) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s) "
        "ON DUPLICATE KEY UPDATE shift_type = %s, updated_by = %s, updated_at = %s"
    )
    n = db.execute_many(sql, params_list)
    if n < 0:
        raise HTTPException(status_code=500, detail="复制上月排班失败，请稍后重试")
    return {"success": True, "message": f"已复制 {len(params_list)} 条排班记录"}


# ==================== 清空当月排班 ====================

class ClearScheduleRequest(BaseModel):
    department: str
    year: Optional[int] = None
    month: Optional[int] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None


@router.post("/clear-schedule")
def clear_schedule(req: ClearScheduleRequest):
    """清空该科室排班：按日期区间或整月"""
    _ensure_tables()
    d0 = _parse_iso_date(req.start_date) if req.start_date else None
    d1 = _parse_iso_date(req.end_date) if req.end_date else None
    if d0 and d1:
        if d1 < d0:
            raise HTTPException(status_code=400, detail="结束日期不能早于起始日期")
        n = db.execute_update(
            "DELETE FROM shift_schedule WHERE department = %s AND shift_date >= %s AND shift_date <= %s",
            (req.department, d0.strftime("%Y-%m-%d"), d1.strftime("%Y-%m-%d")),
        )
    elif req.year is not None and req.month is not None:
        n = db.execute_update(
            "DELETE FROM shift_schedule WHERE department = %s AND year = %s AND month = %s",
            (req.department, req.year, req.month),
        )
    else:
        raise HTTPException(status_code=400, detail="请提供 start_date 与 end_date，或 year 与 month")
    if n < 0:
        raise HTTPException(status_code=500, detail="清空排班失败，请稍后重试")
    return {"success": True, "message": f"已清空 {n} 条排班记录"}


# ==================== 科室列表 ====================

@router.get("/departments")
def get_departments():
    """获取所有科室列表"""
    return {"success": True, "departments": _get_shift_departments()}


def _holiday_options_for_year(year: int) -> List[dict]:
    """某年的假期选项列表，按 holiday.festival 分组（供导出选择与节假日邮件共用）。"""
    grouped = {}
    for r in load_holidays_for_year(str(year)):
        ds = _normalize_date_key(r.get("date"))
        if not ds:
            continue
        ht = (r.get("type") or "").strip()
        if "班" in ht or ("假" not in ht and "休" not in ht):
            continue
        festival = (r.get("festival") or "").strip() or "未命名假期"
        item = grouped.setdefault(festival, {"name": festival, "dates": []})
        item["dates"].append(ds)
    options = []
    for item in grouped.values():
        dates_sorted = sorted(set(item["dates"]))
        options.append({
            "name": item["name"],
            "startDate": dates_sorted[0],
            "endDate": dates_sorted[-1],
            "days": len(dates_sorted),
            "dates": dates_sorted,
        })
    options.sort(key=lambda x: x["startDate"])
    return options


@router.get("/holiday-options")
def get_shift_holiday_options(year: int = Query(...)):
    """返回某年可导出的假期选项，按 holiday.festival 分组。"""
    return {"success": True, "year": year, "options": _holiday_options_for_year(year)}


def build_week_schedule_report(department: str, anchor: Optional[date] = None) -> dict:
    """生成周排班每日汇总，供导出与自动邮件共用。"""
    try:
        from html import escape
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl，无法生成周排班表")

    _ensure_tables()
    if not department or department == "__ALL__":
        raise HTTPException(status_code=400, detail="周排班表请选择具体科室")

    start_offset = _get_shift_email_start_offset_days(department)
    week_start, week_end = _week_range_for_send_day(
        (anchor or date.today()) + timedelta(days=start_offset),
        _get_shift_email_send_weekday(department),
        start_offset == 0,
        start_offset,
    )
    dates = _daterange(week_start, week_end)
    employees = _get_dept_employees(department)
    ds_lo = week_start.strftime("%Y-%m-%d")
    ds_hi = week_end.strftime("%Y-%m-%d")

    holiday_by_date = _holiday_map_for_years({d.year for d in dates})
    holidays = {k: v["type"] for k, v in holiday_by_date.items()}
    date_info = []
    for d in dates:
        ds = d.strftime("%Y-%m-%d")
        h = holiday_by_date.get(ds)
        ht = h["type"] if h else ""
        fest = h["festival"] if h else ""
        date_info.append({
            "date": ds,
            "weekday": d.weekday(),
            "isWorkday": _is_workday(d, holidays),
            "label": ["一", "二", "三", "四", "五", "六", "日"][d.weekday()],
            "holidayType": ht,
            "holidayFestival": fest,
            "holidayMark": _holiday_header_mark(ht, fest) if h else "",
        })

    schedule = {}
    locations = {}
    if employees:
        ph = ",".join(["%s"] * len(employees))
        rows = db.execute_query(
            f"SELECT employee_name, shift_date, shift_type, shift_location FROM shift_schedule "
            f"WHERE department = %s AND shift_date >= %s AND shift_date <= %s AND employee_name IN ({ph})",
            (department, ds_lo, ds_hi) + tuple(employees),
        )
        for r in rows or []:
            name = (r.get("employee_name") or "").strip()
            sd = r.get("shift_date")
            sd = sd.strftime("%Y-%m-%d") if hasattr(sd, "strftime") else str(sd)[:10]
            st = (r.get("shift_type") or "").strip()
            sl = (r.get("shift_location") or "").strip()
            if name and sd:
                schedule.setdefault(name, {})[sd] = st
                if sl:
                    locations.setdefault(name, {})[sd] = sl

    day_plans = {}
    try:
        plan_rows = db.execute_query(
            "SELECT plan_date, content FROM shift_day_plan WHERE department = %s AND plan_date >= %s AND plan_date <= %s",
            (department, ds_lo, ds_hi),
        )
        for pr in plan_rows or []:
            pd = pr.get("plan_date")
            pds = pd.strftime("%Y-%m-%d") if hasattr(pd, "strftime") else str(pd)[:10]
            if pds:
                day_plans[pds] = (pr.get("content") or "").strip()
    except Exception:
        pass

    no_duty_dates = _load_no_duty_dates(department, ds_lo, ds_hi)
    people = _get_dept_people(department)
    people_by_name = {p["name"]: p for p in people}

    def _person(name: str) -> dict:
        return people_by_name.get(name) or {"name": name, "jb": "", "mobile": "", "telephone": ""}

    def _phone_line(name: str) -> str:
        p = _person(name)
        phones = [x for x in [(p.get("mobile") or "").strip(), (p.get("telephone") or "").strip()] if x]
        return f"{name}（{' / '.join(phones)}）" if phones else name

    def _uniq(values: List[str]) -> List[str]:
        return list(dict.fromkeys([v for v in values if v]))

    rows_data = []
    for di in date_info:
        ds = di["date"]
        is_no_duty = ds in no_duty_dates
        day_prepare, day_service, night_prepare, night_service = [], [], [], []
        duty_names = []
        for emp in employees:
            st = schedule.get(emp, {}).get(ds, "")
            loc = locations.get(emp, {}).get(ds, "")
            if st in ("白班", "白+夜"):
                duty_names.append(emp)
                if loc == "服务组":
                    day_service.append(emp)
                else:
                    day_prepare.append(emp)
            if st in ("夜班", "白+夜"):
                duty_names.append(emp)
                if loc == "准备组":
                    night_prepare.append(emp)
                else:
                    night_service.append(emp)
        contact_names = _uniq(duty_names)
        remark = "当日不设置值班" if is_no_duty else ""
        rows_data.append({
            "date": ds,
            "weekday": "\n".join([x for x in [f"星期{di['label']}", di.get("holidayFestival"), di.get("holidayType")] if x]),
            "dayPrepare": "\n".join(_phone_line(n) for n in _uniq(day_prepare)),
            "dayService": "\n".join(_phone_line(n) for n in _uniq(day_service)),
            "nightPrepare": "\n".join(_phone_line(n) for n in _uniq(night_prepare)),
            "nightService": "\n".join(_phone_line(n) for n in _uniq(night_service)),
            "contacts": "\n".join(_phone_line(n) for n in contact_names),
            "plan": day_plans.get(ds, ""),
            "count": len(set(duty_names)),
            "remark": remark,
            "noDuty": is_no_duty,
            "dateInfo": di,
        })

    thin_side = Side(style="thin", color="B0B0B0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    font_title = Font(name="微软雅黑", size=14, bold=True, color="1E293B")
    font_header = Font(name="微软雅黑", size=10, bold=True, color="1E293B")
    font_body = Font(name="微软雅黑", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_header = PatternFill("solid", fgColor="F1F5F9")
    fill_weekend = PatternFill("solid", fgColor="FEF9C3")
    fill_holiday_rest = PatternFill("solid", fgColor="FDE68A")
    fill_day_shift = PatternFill("solid", fgColor="DBEAFE")
    fill_night_shift = PatternFill("solid", fgColor="FEF3C7")
    fill_summary = PatternFill("solid", fgColor="F1F5F9")

    def _cell_fill_for_date(di):
        if not di["isWorkday"]:
            ht = di.get("holidayType", "")
            if "假" in ht or "休" in ht or di.get("holidayFestival"):
                return fill_holiday_rest
            return fill_weekend
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "每日汇总"
    ws.sheet_properties.tabColor = "10B981"
    title = f"{department} {ds_lo} 至 {ds_hi} 周排班每日汇总"
    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["日期", "星期/节假日", "白班准备组", "白班服务组", "夜班准备组", "夜班服务组", "联系方式", "工作计划", "当日值班人数", "备注"]
    keys = ["date", "weekday", "dayPrepare", "dayService", "nightPrepare", "nightService", "contacts", "plan", "count", "remark"]
    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=header)
        c.font = font_header
        c.alignment = align_center
        c.fill = fill_header
        c.border = thin_border
    widths = [13, 18, 30, 30, 30, 30, 42, 48, 12, 24]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row_idx, row in enumerate(rows_data, start=3):
        for col, key in enumerate(keys, start=1):
            c = ws.cell(row=row_idx, column=col, value=row[key])
            c.font = font_body
            c.alignment = Alignment(horizontal="left" if col in (3, 4, 5, 6, 7, 8, 10) else "center", vertical="center", wrap_text=True)
            c.border = thin_border
            cfill = _cell_fill_for_date(row["dateInfo"])
            if cfill and col in (1, 2):
                c.fill = cfill
            if col in (3, 4):
                c.fill = fill_day_shift
            elif col in (5, 6):
                c.fill = fill_night_shift
            elif col == 9:
                c.fill = fill_summary
        ws.row_dimensions[row_idx].height = 72
    ws.freeze_panes = "A3"

    bio = BytesIO()
    wb.save(bio)
    filename = f"{department}_{ds_lo}_至_{ds_hi}_周排班每日汇总.xlsx"

    html_headers = "".join(f"<th>{escape(h)}</th>" for h in headers)
    html_rows = []
    for row in rows_data:
        tds = []
        for key in keys:
            value = str(row[key] if row[key] is not None else "")
            tds.append(f"<td>{escape(value).replace(chr(10), '<br>')}</td>")
        html_rows.append("<tr>" + "".join(tds) + "</tr>")
    html_table = (
        '<table border="1" cellspacing="0" cellpadding="6" '
        'style="border-collapse:collapse;font-family:Microsoft YaHei,Arial,sans-serif;font-size:13px;">'
        f"<thead><tr>{html_headers}</tr></thead><tbody>{''.join(html_rows)}</tbody></table>"
    )

    return {
        "department": department,
        "week_start": week_start,
        "week_end": week_end,
        "title": title,
        "filename": filename,
        "excel_bytes": bio.getvalue(),
        "html_table": html_table,
        "rows_data": rows_data,
    }


# ==================== 导出排班 Excel ====================

@router.get("/export-excel")
def export_schedule_excel(
    department: str = Query(...),
    year: int = Query(...),
    month: Optional[int] = Query(None, ge=1, le=12),
    export_format: str = Query("month", alias="format"),
    holiday: str = Query("", description="format=holiday 时的假期名称"),
    week_date: Optional[str] = Query(None, description="format=week 时用于定位周六-周五排班周的日期"),
):
    """导出科室排班表 Excel。format=holiday 时按节假日值班表样式导出。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
    except ImportError:
        raise HTTPException(status_code=500, detail="服务端未安装 openpyxl，无法导出")

    _ensure_tables()
    fmt = (export_format or "").strip().lower()
    is_holiday_export = fmt in ("holiday", "festival", "duty")
    is_week_export = fmt in ("week", "weekly")

    if is_week_export:
        anchor = _parse_iso_date(week_date or "")
        if not anchor:
            try:
                anchor = date(year, int(month or 1), 1) if month else date.today()
            except Exception:
                anchor = date.today()
        report = build_week_schedule_report(department, anchor)
        return Response(
            content=report["excel_bytes"],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(report['filename'])}"},
        )

    if is_holiday_export:
        holiday_by_date = _holiday_map_for_years({year})
        holiday_name = (holiday or "").strip()
        if not holiday_name:
            raise HTTPException(status_code=400, detail="请选择要导出的假期")
        selected_dates = []
        for ds, h in sorted(holiday_by_date.items()):
            if (h.get("festival") or "").strip() != holiday_name:
                continue
            ht = (h.get("type") or "").strip()
            if "班" in ht:
                continue
            if "假" in ht or "休" in ht:
                d = _parse_iso_date(ds)
                if d:
                    selected_dates.append(d)
        dates = selected_dates
        employees = _get_dept_employees(department) if department and department != "__ALL__" else []
    elif is_week_export:
        if not department or department == "__ALL__":
            raise HTTPException(status_code=400, detail="周排班表请选择具体科室")
        anchor = _parse_iso_date(week_date or "")
        if not anchor:
            try:
                anchor = date(year, int(month or 1), 1) if month else date.today()
            except Exception:
                anchor = date.today()
        start_offset = _get_shift_email_start_offset_days(department)
        week_start, week_end = _week_range_for_send_day(
            anchor + timedelta(days=start_offset),
            _get_shift_email_send_weekday(department),
            start_offset == 0,
            start_offset,
        )
        dates = _daterange(week_start, week_end)
        employees = _get_dept_employees(department)
    else:
        if month is None:
            raise HTTPException(status_code=400, detail="请选择月份")
        if not department or department == "__ALL__":
            raise HTTPException(status_code=400, detail="月排班表请选择具体科室")
        employees = _get_dept_employees(department)
        dates = _month_dates(year, month)
    if not dates:
        raise HTTPException(status_code=400, detail="未找到可导出的日期")

    years_set = {d.year for d in dates}
    holiday_by_date = _holiday_map_for_years(years_set)
    holidays = {k: v["type"] for k, v in holiday_by_date.items()}

    date_info = []
    for d in dates:
        ds = d.strftime("%Y-%m-%d")
        wd = _is_workday(d, holidays)
        h = holiday_by_date.get(ds)
        ht = h["type"] if h else ""
        fest = h["festival"] if h else ""
        mark = _holiday_header_mark(ht, fest) if h else ""
        date_info.append({
            "date": ds, "weekday": d.weekday(), "isWorkday": wd,
            "label": ["一", "二", "三", "四", "五", "六", "日"][d.weekday()],
            "holidayType": ht, "holidayFestival": fest, "holidayMark": mark,
        })

    ds_lo, ds_hi = dates[0].strftime("%Y-%m-%d"), dates[-1].strftime("%Y-%m-%d")
    schedule = {}
    export_locations = {}
    if employees:
        ph = ",".join(["%s"] * len(employees))
        rows = db.execute_query(
            f"SELECT employee_name, shift_date, shift_type, shift_location FROM shift_schedule "
            f"WHERE department = %s AND shift_date >= %s AND shift_date <= %s AND employee_name IN ({ph})",
            (department, ds_lo, ds_hi) + tuple(employees),
        )
        for r in rows:
            name = (r.get("employee_name") or "").strip()
            sd = r.get("shift_date")
            sd = sd.strftime("%Y-%m-%d") if hasattr(sd, "strftime") else str(sd)[:10]
            st = (r.get("shift_type") or "").strip()
            sl = (r.get("shift_location") or "").strip()
            if name and sd:
                schedule.setdefault(name, {})[sd] = st
                if sl:
                    export_locations.setdefault(name, {})[sd] = sl

    day_plans = {}
    try:
        plan_rows = db.execute_query(
            "SELECT plan_date, content FROM shift_day_plan WHERE department = %s AND plan_date >= %s AND plan_date <= %s",
            (department, ds_lo, ds_hi),
        )
        for pr in plan_rows or []:
            pd = pr.get("plan_date")
            pds = pd.strftime("%Y-%m-%d") if hasattr(pd, "strftime") else str(pd)[:10]
            if pds:
                day_plans[pds] = (pr.get("content") or "").strip()
    except Exception:
        pass

    def _phone_text(person: Optional[dict]) -> str:
        if not person:
            return ""
        phone = (person.get("mobile") or person.get("telephone") or "").strip()
        return f"{person.get('name') or ''}\n{phone}" if phone else (person.get("name") or "")

    def _festival_export_name(selected_dates: List[dict]) -> str:
        festivals = []
        for di in selected_dates:
            fest = (di.get("holidayFestival") or "").strip()
            if fest and fest not in festivals:
                festivals.append(fest)
        if festivals:
            return "、".join(festivals[:2])
        return f"{month}月节假日"

    def _compact_plan(plan: str) -> str:
        plan = (plan or "").strip()
        if plan:
            return plan
        return "按当日生产、技术准备、工地服务及专项工作安排执行。"

    def _format_trip_line(trip_entries: List[tuple]) -> str:
        if not trip_entries:
            return ""
        parts = []
        for name, project in trip_entries:
            project = (project or "").strip()
            parts.append(f"{name}（{project}）" if project and project != "公出" else name)
        return f"公出人员：{'、'.join(parts)}"

    def _format_group_lines(title: str, names: List[str], people_by_name: dict, trip_entries: Optional[List[tuple]] = None) -> str:
        trip_line = _format_trip_line(trip_entries or [])
        if not names and not trip_line:
            return ""
        body = "、".join(names)
        phones = []
        for name in names:
            person = people_by_name.get(name)
            phone = (person or {}).get("telephone") or (person or {}).get("mobile") or ""
            if phone:
                phones.append(phone)
        lines = [title]
        if body:
            lines.append(body)
        if phones:
            lines.append("、".join(phones))
        if trip_line:
            lines.append(trip_line)
        return "\n".join(lines)

    def _format_group_rich(title: str, text: str):
        if not text:
            return ""
        rest = text[len(title):] if text.startswith(title) else f"\n{text}"
        return CellRichText([
            TextBlock(InlineFont(rFont="宋体", sz=9, b=True), title),
            rest,
        ])

    def _is_duty_contact_candidate(person: Optional[dict]) -> bool:
        jb = ((person or {}).get("jb") or "").strip()
        return ("主任" in jb) or ("组长" in jb)

    def _duty_contact_priority(person: Optional[dict]) -> int:
        """值班联系人择优：主任 > 副主任 > 组长 > 其他当天值班人员。"""
        jb = ((person or {}).get("jb") or "").strip()
        if "副主任" in jb:
            return 2
        if "主任" in jb:
            return 1
        if "组长" in jb:
            return 3
        return 9

    def _export_holiday_duty_excel():
        """按参考表样式导出节假日期间值班值宿人员安排表。"""
        holiday_dates = date_info
        scoped_departments = _get_shift_departments() if department == "__ALL__" else [department]
        scoped_departments = [d for d in scoped_departments if d]
        if not scoped_departments:
            raise HTTPException(status_code=400, detail="未找到可导出的科室")

        ds_lo = holiday_dates[0]["date"]
        ds_hi = holiday_dates[-1]["date"]
        dept_cache = {}
        for dept_name in scoped_departments:
            people = _get_dept_people(dept_name)
            employees_local = [p["name"] for p in people]
            people_by_name = {p["name"]: p for p in people}
            managers = [
                p for p in people
                if ("主任" in (p.get("jb") or "")) or ("组长" in (p.get("jb") or ""))
            ]
            if not managers:
                managers = people[:2]

            dept_schedule = {}
            dept_locations = {}
            if employees_local:
                ph = ",".join(["%s"] * len(employees_local))
                rows_local = db.execute_query(
                    f"SELECT employee_name, shift_date, shift_type, shift_location FROM shift_schedule "
                    f"WHERE department = %s AND shift_date >= %s AND shift_date <= %s AND employee_name IN ({ph})",
                    (dept_name, ds_lo, ds_hi) + tuple(employees_local),
                )
                for r in rows_local or []:
                    name = (r.get("employee_name") or "").strip()
                    sd = r.get("shift_date")
                    sd = sd.strftime("%Y-%m-%d") if hasattr(sd, "strftime") else str(sd)[:10]
                    st = (r.get("shift_type") or "").strip()
                    sl = (r.get("shift_location") or "").strip()
                    if name and sd:
                        dept_schedule.setdefault(name, {})[sd] = st
                        if sl:
                            dept_locations.setdefault(name, {})[sd] = sl

            dept_plans = {}
            plan_rows_local = db.execute_query(
                "SELECT plan_date, content FROM shift_day_plan WHERE department = %s AND plan_date >= %s AND plan_date <= %s",
                (dept_name, ds_lo, ds_hi),
            )
            for pr in plan_rows_local or []:
                pd = pr.get("plan_date")
                pds = pd.strftime("%Y-%m-%d") if hasattr(pd, "strftime") else str(pd)[:10]
                if pds:
                    dept_plans[pds] = (pr.get("content") or "").strip()
            dept_trips = _get_dept_business_trips(
                dept_name,
                datetime.strptime(ds_lo, "%Y-%m-%d").date(),
                datetime.strptime(ds_hi, "%Y-%m-%d").date(),
            )

            dept_cache[dept_name] = {
                "employees": employees_local,
                "peopleByName": people_by_name,
                "managers": managers,
                "schedule": dept_schedule,
                "locations": dept_locations,
                "plans": dept_plans,
                "businessTrips": dept_trips,
            }

        wb_h = Workbook()
        ws = wb_h.active
        sheet_name = "全部门汇总" if department == "__ALL__" else department
        ws.title = sheet_name[:31] if sheet_name else "节假日值班表"
        ws.sheet_properties.tabColor = "C00000"

        title_font = Font(name="宋体", size=18, bold=True)
        unit_font = Font(name="宋体", size=12, bold=True)
        header_font = Font(name="宋体", size=11, bold=True)
        body_font = Font(name="宋体", size=10)
        small_font = Font(name="宋体", size=9)
        red_fill = PatternFill("solid", fgColor="F4CCCC")
        grey_fill = PatternFill("solid", fgColor="D9EAD3")
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

        festival_name = _festival_export_name(holiday_dates)
        ws.merge_cells("A1:H1")
        ws["A1"] = f"{year}年{festival_name}期间科室（专业）值班值宿人员安排表"
        ws["A1"].font = title_font
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:H2")
        unit_name = "全部门（所有科室）" if department == "__ALL__" else department
        ws["A2"] = f"单位：{unit_name}"
        ws["A2"].font = unit_font
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 28

        headers = ["时    间", "值班领导\n联系方式", "值班负责人\n联系方式", "准备组值班人员\n联系方式", "服务组值班人员\n联系方式", "工作内容(包括技术准备和生产服务)", "", "当日出勤人数"]
        headers[2] = "值班联系人\n联系方式"
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(3, col, header)
            cell.font = header_font
            cell.alignment = center
            cell.fill = red_fill
            cell.border = border
        ws.merge_cells("F3:G3")
        ws.row_dimensions[3].height = 40

        for col, width in {"A": 17.5, "B": 20, "C": 20, "D": 42, "E": 42, "F": 58, "G": 12, "H": 14}.items():
            ws.column_dimensions[col].width = width

        current_row = 4

        for idx, di in enumerate(holiday_dates):
            ds = di["date"]
            date_obj = datetime.strptime(ds, "%Y-%m-%d").date()
            date_label = f"{date_obj.month}月{date_obj.day}日\n星期{di['label']}"
            if di.get("holidayMark"):
                date_label += f"\n{di['holidayMark']}"
            date_start_row = current_row

            daily_contact_candidates = []
            for dept_idx, dept_name in enumerate(scoped_departments):
                cached = dept_cache.get(dept_name) or {}
                people_by_name = cached.get("peopleByName") or {}
                dept_schedule = cached.get("schedule") or {}
                for emp_idx, emp in enumerate(cached.get("employees") or []):
                    if dept_schedule.get(emp, {}).get(ds, ""):
                        person = people_by_name.get(emp) or {"name": emp, "jb": ""}
                        daily_contact_candidates.append((_duty_contact_priority(person), dept_idx, emp_idx, person))
            daily_contact = sorted(daily_contact_candidates, key=lambda x: (x[0], x[1], x[2]))[0][3] if daily_contact_candidates else None

            daily_total = 0
            for dept_index, dept_name in enumerate(scoped_departments):
                cached = dept_cache.get(dept_name) or {}
                employees_local = cached.get("employees") or []
                people_by_name = cached.get("peopleByName") or {}
                dept_schedule = cached.get("schedule") or {}
                dept_locations = cached.get("locations") or {}
                dept_plans = cached.get("plans") or {}
                dept_trips = cached.get("businessTrips") or {}

                day_shift = []
                night_shift = []
                prepare_names = []
                service_names = []
                trip_entries = []
                for emp in employees_local:
                    st = dept_schedule.get(emp, {}).get(ds, "")
                    loc = dept_locations.get(emp, {}).get(ds, "")
                    trip_project = dept_trips.get(emp, {}).get(ds, "")
                    if st:
                        if loc == "准备组":
                            prepare_names.append(emp)
                        elif loc == "服务组":
                            service_names.append(emp)
                        elif st == "白班":
                            prepare_names.append(emp)
                        elif st == "夜班":
                            service_names.append(emp)
                        if st == "白班":
                            day_shift.append(emp)
                        elif st == "夜班":
                            night_shift.append(emp)
                    if trip_project:
                        trip_entries.append((emp, trip_project))

                trip_names = [name for name, _project in trip_entries]
                attendance_count = len(set(day_shift + night_shift + trip_names))
                daily_total += attendance_count
                prepare_text = _format_group_lines(dept_name, prepare_names, people_by_name, trip_entries)
                service_text = _format_group_lines(dept_name, service_names, people_by_name)
                plan_text = _compact_plan(dept_plans.get(ds, ""))
                trip_line = _format_trip_line(trip_entries)
                if trip_line:
                    plan_text = f"{plan_text}\n{trip_line}"

                values = [
                    date_label,
                    "",
                    _phone_text(daily_contact),
                    _format_group_rich(dept_name, prepare_text),
                    _format_group_rich(dept_name, service_text),
                    plan_text,
                    "",
                    attendance_count,
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(current_row, col, value)
                    cell.font = small_font if col in (4, 5, 6) else body_font
                    cell.alignment = center if col in (4, 5) else (left_top if col == 6 else center)
                    cell.border = border
                    if col == 8:
                        cell.fill = grey_fill
                ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
                ws.row_dimensions[current_row].height = 120
                current_row += 1

            total_row = current_row
            for col in range(1, 9):
                cell = ws.cell(total_row, col, "")
                cell.border = border
                cell.fill = grey_fill
                cell.alignment = center
                cell.font = header_font if col in (4, 8) else body_font
            ws.cell(total_row, 4, "当日总人数")
            ws.cell(total_row, 8, daily_total)
            ws.merge_cells(start_row=total_row, start_column=4, end_row=total_row, end_column=7)
            ws.row_dimensions[total_row].height = 28
            current_row += 1

            if current_row - date_start_row > 1:
                for merge_col in (1, 2, 3):
                    ws.merge_cells(start_row=date_start_row, start_column=merge_col, end_row=current_row - 1, end_column=merge_col)
                    ws.cell(date_start_row, merge_col).alignment = center

        ws.freeze_panes = "A4"
        bio_h = BytesIO()
        wb_h.save(bio_h)
        fname_h = f"{unit_name}_{year}年{festival_name}期间值班值宿人员安排表.xlsx"
        return Response(
            content=bio_h.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname_h)}"},
        )

    if (export_format or "").strip().lower() in ("holiday", "festival", "duty"):
        return _export_holiday_duty_excel()

    # ---- 公用样式 ----
    thin_side = Side(style="thin", color="B0B0B0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    font_title = Font(name="微软雅黑", size=14, bold=True, color="1E293B")
    font_header = Font(name="微软雅黑", size=10, bold=True, color="1E293B")
    font_body = Font(name="微软雅黑", size=10)
    font_plan_label = Font(name="微软雅黑", size=9, italic=True, color="0369A1")
    font_plan = Font(name="微软雅黑", size=8, color="334155")
    font_stat_label = Font(name="微软雅黑", size=9, bold=True, color="64748B")
    font_holiday_mark = Font(name="微软雅黑", size=8, bold=True, color="9333EA")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    fill_header = PatternFill("solid", fgColor="F1F5F9")
    fill_weekend = PatternFill("solid", fgColor="FEF9C3")
    fill_holiday_rest = PatternFill("solid", fgColor="FDE68A")
    fill_day_shift = PatternFill("solid", fgColor="DBEAFE")
    fill_night_shift = PatternFill("solid", fgColor="FEF3C7")
    fill_summary = PatternFill("solid", fgColor="F1F5F9")
    fill_plan_row = PatternFill("solid", fgColor="EFF6FF")

    def _apply_border(ws, min_row, max_row, min_col, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for c in row:
                c.border = thin_border

    def _cell_fill_for_date(di):
        """根据日期信息返回背景填充色"""
        if not di["isWorkday"]:
            ht = di.get("holidayType", "")
            if "假" in ht or "休" in ht or di.get("holidayFestival"):
                return fill_holiday_rest
            return fill_weekend
        return None

    def _export_week_schedule_excel():
        people = _get_dept_people(department)
        people_by_name = {p["name"]: p for p in people}

        def _person(name: str) -> dict:
            return people_by_name.get(name) or {"name": name, "jb": "", "mobile": "", "telephone": ""}

        def _mobile(name: str) -> str:
            return (_person(name).get("mobile") or "").strip()

        def _telephone(name: str) -> str:
            return (_person(name).get("telephone") or "").strip()

        def _phone_line(name: str) -> str:
            p = _person(name)
            phones = [x for x in [(p.get("mobile") or "").strip(), (p.get("telephone") or "").strip()] if x]
            return f"{name}（{' / '.join(phones)}）" if phones else name

        wb_w = Workbook()
        ws = wb_w.active
        ws.title = "每日汇总"
        ws.sheet_properties.tabColor = "10B981"

        week_start_text = dates[0].strftime("%Y-%m-%d")
        week_end_text = dates[-1].strftime("%Y-%m-%d")
        ws.merge_cells("A1:J1")
        t = ws["A1"]
        t.value = f"{department} {week_start_text} 至 {week_end_text} 周排班每日汇总"
        t.font = font_title
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        headers = ["日期", "星期/节假日", "白班准备组", "白班服务组", "夜班准备组", "夜班服务组", "联系方式", "工作计划", "当日值班人数", "备注"]
        for col, header in enumerate(headers, start=1):
            c = ws.cell(row=2, column=col, value=header)
            c.font = font_header
            c.alignment = align_center
            c.fill = fill_header
            c.border = thin_border
        widths = [13, 18, 30, 30, 30, 30, 42, 48, 12, 24]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        for row_idx, di in enumerate(date_info, start=3):
            ds = di["date"]
            day_prepare, day_service, night_prepare, night_service = [], [], [], []
            duty_names = []
            for emp in employees:
                st = schedule.get(emp, {}).get(ds, "")
                loc = export_locations.get(emp, {}).get(ds, "")
                if st in ("白班", "白+夜"):
                    duty_names.append(emp)
                    if loc == "服务组":
                        day_service.append(emp)
                    else:
                        day_prepare.append(emp)
                if st in ("夜班", "白+夜"):
                    duty_names.append(emp)
                    if loc == "准备组":
                        night_prepare.append(emp)
                    else:
                        night_service.append(emp)
            contact_names = []
            for name in duty_names:
                if name not in contact_names:
                    contact_names.append(name)
            values = [
                ds,
                "\n".join([x for x in [f"星期{di['label']}", di.get("holidayFestival"), di.get("holidayType")] if x]),
                "\n".join(_phone_line(n) for n in dict.fromkeys(day_prepare)),
                "\n".join(_phone_line(n) for n in dict.fromkeys(day_service)),
                "\n".join(_phone_line(n) for n in dict.fromkeys(night_prepare)),
                "\n".join(_phone_line(n) for n in dict.fromkeys(night_service)),
                "\n".join(_phone_line(n) for n in contact_names),
                day_plans.get(ds, ""),
                len(set(duty_names)),
                "",
            ]
            for col, value in enumerate(values, start=1):
                c = ws.cell(row=row_idx, column=col, value=value)
                c.font = font_body
                c.alignment = Alignment(horizontal="left" if col in (3, 4, 5, 6, 7, 8, 10) else "center", vertical="center", wrap_text=True)
                c.border = thin_border
                cfill = _cell_fill_for_date(di)
                if cfill and col in (1, 2):
                    c.fill = cfill
                if col in (3, 4):
                    c.fill = fill_day_shift
                elif col in (5, 6):
                    c.fill = fill_night_shift
                elif col == 9:
                    c.fill = fill_summary
            ws.row_dimensions[row_idx].height = 72
        ws.freeze_panes = "A3"

        bio_w = BytesIO()
        wb_w.save(bio_w)
        fname_w = f"{department}_{week_start_text}_至_{week_end_text}_周排班每日汇总.xlsx"
        return Response(
            content=bio_w.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname_w)}"},
        )

    if is_week_export:
        return _export_week_schedule_excel()

    wb = Workbook()

    # ============== Sheet 1: 表格形式 ==============
    ws1 = wb.active
    ws1.title = "排班表（表格）"
    ws1.sheet_properties.tabColor = "3B82F6"

    n_emp = len(employees)
    n_days = len(date_info)

    # -- 标题行 --
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + n_days)
    title_cell = ws1.cell(row=1, column=1, value=f"{department}  {year}年{month}月 排班表")
    title_cell.font = font_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # -- 表头行1: 日期 --
    header_row = 2
    ws1.cell(row=header_row, column=1, value="姓名").font = font_header
    ws1.cell(row=header_row, column=1).fill = fill_header
    ws1.cell(row=header_row, column=1).alignment = align_center
    ws1.cell(row=header_row, column=2, value="统计\n白/夜").font = font_header
    ws1.cell(row=header_row, column=2).fill = fill_header
    ws1.cell(row=header_row, column=2).alignment = align_center
    ws1.column_dimensions[get_column_letter(1)].width = 10
    ws1.column_dimensions[get_column_letter(2)].width = 8

    for ci, di in enumerate(date_info):
        col = 3 + ci
        header_text = f"{int(di['date'][8:])}\n{di['label']}"
        if di["holidayMark"]:
            header_text += f"\n{di['holidayMark']}"
        cell = ws1.cell(row=header_row, column=col, value=header_text)
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_header
        cfill = _cell_fill_for_date(di)
        if cfill:
            cell.fill = cfill
        ws1.column_dimensions[get_column_letter(col)].width = 6.5
    ws1.row_dimensions[header_row].height = 48

    # -- 员工行 --
    for ri, emp in enumerate(employees):
        row = header_row + 1 + ri
        ws1.cell(row=row, column=1, value=emp).font = font_body
        ws1.cell(row=row, column=1).alignment = align_left
        day_cnt, night_cnt = 0, 0
        for ci, di in enumerate(date_info):
            col = 3 + ci
            v = schedule.get(emp, {}).get(di["date"], "")
            loc = export_locations.get(emp, {}).get(di["date"], "")
            loc_short = "准" if loc == "准备组" else ("服" if loc == "服务组" else "")
            label = ""
            if v == "白班":
                label = f"白{loc_short}" if loc_short else "白"
                day_cnt += 1
            elif v == "夜班":
                label = f"夜{loc_short}" if loc_short else "夜"
                night_cnt += 1
            cell = ws1.cell(row=row, column=col, value=label)
            cell.font = font_body
            cell.alignment = align_center
            if v == "白班":
                cell.fill = fill_day_shift
            elif v == "夜班":
                cell.fill = fill_night_shift
            else:
                cfill = _cell_fill_for_date(di)
                if cfill:
                    cell.fill = cfill
        stat_cell = ws1.cell(row=row, column=2, value=f"{day_cnt}/{night_cnt}")
        stat_cell.font = font_body
        stat_cell.alignment = align_center

    # -- 合计行 --
    summary_row = header_row + 1 + n_emp
    ws1.cell(row=summary_row, column=1, value="当日合计").font = font_stat_label
    ws1.cell(row=summary_row, column=1).fill = fill_summary
    ws1.cell(row=summary_row, column=1).alignment = align_center
    ws1.cell(row=summary_row, column=2, value="—").font = font_stat_label
    ws1.cell(row=summary_row, column=2).fill = fill_summary
    ws1.cell(row=summary_row, column=2).alignment = align_center
    for ci, di in enumerate(date_info):
        col = 3 + ci
        day_cnt = sum(1 for e in employees if schedule.get(e, {}).get(di["date"]) == "白班")
        night_cnt = sum(1 for e in employees if schedule.get(e, {}).get(di["date"]) == "夜班")
        cell = ws1.cell(row=summary_row, column=col, value=f"{day_cnt}/{night_cnt}")
        cell.font = font_stat_label
        cell.alignment = align_center
        cell.fill = fill_summary

    # -- 值班计划行 --
    plan_row = summary_row + 1
    ws1.cell(row=plan_row, column=1, value="值班计划").font = font_plan_label
    ws1.cell(row=plan_row, column=1).fill = fill_plan_row
    ws1.cell(row=plan_row, column=1).alignment = align_center
    ws1.cell(row=plan_row, column=2, value="").fill = fill_plan_row
    for ci, di in enumerate(date_info):
        col = 3 + ci
        plan = day_plans.get(di["date"], "")
        cell = ws1.cell(row=plan_row, column=col, value=plan)
        cell.font = font_plan
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = fill_plan_row
    ws1.row_dimensions[plan_row].height = 60

    # 边框
    _apply_border(ws1, header_row, plan_row, 1, 2 + n_days)

    # 冻结首列+表头
    ws1.freeze_panes = "C3"

    # ============== Sheet 2: 日历形式 ==============
    ws2 = wb.create_sheet(title="排班表（日历）")
    ws2.sheet_properties.tabColor = "10B981"

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    t2 = ws2.cell(row=1, column=1, value=f"{department}  {year}年{month}月 排班日历")
    t2.font = font_title
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    weekday_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    for ci, wn in enumerate(weekday_names):
        cell = ws2.cell(row=2, column=1 + ci, value=wn)
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_header
        if ci >= 5:
            cell.fill = fill_weekend
        ws2.column_dimensions[cell.column_letter].width = 22
    ws2.row_dimensions[2].height = 22

    first_weekday = dates[0].weekday()
    cal_row = 3
    cal_col = first_weekday

    for idx, di in enumerate(date_info):
        ds = di["date"]
        day_num = int(ds[8:])

        lines = [f"{day_num}日  星期{di['label']}"]
        if di["holidayMark"]:
            lines[0] += f"  [{di['holidayMark']}]"

        day_emps = []
        night_emps = []
        for emp in employees:
            v = schedule.get(emp, {}).get(ds, "")
            loc = export_locations.get(emp, {}).get(ds, "")
            loc_tag = f"({loc})" if loc else ""
            if v == "白班":
                day_emps.append(emp + loc_tag)
            elif v == "夜班":
                night_emps.append(emp + loc_tag)
        if day_emps:
            lines.append(f"白班({len(day_emps)})：{'、'.join(day_emps)}")
        if night_emps:
            lines.append(f"夜班({len(night_emps)})：{'、'.join(night_emps)}")
        if not day_emps and not night_emps:
            lines.append("无排班")

        plan = day_plans.get(ds, "").strip()
        if plan:
            plan_display = plan if len(plan) <= 60 else plan[:57] + "…"
            lines.append(f"[计划] {plan_display}")

        cell = ws2.cell(row=cal_row, column=1 + cal_col, value="\n".join(lines))
        cell.font = font_body
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = thin_border

        cfill = _cell_fill_for_date(di)
        if cfill:
            cell.fill = cfill

        if plan:
            cell.font = Font(name="微软雅黑", size=9, color="1E293B")

        cal_col += 1
        if cal_col >= 7:
            ws2.row_dimensions[cal_row].height = 90
            cal_col = 0
            cal_row += 1

    if cal_col != 0:
        ws2.row_dimensions[cal_row].height = 90
        for empty_col in range(cal_col, 7):
            ws2.cell(row=cal_row, column=1 + empty_col).border = thin_border
    _apply_border(ws2, 2, 2, 1, 7)

    # -- 图例 --
    legend_row = cal_row + (1 if cal_col == 0 else 2)
    ws2.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=7)
    legend_cell = ws2.cell(row=legend_row, column=1, value="图例：  ■ 黄色底 = 休息日/节假日    ■ 蓝色底 = 白班    ■ 橙色底 = 夜班    [计划] = 当日值班工作计划")
    legend_cell.font = Font(name="微软雅黑", size=9, color="64748B")
    legend_cell.alignment = Alignment(horizontal="left", vertical="center")

    # -- 输出 --
    bio = BytesIO()
    wb.save(bio)
    data = bio.getvalue()
    fname = f"{department}_{year}年{month}月_排班表.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )
