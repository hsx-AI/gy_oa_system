# -*- coding: utf-8 -*-
"""
Import paid annual leave records from qj_带薪年休假_2026起.xlsx into MySQL qj.

Usage:
    cd fastapi_backend
    python scripts/import_paid_annual_leave_qj_from_xlsx.py
    python scripts/import_paid_annual_leave_qj_from_xlsx.py --commit

Default mode is dry-run. Add --commit to write rows.
When an Excel id already exists in qj, the script updates that row.
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any


SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
DEFAULT_XLSX = SCRIPT_DIR / "qj_带薪年休假_2026起.xlsx"

# Some local environments set DEBUG=release, which is not parseable by the
# current pydantic Settings.DEBUG bool. Keep this import-only script runnable.
_debug = (os.environ.get("DEBUG") or "").strip().lower()
if _debug and _debug not in {"1", "0", "true", "false", "yes", "no", "on", "off"}:
    os.environ["DEBUG"] = "true"

sys.path.insert(0, str(BACKEND_DIR))

from database import db  # noqa: E402


DEFAULT_VALUES: dict[str, Any] = {
    "jb": "带薪年休假",
    "bc": "行政白班",
    "jy": "带薪休假",
    "smclwj": "",
    "bhyy": None,
}

EMPLOYEE_FIELD_MAP = {
    "sfzh": "sfzh",
    "xb": "xbie",
    "lsys": "lsys",
}

INT_COLUMNS = {"qjzt", "2j", "sp2zt", "bianhaoweishu", "hxwc"}
FLOAT_COLUMNS = {"hxpxh", "hxps", "hxpsy"}
DATETIME_COLUMNS = {"timefrom", "timeto"}
DATE_TEXT_COLUMNS = {"timefromdate", "qjtime", "sptime", "sctime", "sp2time"}


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_datetime(value: Any) -> datetime | None:
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%y %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%y",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _format_datetime(value: Any) -> str | None:
    dt = _parse_datetime(value)
    if dt is None:
        return _clean_text(value)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _format_date_text(value: Any) -> str | None:
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return _clean_text(value)


def _to_int(value: Any) -> int | None:
    if _is_blank(value):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _to_float(value: Any) -> float | None:
    if _is_blank(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_value(column: str, value: Any) -> Any:
    if column == "id":
        return _clean_text(value)
    if column in DATETIME_COLUMNS:
        return _format_datetime(value)
    if column in DATE_TEXT_COLUMNS:
        return _format_date_text(value)
    if column in INT_COLUMNS:
        return _to_int(value)
    if column in FLOAT_COLUMNS:
        return _to_float(value)
    if isinstance(value, (datetime, date)):
        return _format_date_text(value)
    if isinstance(value, Decimal):
        return float(value)
    return _clean_text(value)


def load_xlsx(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] 缺少 openpyxl，请先执行：pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not raw_headers:
        raise RuntimeError("Excel 第一行为空，无法识别表头")

    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    rows: list[dict[str, Any]] = []
    for row_no, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rec = {h: values[idx] if idx < len(values) else None for idx, h in enumerate(headers) if h}
        if any(not _is_blank(v) for v in rec.values()):
            rec["_excel_row"] = row_no
            rows.append(rec)
    wb.close()
    return headers, rows


def load_table_columns() -> list[str]:
    rows = db.execute_query("SHOW COLUMNS FROM qj")
    if not rows:
        raise RuntimeError("无法读取 qj 表结构，请检查数据库连接")
    return [r["Field"] for r in rows]


def load_employee_map() -> dict[str, dict[str, Any]]:
    rows = db.execute_query(
        """
        SELECT TRIM(name) AS name, sfzh, xbie, lsys, jb
        FROM yggl
        WHERE name IS NOT NULL AND TRIM(name) != '' AND COALESCE(zaizhi, 0) = 0
        """
    )
    result: dict[str, dict[str, Any]] = {}
    duplicate_names: set[str] = set()
    for row in rows:
        name = (row.get("name") or "").strip()
        if not name:
            continue
        if name in result:
            duplicate_names.add(name)
        result[name] = row
    if duplicate_names:
        print(f"[WARN] yggl 在职员工姓名重复 {len(duplicate_names)} 个，将使用最后一条：{sorted(duplicate_names)}")
    return result


def load_existing_ids() -> set[str]:
    rows = db.execute_query("SELECT id FROM qj")
    return {(str(r.get("id")).strip()) for r in rows if not _is_blank(r.get("id"))}


def build_records(
    excel_rows: list[dict[str, Any]],
    qj_columns: list[str],
    employee_map: dict[str, dict[str, Any]],
    skip_unmatched_employee: bool,
) -> tuple[list[dict[str, Any]], dict[str, list[Any]]]:
    qj_column_set = set(qj_columns)
    insert_columns = [c for c in qj_columns if c in qj_column_set]
    records: list[dict[str, Any]] = []
    report: dict[str, list[Any]] = {
        "unmatched_employee": [],
        "missing_required": [],
        "generated_id": [],
    }

    for row in excel_rows:
        excel_row = row.get("_excel_row")
        name = _clean_text(row.get("xm")) or ""
        if not name or _is_blank(row.get("timefrom")) or _is_blank(row.get("timeto")):
            report["missing_required"].append(
                {
                    "row": excel_row,
                    "xm": name,
                    "timefrom": row.get("timefrom"),
                    "timeto": row.get("timeto"),
                }
            )
            continue

        emp = employee_map.get(name)
        if not emp:
            report["unmatched_employee"].append({"row": excel_row, "xm": name})
            if skip_unmatched_employee:
                continue

        rec: dict[str, Any] = {}
        for column in insert_columns:
            if column in row:
                rec[column] = _normalize_value(column, row.get(column))
            elif column in DEFAULT_VALUES:
                rec[column] = DEFAULT_VALUES[column]
            else:
                rec[column] = None

        if emp:
            for qj_field, emp_field in EMPLOYEE_FIELD_MAP.items():
                if qj_field in qj_column_set and _is_blank(rec.get(qj_field)):
                    rec[qj_field] = _normalize_value(qj_field, emp.get(emp_field))

        # This source workbook stores the leave type in jb. The qj table uses
        # qjfs for statistics/display, so jb must override qjfs.
        if "qjfs" in qj_column_set and not _is_blank(row.get("jb")):
            rec["qjfs"] = _normalize_value("qjfs", row.get("jb"))
        elif "qjfs" in qj_column_set and _is_blank(rec.get("qjfs")):
            rec["qjfs"] = "带薪年休假"

        if _is_blank(rec.get("bz")) and not _is_blank(rec.get("lsys")):
            rec["bz"] = rec["lsys"]
        if _is_blank(rec.get("qjr")):
            rec["qjr"] = name
        if _is_blank(rec.get("scr")):
            rec["scr"] = rec.get("spr") or rec.get("spr2") or ""
        if _is_blank(rec.get("content")):
            rec["content"] = ""
        if _is_blank(rec.get("smcl")):
            rec["smcl"] = "带薪年休假"
        if _is_blank(rec.get("tian")):
            rec["tian"] = "0"
        if _is_blank(rec.get("xiaoshi")):
            rec["xiaoshi"] = "0"
        if rec.get("qjzt") is None:
            rec["qjzt"] = 4

        if _is_blank(rec.get("id")):
            rec["id"] = f"xlsx_{excel_row}"
            report["generated_id"].append({"row": excel_row, "id": rec["id"], "xm": name})

        records.append(rec)

    return records, report


def print_list(title: str, rows: list[Any], limit: int = 30) -> None:
    if not rows:
        return
    print(f"[WARN] {title}: {len(rows)}")
    for item in rows[:limit]:
        print(f"  - {item}")
    if len(rows) > limit:
        print(f"  ... 还有 {len(rows) - limit} 条未显示")


def main() -> None:
    parser = argparse.ArgumentParser(description="导入 qj_带薪年休假_2026起.xlsx 到 qj 表")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Excel 路径，默认同目录 qj_带薪年休假_2026起.xlsx")
    parser.add_argument("--commit", action="store_true", help="实际写入数据库；默认仅预览")
    parser.add_argument("--skip-unmatched-employee", action="store_true", help="xm 无法匹配 yggl 在职员工时跳过该行")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).resolve()
    if not xlsx_path.exists():
        print(f"[ERROR] Excel 文件不存在：{xlsx_path}")
        sys.exit(1)

    print(f"[INFO] Excel: {xlsx_path}")
    print(f"[INFO] 模式: {'写入数据库' if args.commit else 'DRY-RUN（不写库）'}")

    headers, excel_rows = load_xlsx(xlsx_path)
    qj_columns = load_table_columns()
    qj_column_set = set(qj_columns)
    employee_map = load_employee_map()
    existing_ids = load_existing_ids()

    excel_columns = [h for h in headers if h]
    direct_columns = [c for c in excel_columns if c in qj_column_set]
    dropped_excel_columns = [c for c in excel_columns if c not in qj_column_set]
    filled_columns = [c for c in qj_columns if c not in direct_columns and (c in DEFAULT_VALUES or c in EMPLOYEE_FIELD_MAP or c == "qjfs")]
    untouched_qj_columns = [c for c in qj_columns if c not in direct_columns and c not in filled_columns]

    print(f"[INFO] Excel 行数: {len(excel_rows)}")
    print(f"[INFO] Excel 同名写入字段: {len(direct_columns)} 个：{direct_columns}")
    print(f"[INFO] 通过默认值或 xm->yggl 补充字段: {filled_columns}")
    if dropped_excel_columns:
        print(f"[WARN] Excel 有但 qj 表没有的字段，将忽略：{dropped_excel_columns}")
    print(f"[INFO] qj 中未由 Excel/默认值/yggl 填充的字段: {untouched_qj_columns}")

    records, report = build_records(excel_rows, qj_columns, employee_map, args.skip_unmatched_employee)

    to_insert: list[dict[str, Any]] = []
    to_update: list[dict[str, Any]] = []
    duplicate_ids: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for rec in records:
        rid = str(rec.get("id") or "").strip()
        name = rec.get("xm")
        if rid in seen_ids:
            duplicate_ids.append({"id": rid, "xm": name, "timefrom": rec.get("timefrom")})
            continue
        if rid in existing_ids:
            to_update.append(rec)
            seen_ids.add(rid)
            continue
        seen_ids.add(rid)
        to_insert.append(rec)

    print_list("xm 未匹配到 yggl 在职员工", report["unmatched_employee"])
    print_list("缺少 xm/timefrom/timeto，已跳过", report["missing_required"])
    print_list("id 为空，已生成临时 id", report["generated_id"])
    print_list("id 已存在或 Excel 内重复，已跳过", duplicate_ids)

    print(f"[INFO] 可新增记录数: {len(to_insert)}；可更新记录数: {len(to_update)} / 原始 Excel 有效行 {len(excel_rows)}")
    print("[PREVIEW] 前 5 条新增：")
    for idx, rec in enumerate(to_insert[:5], start=1):
        print(
            f"  [{idx}] id={rec.get('id')} xm={rec.get('xm')} qjfs={rec.get('qjfs')} "
            f"timefrom={rec.get('timefrom')} timeto={rec.get('timeto')} tian={rec.get('tian')} lsys={rec.get('lsys')}"
        )
    if to_update:
        print("[PREVIEW] 前 5 条更新：")
        for idx, rec in enumerate(to_update[:5], start=1):
            print(
                f"  [{idx}] id={rec.get('id')} xm={rec.get('xm')} qjfs={rec.get('qjfs')} "
                f"timefrom={rec.get('timefrom')} timeto={rec.get('timeto')} tian={rec.get('tian')} lsys={rec.get('lsys')}"
            )

    if not args.commit:
        print("[DRY-RUN] 未写入数据库。确认无误后执行：python scripts/import_paid_annual_leave_qj_from_xlsx.py --commit")
        return

    if not to_insert and not to_update:
        print("[DONE] 没有需要写入的记录。")
        return

    insert_columns = qj_columns
    col_sql = ", ".join(f"`{c}`" for c in insert_columns)
    placeholders = ", ".join(["%s"] * len(insert_columns))
    insert_sql = f"INSERT INTO qj ({col_sql}) VALUES ({placeholders})"

    inserted = 0
    failed: list[dict[str, Any]] = []
    for rec in to_insert:
        params = tuple(rec.get(c) for c in insert_columns)
        affected = db.execute_update(insert_sql, params)
        if affected > 0:
            inserted += 1
        else:
            failed.append({"id": rec.get("id"), "xm": rec.get("xm"), "timefrom": rec.get("timefrom")})

    updated = 0
    if to_update:
        update_columns = [c for c in qj_columns if c != "id"]
        set_sql = ", ".join(f"`{c}` = %s" for c in update_columns)
        update_sql = f"UPDATE qj SET {set_sql} WHERE id = %s"
        for rec in to_update:
            params = tuple(rec.get(c) for c in update_columns) + (rec.get("id"),)
            affected = db.execute_update(update_sql, params)
            if affected >= 0:
                updated += affected
            else:
                failed.append({"id": rec.get("id"), "xm": rec.get("xm"), "timefrom": rec.get("timefrom")})

    print(f"[DONE] 成功新增: {inserted}，成功更新/影响: {updated}，失败/未影响: {len(failed)}")
    print_list("写入失败/未影响记录", failed)


if __name__ == "__main__":
    main()
