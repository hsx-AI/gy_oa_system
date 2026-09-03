# -*- coding: utf-8 -*-
"""工艺码上办月报：月底自动邮件（含全年图表内嵌 + 本月 Excel 附件）。"""
from __future__ import annotations

import asyncio
import base64
import calendar
import io
import json
import logging
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, HTTPException, Query
from openpyxl import Workbook
from pydantic import BaseModel, Field

from database import db

logger = logging.getLogger(__name__)
router = APIRouter(tags=["工艺码上办邮件"])

MSB_EMAIL_CHECK_SECONDS = 60
MSB_EMAIL_CATCHUP_HOURS = 6
MSB_PORTAL_URL = "http://10.42.60.230"
MSB_PORTAL_PATH = "/admin/mashangban"

_CONFIG_COLUMNS = [
    ("mashangban_email_enabled", "TINYINT DEFAULT 0"),
    ("mashangban_email_hour", "INT DEFAULT 17"),
    ("mashangban_email_minute", "INT DEFAULT 0"),
    ("mashangban_email_recipients", "MEDIUMTEXT"),
    ("mashangban_email_use_auto_recipients", "TINYINT DEFAULT 1"),
    ("mashangban_email_log", "MEDIUMTEXT"),
]

# -------------------- config / admin --------------------

def _get_admin1() -> Optional[str]:
    try:
        rows = db.execute_query("SELECT admin1 FROM webconfig WHERE id = %s LIMIT 1", ("1",))
        if rows and rows[0].get("admin1") is not None:
            return (rows[0]["admin1"] or "").strip() or None
    except Exception:
        pass
    return None

def _require_admin(current_user: str):
    admin1 = _get_admin1()
    if not admin1 or (current_user or "").strip() != admin1:
        raise HTTPException(status_code=403, detail="仅系统管理员（webconfig.admin1）可操作")

def _ensure_email_config_columns():
    for col, typedef in _CONFIG_COLUMNS:
        try:
            db.execute_update(f"ALTER TABLE webconfig ADD COLUMN {col} {typedef}", ())
        except Exception:
            pass

def _ensure_email_log_table():
    try:
        db.execute_update(
            """
            CREATE TABLE IF NOT EXISTS mashangban_email_log (
              id INT AUTO_INCREMENT PRIMARY KEY,
              report_month CHAR(7) NOT NULL,
              trigger_label VARCHAR(100) NULL,
              recipient_count INT NOT NULL DEFAULT 0,
              status VARCHAR(20) NOT NULL DEFAULT 'ok',
              message VARCHAR(500) NULL,
              sent_at DATETIME NOT NULL,
              INDEX idx_msb_mail_month (report_month, status),
              INDEX idx_msb_mail_sent_at (sent_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """,
            (),
        )
    except Exception as e:
        logger.warning("mashangban_email_log 建表跳过: %s", e)

def _parse_json_list(raw) -> list:
    if raw is None or raw == "":
        return []
    if isinstance(raw, (bytes, bytearray)):
        raw = raw.decode("utf-8", errors="ignore")
    if isinstance(raw, list):
        return raw
    try:
        data = json.loads(raw)
        return data if isinstance(data, list) else []
    except Exception:
        return []

def _get_email_feature_config() -> dict:
    _ensure_email_config_columns()
    rows = db.execute_query(
        """
        SELECT mashangban_email_enabled, mashangban_email_hour, mashangban_email_minute,
               mashangban_email_recipients, mashangban_email_use_auto_recipients,
               mashangban_email_log
        FROM webconfig WHERE id = %s LIMIT 1
        """,
        ("1",),
    )
    r = rows[0] if rows else {}
    hour = int(r.get("mashangban_email_hour") if r.get("mashangban_email_hour") is not None else 17)
    minute = int(r.get("mashangban_email_minute") if r.get("mashangban_email_minute") is not None else 0)
    return {
        "enabled": bool(r.get("mashangban_email_enabled")),
        "hour": max(0, min(23, hour)),
        "minute": max(0, min(59, minute)),
        "useAutoRecipients": bool(
            r.get("mashangban_email_use_auto_recipients")
            if r.get("mashangban_email_use_auto_recipients") is not None
            else 1
        ),
        "recipients": [
            {
                "name": str(item.get("name") or "").strip(),
                "email": str(item.get("email") or "").strip(),
            }
            for item in _parse_json_list(r.get("mashangban_email_recipients"))
            if isinstance(item, dict) and (item.get("email") or "").strip()
        ],
        "log": _parse_json_list(r.get("mashangban_email_log"))[-30:],
    }

def _save_email_feature_config(
    enabled: bool,
    hour: int,
    minute: int,
    use_auto_recipients: bool,
    recipients: List[dict],
):
    _ensure_email_config_columns()
    cleaned = []
    for item in recipients or []:
        email = str((item or {}).get("email") or "").strip()
        if not email:
            continue
        cleaned.append({
            "name": str((item or {}).get("name") or "").strip(),
            "email": email,
        })
    db.execute_update(
        """
        UPDATE webconfig SET
          mashangban_email_enabled = %s,
          mashangban_email_hour = %s,
          mashangban_email_minute = %s,
          mashangban_email_use_auto_recipients = %s,
          mashangban_email_recipients = %s
        WHERE id = %s
        """,
        (
            1 if enabled else 0,
            max(0, min(23, int(hour))),
            max(0, min(59, int(minute))),
            1 if use_auto_recipients else 0,
            json.dumps(cleaned, ensure_ascii=False),
            "1",
        ),
    )

def _append_config_log(entry: dict):
    _ensure_email_config_columns()
    try:
        rows = db.execute_query(
            "SELECT mashangban_email_log FROM webconfig WHERE id = %s LIMIT 1", ("1",)
        )
        raw = (rows[0].get("mashangban_email_log") if rows else None) or "[]"
        log_list = _parse_json_list(raw)
        log_list.append(entry)
        db.execute_update(
            "UPDATE webconfig SET mashangban_email_log = %s WHERE id = %s",
            (json.dumps(log_list[-80:], ensure_ascii=False), "1"),
        )
    except Exception as e:
        logger.warning("写入码上办邮件配置日志失败: %s", e)

def _record_email_log(report_month: str, trigger_label: str, recipient_count: int, status: str, message: str):
    _ensure_email_log_table()
    db.execute_update(
        """
        INSERT INTO mashangban_email_log
          (report_month, trigger_label, recipient_count, status, message, sent_at)
        VALUES (%s,%s,%s,%s,%s,%s)
        """,
        (
            report_month,
            (trigger_label or "")[:100],
            int(recipient_count or 0),
            status,
            (message or "")[:500],
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ),
    )

def _has_sent_ok(report_month: str) -> bool:
    _ensure_email_log_table()
    rows = db.execute_query(
        "SELECT id FROM mashangban_email_log WHERE report_month=%s AND status='ok' LIMIT 1",
        (report_month,),
    )
    return bool(rows)

# -------------------- recipients --------------------

def _is_manager_level(jb: str) -> bool:
    j = (jb or "").strip()
    return j in {"经理", "副经理", "经理助理"}

def _is_dept_leader(jb: str) -> bool:
    from routers.approvers import _jb_match
    return _jb_match(jb, "主任") or _jb_match(jb, "副主任") or _jb_match(jb, "组长")

def _can_access_mashangban(name: str, jb: str, admin1: Optional[str]) -> bool:
    n = (name or "").strip()
    if admin1 and n == admin1.strip():
        return True
    if _is_manager_level(jb):
        return True
    return _is_dept_leader(jb)

def _collect_auto_recipients() -> List[dict]:
    """对齐 canAccessMashangban：admin1 + 经理层 + 主任/副主任/班组长，且有企业邮箱。"""
    admin1 = _get_admin1()
    rows = db.execute_query(
        """
        SELECT name, jb, lsys, enterprise_email
        FROM yggl
        WHERE name IS NOT NULL AND TRIM(name) != ''
          AND COALESCE(zaizhi, 0) = 0
          AND enterprise_email IS NOT NULL AND TRIM(enterprise_email) != ''
        """,
        (),
    ) or []
    result = []
    seen = set()
    for r in rows:
        name = (r.get("name") or "").strip()
        jb = (r.get("jb") or "").strip()
        email = (r.get("enterprise_email") or "").strip()
        if not name or not email:
            continue
        if not _can_access_mashangban(name, jb, admin1):
            continue
        key = email.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append({"name": name, "email": email, "jb": jb, "dept": (r.get("lsys") or "").strip()})
    return result

def _resolve_recipients(cfg: Optional[dict] = None) -> List[dict]:
    cfg = cfg or _get_email_feature_config()
    by_email: Dict[str, dict] = {}
    if cfg.get("useAutoRecipients", True):
        for item in _collect_auto_recipients():
            by_email[item["email"].lower()] = item
    for item in cfg.get("recipients") or []:
        email = (item.get("email") or "").strip()
        if not email:
            continue
        by_email[email.lower()] = {
            "name": (item.get("name") or "").strip(),
            "email": email,
        }
    return list(by_email.values())

# -------------------- data helpers --------------------

def _list_available_months() -> List[str]:
    rows = db.execute_query(
        """
        SELECT DISTINCT report_month
        FROM mashangban_dept_monthly
        ORDER BY report_month ASC
        """,
        (),
    ) or []
    return [(r.get("report_month") or "").strip() for r in rows if (r.get("report_month") or "").strip()]

def _fetch_dept(year_month: str) -> List[dict]:
    return db.execute_query(
        """
        SELECT dept_name AS deptName, order_count AS orderCount,
               total_service_hours AS totalServiceHours, avg_service_hours AS avgServiceHours,
               avg_accept_hours AS avgAcceptHours, avg_arrive_hours AS avgArriveHours,
               pending_accept AS pendingAccept, pending_arrive AS pendingArrive,
               processing, pending_confirm AS pendingConfirm
        FROM mashangban_dept_monthly
        WHERE report_month=%s
        ORDER BY order_count DESC, dept_name ASC
        """,
        (year_month,),
    ) or []

def _fetch_person(year_month: str) -> List[dict]:
    return db.execute_query(
        """
        SELECT dept_name AS deptName, employee_name AS employeeName,
               service_count AS serviceCount, total_service_hours AS totalServiceHours,
               type_simple AS typeSimple, type_normal AS typeNormal, type_complex AS typeComplex,
               type_hard AS typeHard, type_improve AS typeImprove,
               avg_service_hours AS avgServiceHours, avg_accept_hours AS avgAcceptHours,
               avg_arrive_hours AS avgArriveHours,
               patrol_count AS patrolCount,
               rate_excellent AS rateExcellent, rate_good AS rateGood, rate_normal AS rateNormal,
               rate_poor AS ratePoor, rate_bad AS rateBad
        FROM mashangban_person_monthly
        WHERE report_month=%s
        ORDER BY service_count DESC, employee_name ASC
        """,
        (year_month,),
    ) or []

def _fetch_orders(year_month: str) -> List[dict]:
    return db.execute_query(
        """
        SELECT order_no AS orderNo, dept_name AS deptName, order_type AS orderType,
               order_status AS orderStatus, process_engineer_name AS processEngineerName,
               operator_name AS operatorName, workpiece_name AS workpieceName,
               work_no AS workNo, created_at_src AS createdAt, rating_label AS ratingLabel
        FROM mashangban_work_orders
        WHERE report_month=%s
        ORDER BY created_at_src DESC, order_no DESC
        """,
        (year_month,),
    ) or []

def _num(v, default=0.0) -> float:
    try:
        if v is None:
            return default
        return float(v)
    except Exception:
        return default

def _sum_field(rows: List[dict], field: str) -> float:
    return sum(_num(r.get(field)) for r in rows)

def _weighted_avg(rows: List[dict], field: str, weight_field: str = "orderCount") -> Optional[float]:
    w_sum = 0.0
    v_sum = 0.0
    for r in rows:
        w = _num(r.get(weight_field))
        if w <= 0:
            continue
        v = r.get(field)
        if v is None:
            continue
        try:
            n = float(v)
        except Exception:
            continue
        w_sum += w
        v_sum += n * w
    if w_sum <= 0:
        return None
    return v_sum / w_sum

def _year_months_upto(report_month: str) -> List[str]:
    year = report_month[:4]
    available = [m for m in _list_available_months() if m.startswith(year) and m <= report_month]
    return available

# -------------------- charts --------------------

def _png_to_b64(fig) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("ascii")

def _new_axes(title: str, ylabel: str = ""):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from utils.chart_font import configure_matplotlib_cjk, get_cjk_font_properties

    configure_matplotlib_cjk()
    cjk = get_cjk_font_properties()
    text_kw = {"fontproperties": cjk} if cjk else {}
    fig, ax = plt.subplots(figsize=(8.2, 3.6), dpi=120)
    ax.set_title(title, fontsize=13, fontweight="bold", **text_kw)
    if ylabel:
        ax.set_ylabel(ylabel, **text_kw)
    ax.grid(True, alpha=0.28)
    return fig, ax, text_kw, cjk

def _style_xticks(ax, labels, text_kw, cjk):
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=0 if len(labels) <= 8 else 25, ha="center" if len(labels) <= 8 else "right", **text_kw)
    if cjk:
        for lbl in ax.get_yticklabels():
            lbl.set_fontproperties(cjk)

def _render_line_chart(title: str, labels: List[str], values: List[float], ylabel: str) -> str:
    import matplotlib.pyplot as plt

    fig, ax, text_kw, cjk = _new_axes(title, ylabel)
    ax.plot(range(len(labels)), values, marker="o", linewidth=2, markersize=5, color="#1677ff")
    _style_xticks(ax, labels, text_kw, cjk)
    fig.tight_layout()
    data = _png_to_b64(fig)
    plt.close(fig)
    return data

def _render_multi_line_chart(title: str, labels: List[str], series: List[dict], ylabel: str = "") -> str:
    import matplotlib.pyplot as plt

    fig, ax, text_kw, cjk = _new_axes(title, ylabel)
    colors = ["#1677ff", "#10b981", "#f59e0b", "#ef4444", "#8b5cf6"]
    for i, s in enumerate(series):
        ax.plot(
            range(len(labels)),
            s["values"],
            marker="o",
            linewidth=2,
            markersize=4,
            label=s["name"],
            color=colors[i % len(colors)],
        )
    if len(series) > 1:
        ax.legend(loc="best", fontsize=9, prop=cjk)
    _style_xticks(ax, labels, text_kw, cjk)
    fig.tight_layout()
    data = _png_to_b64(fig)
    plt.close(fig)
    return data

def _render_bar_chart(title: str, labels: List[str], values: List[float], ylabel: str) -> str:
    import matplotlib.pyplot as plt

    fig, ax, text_kw, cjk = _new_axes(title, ylabel)
    colors = ["#1677ff", "#36cfc9", "#73d13d", "#ffc53d", "#ff7a45", "#9254de", "#f759ab"]
    bar_colors = [colors[i % len(colors)] for i in range(len(labels))]
    ax.bar(range(len(labels)), values, color=bar_colors, width=0.68)
    _style_xticks(ax, labels, text_kw, cjk)
    fig.tight_layout()
    data = _png_to_b64(fig)
    plt.close(fig)
    return data

def _build_inline_charts(report_month: str) -> Tuple[List[dict], dict]:
    """返回 (inline_images, kpi_summary)。图表对齐页面「全年」视图 + 科室对比。"""
    months = _year_months_upto(report_month)
    if not months:
        return [], {}

    snaps = []
    for ym in months:
        snaps.append({"yearMonth": ym, "dept": _fetch_dept(ym), "person": _fetch_person(ym)})

    axis = [f"{int(ym[5:7])}月" for ym in months]
    order_vals = [_sum_field(s["dept"], "orderCount") for s in snaps]
    hours_vals = [round(_sum_field(s["dept"], "totalServiceHours"), 2) for s in snaps]
    avg_service = []
    avg_accept = []
    avg_arrive = []
    for s in snaps:
        vs = _weighted_avg(s["dept"], "avgServiceHours")
        va = _weighted_avg(s["dept"], "avgAcceptHours")
        vr = _weighted_avg(s["dept"], "avgArriveHours")
        avg_service.append(0 if vs is None else round(vs, 3))
        avg_accept.append(0 if va is None else round(va, 3))
        avg_arrive.append(0 if vr is None else round(vr, 3))

    pending_accept = [_sum_field(s["dept"], "pendingAccept") for s in snaps]
    pending_arrive = [_sum_field(s["dept"], "pendingArrive") for s in snaps]
    processing = [_sum_field(s["dept"], "processing") for s in snaps]
    pending_confirm = [_sum_field(s["dept"], "pendingConfirm") for s in snaps]

    # 当年累计（截至 report_month）科室对比：聚合
    dept_map: Dict[str, dict] = {}
    for s in snaps:
        for row in s["dept"]:
            name = (row.get("deptName") or "").strip()
            if not name:
                continue
            bucket = dept_map.setdefault(name, {"deptName": name, "orderCount": 0, "totalServiceHours": 0.0})
            bucket["orderCount"] += int(_num(row.get("orderCount")))
            bucket["totalServiceHours"] += _num(row.get("totalServiceHours"))
    dept_rows = sorted(dept_map.values(), key=lambda x: x["orderCount"], reverse=True)[:16]

    images = [
        {
            "cid": "msb_year_orders",
            "filename": "月度工单总量趋势.png",
            "content_base64": _render_line_chart("月度工单总量趋势", axis, order_vals, "工单数"),
        },
        {
            "cid": "msb_year_hours",
            "filename": "月度总服务时长趋势.png",
            "content_base64": _render_line_chart("月度总服务时长趋势", axis, hours_vals, "小时"),
        },
        {
            "cid": "msb_year_avg",
            "filename": "月度响应效率趋势.png",
            "content_base64": _render_multi_line_chart(
                "月度响应效率趋势",
                axis,
                [
                    {"name": "平均服务(h)", "values": avg_service},
                    {"name": "平均接单(h)", "values": avg_accept},
                    {"name": "平均到场(h)", "values": avg_arrive},
                ],
                "小时",
            ),
        },
        {
            "cid": "msb_year_status",
            "filename": "月度工单状态趋势.png",
            "content_base64": _render_multi_line_chart(
                "月度工单状态趋势",
                axis,
                [
                    {"name": "未接单", "values": pending_accept},
                    {"name": "待到场", "values": pending_arrive},
                    {"name": "处理中", "values": processing},
                    {"name": "待确认", "values": pending_confirm},
                ],
                "工单数",
            ),
        },
        {
            "cid": "msb_dept_orders",
            "filename": "各科室工单总数对比.png",
            "content_base64": _render_bar_chart(
                "各科室工单总数对比",
                [d["deptName"] for d in dept_rows],
                [d["orderCount"] for d in dept_rows],
                "工单数",
            ),
        },
        {
            "cid": "msb_dept_hours",
            "filename": "各科室总服务时长对比.png",
            "content_base64": _render_bar_chart(
                "各科室总服务时长对比",
                [d["deptName"] for d in dept_rows],
                [round(d["totalServiceHours"], 2) for d in dept_rows],
                "小时",
            ),
        },
    ]

    total_orders = int(sum(order_vals))
    total_hours = round(sum(hours_vals), 3)
    kpi = {
        "year": report_month[:4],
        "months": months,
        "monthCount": len(months),
        "totalOrders": total_orders,
        "totalHours": total_hours,
        "deptCount": len(dept_map),
        "personCount": len({
            f"{(p.get('deptName') or '').strip()}::{(p.get('employeeName') or '').strip()}"
            for s in snaps
            for p in s["person"]
            if (p.get("employeeName") or "").strip()
        }),
        "rangeLabel": f"{months[0]} ~ {months[-1]}" if len(months) > 1 else months[0],
    }
    return images, kpi

# -------------------- excel --------------------

def _build_month_xlsx_bytes(report_month: str) -> bytes:
    dept_rows = _fetch_dept(report_month)
    person_rows = _fetch_person(report_month)
    order_rows = _fetch_orders(report_month)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "科室统计"
    ws1.append([
        "科室", "工单总数", "总服务时长", "平均服务", "平均接单", "平均到场",
        "未接单", "待到场", "处理中", "待确认",
    ])
    for r in dept_rows:
        ws1.append([
            r.get("deptName"), r.get("orderCount"), r.get("totalServiceHours"),
            r.get("avgServiceHours"), r.get("avgAcceptHours"), r.get("avgArriveHours"),
            r.get("pendingAccept"), r.get("pendingArrive"), r.get("processing"), r.get("pendingConfirm"),
        ])

    ws2 = wb.create_sheet("人员服务绩效")
    ws2.append([
        "科室", "姓名", "服务频次", "总服务时长",
        "简单", "一般", "复杂", "疑难", "改进",
        "平均服务", "平均接单", "平均到场", "巡视次数",
        "很好", "较好", "一般", "较差", "很差",
    ])
    for r in person_rows:
        ws2.append([
            r.get("deptName"), r.get("employeeName"), r.get("serviceCount"), r.get("totalServiceHours"),
            r.get("typeSimple"), r.get("typeNormal"), r.get("typeComplex"), r.get("typeHard"), r.get("typeImprove"),
            r.get("avgServiceHours"), r.get("avgAcceptHours"), r.get("avgArriveHours"), r.get("patrolCount"),
            r.get("rateExcellent"), r.get("rateGood"), r.get("rateNormal"), r.get("ratePoor"), r.get("rateBad"),
        ])

    ws3 = wb.create_sheet("工单明细")
    ws3.append([
        "工单号", "科室", "类型", "状态", "工艺员", "操作员", "工件", "工作号", "创建时间", "评价",
    ])
    for r in order_rows:
        created = r.get("createdAt")
        if hasattr(created, "strftime"):
            created = created.strftime("%Y-%m-%d %H:%M:%S")
        ws3.append([
            r.get("orderNo"), r.get("deptName"), r.get("orderType"), r.get("orderStatus"),
            r.get("processEngineerName"), r.get("operatorName"), r.get("workpieceName"),
            r.get("workNo"), created, r.get("ratingLabel"),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _build_html_body(report_month: str, kpi: dict, chart_cids: List[str]) -> str:
    y, m = report_month.split("-")
    month_label = f"{int(y)}年{int(m)}月"
    portal = f"{MSB_PORTAL_URL.rstrip('/')}{MSB_PORTAL_PATH}"
    kpi_html = f"""
    <table cellpadding="8" cellspacing="0" style="border-collapse:collapse;margin:12px 0;font-size:14px;">
      <tr>
        <td style="background:#f0f7ff;border:1px solid #d6e4ff;"><b>年度工单总数</b><br>{kpi.get('totalOrders', 0)}</td>
        <td style="background:#f0fdf4;border:1px solid #bbf7d0;"><b>总服务时长(h)</b><br>{kpi.get('totalHours', 0)}</td>
        <td style="background:#fff7ed;border:1px solid #fed7aa;"><b>涉及科室</b><br>{kpi.get('deptCount', 0)}</td>
        <td style="background:#faf5ff;border:1px solid #e9d5ff;"><b>服务人员</b><br>{kpi.get('personCount', 0)}</td>
      </tr>
    </table>
    <p style="color:#64748b;font-size:13px;">统计区间：{kpi.get('rangeLabel', '')} · 覆盖 {kpi.get('monthCount', 0)} 个月</p>
    """
    imgs = "".join(
        f'<div style="margin:14px 0;"><img src="cid:{cid}" alt="{cid}" style="max-width:100%;border:1px solid #e5e7eb;border-radius:8px;"></div>'
        for cid in chart_cids
    )
    return f"""
    <div style="font-family:Microsoft YaHei,PingFang SC,Arial,sans-serif;color:#1f2937;line-height:1.6;">
      <h2 style="margin:0 0 8px;">{month_label}工艺码上办月报</h2>
      <p>本邮件由系统于月底自动发送，正文包含本年度截至当月的趋势图与科室对比图；附件为<strong>{month_label}</strong>单月导出报表。</p>
      {kpi_html}
      {imgs}
      <p style="margin-top:18px;">详情可访问
        <a href="{portal}">智能制造工艺部集成办公平台</a>
        （{MSB_PORTAL_URL}）码上办月报查看。
      </p>
      <p style="color:#94a3b8;font-size:12px;">本邮件为系统自动发送，请勿直接回复。</p>
    </div>
    """

# -------------------- send --------------------

async def run_mashangban_email_once(
    report_month: Optional[str] = None,
    trigger_label: str = "月底自动发送",
    force: bool = False,
    extra_recipients: Optional[List[str]] = None,
    test_only: bool = False,
) -> dict:
    from routers.email_sender import (
        AttachmentItem,
        _build_email_message,
        _get_email_config,
        _release_mysql_lock,
        _smtp_send,
        _try_acquire_mysql_lock,
    )

    if not report_month:
        today = date.today()
        report_month = f"{today.year:04d}-{today.month:02d}"
    report_month = report_month.strip()
    if len(report_month) != 7 or report_month[4] != "-":
        return {"success": False, "message": f"无效月份: {report_month}"}

    lock_name = f"oa_mashangban_email_{report_month}"
    lock_conn = _try_acquire_mysql_lock(lock_name, "MashangbanEmail")
    if not lock_conn:
        return {"success": True, "message": "已有其他进程正在发送码上办月报邮件，本进程已跳过", "sent": 0}

    try:
        if not force and _has_sent_ok(report_month):
            msg = f"{report_month} 本月邮件已发送过，已跳过"
            _append_config_log({
                "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "reportMonth": report_month,
                "trigger": trigger_label,
                "status": "skipped",
                "message": msg,
            })
            return {"success": True, "message": msg, "sent": 0, "skipped": True}

        available = _list_available_months()
        if report_month not in available:
            msg = f"{report_month} 尚无入库数据，无法发送"
            _record_email_log(report_month, trigger_label, 0, "skipped", msg)
            _append_config_log({
                "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "reportMonth": report_month,
                "trigger": trigger_label,
                "status": "skipped",
                "message": msg,
            })
            return {"success": False, "message": msg, "sent": 0}

        cfg_feature = _get_email_feature_config()
        if test_only and extra_recipients:
            recipients = [{"name": "", "email": (e or "").strip()} for e in extra_recipients if (e or "").strip()]
        else:
            recipients = _resolve_recipients(cfg_feature)
            if extra_recipients:
                for email in extra_recipients:
                    e = (email or "").strip()
                    if e and e.lower() not in {r["email"].lower() for r in recipients}:
                        recipients.append({"name": "", "email": e})
        to_emails = [r["email"] for r in recipients if r.get("email")]
        if not to_emails:
            msg = "无有效收件人（请检查自动收件人或手工名单）"
            _record_email_log(report_month, trigger_label, 0, "error", msg)
            return {"success": False, "message": msg, "sent": 0}

        smtp_cfg = _get_email_config()
        sender = smtp_cfg["address"]
        password = smtp_cfg["auth_code"]
        if not sender or not password:
            msg = "邮箱未配置（webconfig.email_address / email_auth_code）"
            _record_email_log(report_month, trigger_label, 0, "error", msg)
            return {"success": False, "message": msg, "sent": 0}

        try:
            inline_images, kpi = _build_inline_charts(report_month)
        except Exception as e:
            logger.exception("生成码上办图表失败")
            msg = f"生成图表失败: {e}"
            _record_email_log(report_month, trigger_label, 0, "error", msg)
            return {"success": False, "message": msg, "sent": 0}

        xlsx_bytes = _build_month_xlsx_bytes(report_month)
        y, m = report_month.split("-")
        subject = f"{int(y)}年{int(m)}月的工艺码上办月报（自动发送）"
        html = _build_html_body(report_month, kpi, [img["cid"] for img in inline_images])
        att = AttachmentItem(
            filename=f"工艺码上办月报_{report_month}.xlsx",
            content_base64=base64.b64encode(xlsx_bytes).decode("ascii"),
        )
        message = _build_email_message(
            sender,
            to_emails,
            [],
            subject,
            html,
            "html",
            [att],
            inline_images=inline_images,
        )
        # SMTP 可能较慢，放到线程避免阻塞事件循环
        await asyncio.to_thread(_smtp_send, sender, password, to_emails, message)

        ok_msg = f"已发送给 {len(to_emails)} 人，附件 {att.filename}，图表 {len(inline_images)} 张"
        _record_email_log(report_month, trigger_label, len(to_emails), "ok", ok_msg)
        _append_config_log({
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "reportMonth": report_month,
            "trigger": trigger_label,
            "status": "ok",
            "recipientCount": len(to_emails),
            "message": ok_msg,
        })
        logger.info("[MashangbanEmail] %s %s", report_month, ok_msg)
        return {
            "success": True,
            "message": ok_msg,
            "sent": 1,
            "reportMonth": report_month,
            "recipientCount": len(to_emails),
            "recipients": to_emails,
            "chartCount": len(inline_images),
            "kpi": kpi,
        }
    except Exception as e:
        logger.exception("码上办月报邮件发送失败")
        msg = str(e)[:480]
        try:
            _record_email_log(report_month, trigger_label, 0, "error", msg)
        except Exception:
            pass
        return {"success": False, "message": msg, "sent": 0}
    finally:
        _release_mysql_lock(lock_conn, lock_name, "MashangbanEmail")

async def mashangban_email_background_loop():
    """月底按配置时刻发送；窗口内补触发，避免卡顿漏发。"""
    logger.info("[MashangbanEmail] 码上办月报邮件后台任务已启动")
    print("[System] 码上办月报月底自动邮件后台任务已启动")
    last_triggered = set()
    while True:
        try:
            await asyncio.sleep(MSB_EMAIL_CHECK_SECONDS)
            cfg = _get_email_feature_config()
            if not cfg.get("enabled"):
                continue
            now = datetime.now()
            last_day = calendar.monthrange(now.year, now.month)[1]
            if now.day != last_day:
                continue
            send_hour = int(cfg.get("hour", 17))
            send_minute = int(cfg.get("minute", 0))
            # 到点后 CATCHUP 小时内可补发
            send_minutes = send_hour * 60 + send_minute
            now_minutes = now.hour * 60 + now.minute
            if now_minutes < send_minutes:
                continue
            if now_minutes > send_minutes + MSB_EMAIL_CATCHUP_HOURS * 60:
                continue

            report_month = f"{now.year:04d}-{now.month:02d}"
            run_key = f"{report_month}|auto"
            if run_key in last_triggered:
                continue
            last_triggered.add(run_key)
            if len(last_triggered) > 24:
                last_triggered = set(sorted(last_triggered)[-12:])

            logger.info(
                "[MashangbanEmail] 触发月底自动发送 now=%s month=%s",
                now.strftime("%Y-%m-%d %H:%M:%S"),
                report_month,
            )
            await run_mashangban_email_once(
                report_month=report_month,
                trigger_label=f"月底{send_hour:02d}:{send_minute:02d}自动发送",
                force=False,
            )
        except Exception as e:
            logger.error("[MashangbanEmail] 循环异常: %s", e)
            await asyncio.sleep(300)

# -------------------- API --------------------

class MashangbanEmailConfigRequest(BaseModel):
    current_user: str
    enabled: bool = False
    hour: int = 17
    minute: int = 0
    useAutoRecipients: bool = True
    recipients: List[dict] = Field(default_factory=list)

class MashangbanEmailRunRequest(BaseModel):
    current_user: str
    yearMonth: Optional[str] = None
    force: bool = False
    testRecipients: Optional[List[str]] = None
    testOnly: bool = True

@router.get("/email-config")
def get_mashangban_email_config(current_user: str = Query(...)):
    _require_admin(current_user)
    cfg = _get_email_feature_config()
    auto_recipients = _collect_auto_recipients()
    return {
        "success": True,
        **cfg,
        "autoRecipientCount": len(auto_recipients),
        "autoRecipientsPreview": auto_recipients[:30],
        "portalUrl": f"{MSB_PORTAL_URL.rstrip('/')}{MSB_PORTAL_PATH}",
    }

@router.post("/email-config")
def save_mashangban_email_config(req: MashangbanEmailConfigRequest):
    _require_admin(req.current_user)
    _save_email_feature_config(
        enabled=bool(req.enabled),
        hour=int(req.hour),
        minute=int(req.minute),
        use_auto_recipients=bool(req.useAutoRecipients),
        recipients=req.recipients or [],
    )
    return {"success": True, "message": "码上办月报邮件配置已保存", **_get_email_feature_config()}

@router.post("/run-email")
async def run_mashangban_email_api(req: MashangbanEmailRunRequest):
    _require_admin(req.current_user)
    test_only = bool(req.testOnly and req.testRecipients)
    result = await run_mashangban_email_once(
        report_month=req.yearMonth,
        trigger_label="管理员手动发送",
        force=bool(req.force),
        extra_recipients=req.testRecipients,
        test_only=test_only,
    )
    return result

@router.get("/email-log")
def get_mashangban_email_log(current_user: str = Query(...), limit: int = Query(20, ge=1, le=100)):
    _require_admin(current_user)
    _ensure_email_log_table()
    rows = db.execute_query(
        """
        SELECT id, report_month AS reportMonth, trigger_label AS triggerLabel,
               recipient_count AS recipientCount, status, message, sent_at AS sentAt
        FROM mashangban_email_log
        ORDER BY id DESC
        LIMIT %s
        """,
        (limit,),
    ) or []
    return {"success": True, "items": rows}
