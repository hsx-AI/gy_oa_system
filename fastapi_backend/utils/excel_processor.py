# -*- coding: utf-8 -*-
"""
Excel 考勤数据处理模块
"""
import openpyxl
from openpyxl import load_workbook
import xlrd
from typing import List, Dict, Optional
from datetime import datetime, time, timedelta
import logging
from collections import defaultdict
import os

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Excel 处理器"""
    
    def __init__(self, file_path: str):
        """初始化处理器"""
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.is_xls = file_path.lower().endswith('.xls')
        self.xlrd_book = None
        self.xlrd_sheet = None

    @staticmethod
    def _clean_employee_id(raw) -> str:
        """清理工号：去浮点尾巴 (1234.0 → 1234)、去前后空格"""
        if raw is None:
            return ""
        if isinstance(raw, float):
            if raw == int(raw):
                return str(int(raw))
            return str(raw).strip()
        return str(raw).strip()
    
    def load_file(self) -> bool:
        """加载 Excel 文件（支持 .xls 和 .xlsx）"""
        try:
            if self.is_xls:
                # 使用 xlrd 读取 .xls 文件
                self.xlrd_book = xlrd.open_workbook(self.file_path)
                self.xlrd_sheet = self.xlrd_book.sheet_by_index(0)
                logger.info(f"成功加载 .xls 文件: {self.file_path}")
            else:
                # 使用 openpyxl 读取 .xlsx 文件
                self.workbook = load_workbook(self.file_path, data_only=True)
                self.worksheet = self.workbook.active
                logger.info(f"成功加载 .xlsx 文件: {self.file_path}")
            return True
        except Exception as e:
            logger.error(f"加载文件失败: {str(e)}")
            return False
    
    @staticmethod
    def _parse_inout_mark(raw) -> Optional[int]:
        """解析 H 列进出字面量：含「入」「进」→0(进)，含「出」→1(出)；识别不到返回 None（合并时再按时间线与前一条交替推断）。"""
        if raw is None:
            return None
        s = str(raw).strip()
        if not s:
            return None
        if "入" in s or "进" in s:
            return 0
        if "出" in s:
            return 1
        return None

    @staticmethod
    def _dedup_close_times(group: List[Dict], threshold_sec: int = 5) -> List[Dict]:
        """
        去除同日同人、间隔 ≤ threshold_sec 的重复打卡。
        仅当相邻两条均无 H 列进出标识（inout_mark 为 None）时才视为重复并丢弃后一条；
        任一条有明确「进/出」标识则全部保留，供后续按标记生成建议。
        必须在按 attendance_time 排序后调用。
        """
        if len(group) <= 1:
            return group
        result = [group[0]]
        for rec in group[1:]:
            prev = result[-1]
            try:
                t_prev = datetime.strptime(prev["attendance_time"], "%H:%M:%S")
                t_curr = datetime.strptime(rec["attendance_time"], "%H:%M:%S")
                diff = abs((t_curr - t_prev).total_seconds())
            except (ValueError, KeyError):
                diff = threshold_sec + 1
            if diff <= threshold_sec:
                prev_marked = prev.get("inout_mark") is not None
                curr_marked = rec.get("inout_mark") is not None
                if not prev_marked and not curr_marked:
                    continue
            result.append(rec)
        return result

    @staticmethod
    def _resolve_inout_mark_sequence(group: List[Dict], neutral_threshold_sec: int = 600) -> List[int]:
        """
        按已排序的同日打卡序列，为每条生成 0/1。
        有字面量则用字面量；否则：前一条是进则本条为出，前一条是出则本条为进；
        当日第一条仍无法识别时，视为进(0)。

        中立刷卡沿用：当前记录无字面量且与上一条时间差 ≤ neutral_threshold_sec
        时，沿用上一条标记（不交替），无论上一条是明确标记还是中立推断。
        避免门口刷脸进 + 几分钟后内部补刷被错标为一进一出。
        """
        THRESHOLD = neutral_threshold_sec
        out: List[int] = []
        prev: Optional[int] = None
        prev_time: Optional[str] = None

        for record in group:
            raw = record.get("inout_mark")
            cur_time = record.get("attendance_time", "")

            if raw is not None:
                resolved = int(raw)
                if resolved not in (0, 1):
                    resolved = 0 if prev is None else (1 if prev == 0 else 0)
            else:
                same_as_prev = False
                if prev is not None and prev_time and cur_time:
                    try:
                        t_p = datetime.strptime(prev_time, "%H:%M:%S")
                        t_c = datetime.strptime(cur_time, "%H:%M:%S")
                        same_as_prev = abs((t_c - t_p).total_seconds()) <= THRESHOLD
                    except (ValueError, KeyError):
                        pass

                if same_as_prev:
                    resolved = prev
                elif prev is None:
                    resolved = 0
                else:
                    resolved = 1 if prev == 0 else 0

            out.append(resolved)
            prev = resolved
            prev_time = cur_time
        return out

    def parse_time_value(self, value) -> str:
        """解析时间值"""
        if value is None or value == "":
            return ""
        
        # 如果是 datetime 对象
        if isinstance(value, datetime):
            return value.strftime("%H:%M:%S")
        
        # 如果是 time 对象
        if isinstance(value, time):
            return value.strftime("%H:%M:%S")
        
        # 如果是字符串
        if isinstance(value, str):
            value = value.strip()
            # 尝试解析各种时间格式
            for fmt in ["%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"]:
                try:
                    t = datetime.strptime(value, fmt)
                    return t.strftime("%H:%M:%S")
                except:
                    continue
            return value
        
        # 如果是数字（Excel 中时间可能以小数形式存储）
        if isinstance(value, (int, float)):
            try:
                # Excel 时间是以天的分数表示的
                total_seconds = int(value * 24 * 3600)
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            except:
                pass
        
        return str(value)
    
    def parse_date_value(self, value) -> str:
        """解析日期值（含 Excel 序列号，避免 .xlsx 中日期被读成数字导致 25-12 等错误）"""
        if value is None or value == "":
            return ""

        # 如果是 datetime 对象
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")

        # Excel 序列号：整数或小数（Windows 基准 1899-12-30）
        if isinstance(value, (int, float)):
            try:
                days = int(round(float(value)))
                base = datetime(1899, 12, 30)
                d = base + timedelta(days=days)
                return d.strftime("%Y-%m-%d")
            except (ValueError, OverflowError):
                pass
            return ""

        # 如果是字符串
        if isinstance(value, str):
            value = value.strip()
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%m/%d/%Y", "%d/%m/%Y"]:
                try:
                    d = datetime.strptime(value, fmt)
                    return d.strftime("%Y-%m-%d")
                except Exception:
                    continue
            return value

        return ""
    
    def read_attendance_data(self, start_row: int = 6) -> List[Dict]:
        """
        读取考勤数据（从指定行开始）
        返回: 原始记录列表
        """
        if self.is_xls:
            return self._read_xls_data(start_row)
        else:
            return self._read_xlsx_data(start_row)
    
    def _read_xls_data(self, start_row: int) -> List[Dict]:
        """读取 .xls 文件数据"""
        if not self.xlrd_sheet:
            logger.error("工作表未加载")
            return []
        
        records = []
        
        try:
            for row_idx in range(start_row - 1, self.xlrd_sheet.nrows):
                try:
                    row = self.xlrd_sheet.row(row_idx)
                    
                    if len(row) < 6:
                        continue
                    
                    employee_id = row[0].value
                    employee_name = row[1].value
                    department1 = row[2].value
                    department2 = row[3].value
                    attendance_date = row[4].value
                    attendance_time = row[5].value
                    
                    employee_id = self._clean_employee_id(employee_id)

                    if not employee_id or not employee_name:
                        continue
                    
                    if row[4].ctype == 3:  # XL_CELL_DATE
                        date_tuple = xlrd.xldate_as_tuple(attendance_date, self.xlrd_book.datemode)
                        parsed_date = datetime(*date_tuple[:3]).strftime("%Y-%m-%d")
                    else:
                        parsed_date = self.parse_date_value(attendance_date)
                    
                    if row[5].ctype == 3:  # XL_CELL_DATE
                        time_tuple = xlrd.xldate_as_tuple(attendance_time, self.xlrd_book.datemode)
                        parsed_time = f"{time_tuple[3]:02d}:{time_tuple[4]:02d}:{time_tuple[5]:02d}"
                    else:
                        parsed_time = self.parse_time_value(attendance_time)
                    
                    if not parsed_date or not parsed_time:
                        logger.warning(f"第{row_idx+1}行数据不完整（employee_id={employee_id}, date={attendance_date}, time={attendance_time}），跳过")
                        continue

                    inout_raw = row[7].value if len(row) > 7 else None
                    inout_mark = self._parse_inout_mark(inout_raw)
                    
                    record = {
                        'employee_id': employee_id,
                        'employee_name': str(employee_name).strip() if employee_name else "",
                        'department': str(department1).strip() if department1 else "",
                        'attendance_date': parsed_date,
                        'attendance_time': parsed_time,
                        'inout_mark': inout_mark,
                    }
                    
                    records.append(record)
                
                except Exception as e:
                    logger.warning(f"读取第{row_idx+1}行失败: {str(e)}")
                    continue
        
        except Exception as e:
            logger.error(f"读取数据失败: {str(e)}")
        
        logger.info(f"共读取 {len(records)} 条原始记录")
        return records
    
    def _read_xlsx_data(self, start_row: int) -> List[Dict]:
        """读取 .xlsx 文件数据"""
        if not self.worksheet:
            logger.error("工作表未加载")
            return []
        
        records = []
        
        try:
            for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=start_row), start=start_row):
                if len(row) < 6:
                    continue
                
                employee_id = self._clean_employee_id(row[0].value)  # A列
                employee_name = row[1].value  # B列
                department1 = row[2].value  # C列
                department2 = row[3].value  # D列
                attendance_date = row[4].value  # E列
                attendance_time = row[5].value  # F列
                
                if not employee_id or not employee_name:
                    continue
                
                parsed_date = self.parse_date_value(attendance_date)
                parsed_time = self.parse_time_value(attendance_time)
                
                if not parsed_date or not parsed_time:
                    logger.warning(f"第{row_idx}行数据不完整（employee_id={employee_id}, date={attendance_date}, time={attendance_time}），跳过")
                    continue

                inout_raw = row[7].value if len(row) > 7 else None  # H列(index 7)
                inout_mark = self._parse_inout_mark(inout_raw)
                
                record = {
                    'employee_id': employee_id,
                    'employee_name': str(employee_name).strip() if employee_name else "",
                    'department': str(department1).strip() if department1 else "",
                    'attendance_date': parsed_date,
                    'attendance_time': parsed_time,
                    'inout_mark': inout_mark,
                }
                
                records.append(record)
        
        except Exception as e:
            logger.error(f"读取数据失败: {str(e)}")
        
        logger.info(f"共读取 {len(records)} 条原始记录")
        return records
    
    def merge_records_by_employee_and_date(self, records: List[Dict]) -> List[Dict]:
        """
        按员工和日期合并记录（同人同日按时间排序，最多保留 10 次打卡）。
        去重：仅合并无进出标识且间隔 ≤5 秒的重复刷卡；有 H 列「进/出」的全部保留。
        进出标记：H 列有「进/入」「出」则用其值；否则按时间线与前一条交替（首条无法识别时视为进）。
        """
        grouped = defaultdict(list)
        
        for record in records:
            key = (record['employee_id'], record['attendance_date'])
            grouped[key].append(record)
        
        merged_records = []
        
        for (employee_id, attendance_date), group in grouped.items():
            group.sort(key=lambda x: x['attendance_time'])
            group = self._dedup_close_times(group)
            sub = group[:10]
            marks = self._resolve_inout_mark_sequence(sub)
            
            merged = {
                'employee_id': employee_id,
                'employee_name': group[0]['employee_name'],
                'department': group[0]['department'],
                'attendance_date': attendance_date
            }
            
            for i, record in enumerate(sub, start=1):
                merged[f'time_{i}'] = record['attendance_time']
                merged[f'time_{i}_mark'] = marks[i - 1]
            
            for i in range(len(group) + 1, 11):
                merged[f'time_{i}'] = None
                merged[f'time_{i}_mark'] = None
            
            merged_records.append(merged)
        
        logger.info(f"合并后共 {len(merged_records)} 条记录")
        return merged_records
    
    def process_file(self, start_row: int = 6) -> tuple:
        """
        处理文件的完整流程
        返回: (是否成功, 合并后的记录列表, 错误信息)
        """
        try:
            # 1. 加载文件
            if not self.load_file():
                return False, [], "文件加载失败"
            
            # 2. 读取原始数据
            raw_records = self.read_attendance_data(start_row)
            
            if not raw_records:
                return False, [], "未读取到有效数据"
            
            # 3. 合并记录
            merged_records = self.merge_records_by_employee_and_date(raw_records)
            
            return True, merged_records, "处理成功"
        
        except Exception as e:
            error_msg = f"处理文件时出错: {str(e)}"
            logger.error(error_msg)
            return False, [], error_msg
        
        finally:
            # 关闭工作簿
            if self.workbook:
                self.workbook.close()
            # xlrd 不需要显式关闭

