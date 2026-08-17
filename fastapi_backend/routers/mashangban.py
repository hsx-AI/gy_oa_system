# -*- coding: utf-8 -*-
"""
工艺码上办月度综合报表：接收 pusher 推送的 Excel，解析入库并提供查询。
"""
from __future__ import annotations

import io
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, File, Form, HTTPException, Query, UploadFile
from openpyxl import load_workbook

from database import db

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/mashangban", tags=["工艺码上办"])

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data" / "mashangban"
UPLOAD_DIR = DATA_DIR / "uploads"
DATA_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

YEAR_MONTH_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")


def _ensure_tables() -> None:
    statements = [
        """
        CREATE TABLE IF NOT EXISTS mashangban_import_log (
          id BIGINT NOT NULL AUTO_INCREMENT,
          report_month CHAR(7) NOT NULL,
          file_name VARCHAR(255) NULL,
          file_size INT NULL,
          dept_rows INT NOT NULL DEFAULT 0,
          person_rows INT NOT NULL DEFAULT 0,
          order_rows INT NOT NULL DEFAULT 0,
          status VARCHAR(32) NOT NULL DEFAULT 'ok',
          message VARCHAR(500) NULL,
          imported_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
          PRIMARY KEY (id),
          KEY idx_msb_import_ym (report_month),
          KEY idx_msb_import_at (imported_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """,
        """
        CREATE TABLE IF NOT EXISTS mashangban_dept_monthly (
          id BIGINT NOT NULL AUTO_INCREMENT,
          report_month CHAR(7) NOT NULL,
          dept_name VARCHAR(64) NOT NULL,
          order_count INT NULL,
          total_service_hours DECIMAL(12,3) NULL,
          avg_service_hours DECIMAL(12,3) NULL,
          avg_accept_hours DECIMAL(12,3) NULL,
          avg_arrive_hours DECIMAL(12,3) NULL,
          pending_accept INT NULL,
          pending_arrive INT NULL,
          processing INT NULL,
          pending_confirm INT NULL,
          updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (id),
          UNIQUE KEY uk_msb_dept_ym (report_month, dept_name),
          KEY idx_msb_dept_ym (report_month)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """,
        """
        CREATE TABLE IF NOT EXISTS mashangban_person_monthly (
          id BIGINT NOT NULL AUTO_INCREMENT,
          report_month CHAR(7) NOT NULL,
          dept_name VARCHAR(64) NOT NULL,
          employee_name VARCHAR(64) NOT NULL,
          service_count INT NULL,
          total_service_hours DECIMAL(12,3) NULL,
          type_simple INT NULL,
          type_normal INT NULL,
          type_complex INT NULL,
          type_hard INT NULL,
          type_improve INT NULL,
          avg_service_hours DECIMAL(12,3) NULL,
          avg_accept_hours DECIMAL(12,3) NULL,
          avg_arrive_hours DECIMAL(12,3) NULL,
          patrol_factory INT NULL,
          patrol_newtech INT NULL,
          patrol_follow INT NULL,
          patrol_count INT NULL,
          patrol_total_hours DECIMAL(12,3) NULL,
          patrol_avg_hours DECIMAL(12,3) NULL,
          rate_excellent INT NULL,
          rate_good INT NULL,
          rate_normal INT NULL,
          rate_poor INT NULL,
          rate_bad INT NULL,
          updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (id),
          UNIQUE KEY uk_msb_person_ym (report_month, dept_name, employee_name),
          KEY idx_msb_person_ym (report_month),
          KEY idx_msb_person_name (employee_name)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """,
        """
        CREATE TABLE IF NOT EXISTS mashangban_work_orders (
          id BIGINT NOT NULL AUTO_INCREMENT,
          report_month CHAR(7) NOT NULL,
          dept_name VARCHAR(64) NULL,
          operator_name VARCHAR(64) NULL,
          order_type VARCHAR(64) NULL,
          order_no VARCHAR(64) NOT NULL,
          operator_phone VARCHAR(32) NULL,
          workshop_name VARCHAR(128) NULL,
          machine_name VARCHAR(255) NULL,
          workpiece_name VARCHAR(255) NULL,
          work_no VARCHAR(128) NULL,
          assembly_no VARCHAR(128) NULL,
          drawing_no VARCHAR(128) NULL,
          created_at_src DATETIME NULL,
          order_desc TEXT NULL,
          order_status VARCHAR(64) NULL,
          assigner_name VARCHAR(64) NULL,
          assigned_at DATETIME NULL,
          operator_finished_at DATETIME NULL,
          schedule_color VARCHAR(32) NULL,
          process_finished_at DATETIME NULL,
          process_engineer_name VARCHAR(64) NULL,
          process_engineer_phone VARCHAR(32) NULL,
          process_accepted_at DATETIME NULL,
          process_scanned_at DATETIME NULL,
          process_order_status VARCHAR(64) NULL,
          rating_score VARCHAR(16) NULL,
          rating_label VARCHAR(64) NULL,
          rated_at DATETIME NULL,
          department VARCHAR(64) NULL,
          updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
          PRIMARY KEY (id),
          UNIQUE KEY uk_msb_order_ym_no (report_month, order_no),
          KEY idx_msb_order_no (order_no),
          KEY idx_msb_order_ym (report_month),
          KEY idx_msb_order_dept (dept_name),
          KEY idx_msb_order_engineer (process_engineer_name),
          KEY idx_msb_order_status (order_status)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """,
    ]
    for sql in statements:
        try:
            db.execute_update(sql, ())
        except Exception as exc:
            logger.warning("创建码上办表失败: %s", exc)


_ensure_tables()


# 码上办 Excel 科室简称 → yggl.lsys 标准科室名
DEPT_NAME_ALIASES = {
    "水发室": "水发工艺室",
    "水轮机室": "水轮机工艺室",
    "汽发室": "汽发工艺室",
    "焊艺室": "焊接工艺室",
    "焊接室": "焊接工艺室",
    "智能室": "智能制造技术室",
    "智能制造室": "智能制造技术室",
    "综合室": "综合技术室",
    "数控室": "数控编程室",
    "非标室": "非标技术室",
    "工具室": "工具技术室",
    "部办室": "部办",
}


_LSYS_CACHE: Optional[List[str]] = None


def _load_standard_lsys(force_refresh: bool = False) -> List[str]:
    """读取 yggl.lsys 主科室名（优先不含末尾数字的正式名称）。"""
    global _LSYS_CACHE
    if _LSYS_CACHE is not None and not force_refresh:
        return _LSYS_CACHE
    rows = db.execute_query(
        """
        SELECT DISTINCT TRIM(lsys) AS lsys
        FROM yggl
        WHERE lsys IS NOT NULL AND TRIM(lsys) <> ''
        ORDER BY lsys
        """
    )
    names = [(r.get("lsys") or "").strip() for r in rows]
    names = [n for n in names if n]
    # 正式名优先：去掉「水发工艺室1」这类副本名，只保留无数字后缀的。
    primary = [n for n in names if not re.search(r"\d+$", n)]
    _LSYS_CACHE = primary or names
    return _LSYS_CACHE


def _normalize_dept_name(raw: Any) -> str:
    """把码上办科室名标准化为系统 yggl.lsys 用词。"""
    text = ("" if raw is None else str(raw)).strip()
    if isinstance(raw, float) and raw == int(raw):
        text = str(int(raw))
    text = text.strip()
    if not text:
        return ""

    standards = _load_standard_lsys()
    standard_set = set(standards)

    if text in standard_set:
        return text

    aliased = DEPT_NAME_ALIASES.get(text)
    if aliased and aliased in standard_set:
        return aliased
    if aliased:
        return aliased

    # 已是「xxx工艺室 / xxx技术室」但不在别名表时，尽量对齐到正式名。
    for std in standards:
        if text == std:
            return std
        if text.rstrip("0123456789") == std:
            return std

    # 「汽发」→「汽发工艺室」等宽松匹配
    compact = text.replace("工艺室", "").replace("技术室", "").replace("室", "")
    for std in standards:
        std_compact = std.replace("工艺室", "").replace("技术室", "").replace("室", "")
        if compact and compact == std_compact:
            return std

    logger.warning("码上办科室未映射到 yggl.lsys，保留原值: %s", text)
    return text


def _validate_year_month(value: str) -> str:
    text = (value or "").strip()
    if not YEAR_MONTH_RE.match(text):
        raise HTTPException(status_code=422, detail="yearMonth 格式应为 YYYY-MM")
    return text


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _to_int(value: Any) -> Optional[int]:
    text = _cell_str(value)
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _to_float(value: Any) -> Optional[float]:
    text = _cell_str(value)
    if not text:
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _to_datetime(value: Any) -> Optional[datetime]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = _cell_str(value)
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _split_counts(value: Any, expected: int) -> List[Optional[int]]:
    text = _cell_str(value)
    if not text:
        return [None] * expected
    parts = [p.strip() for p in text.split("/")]
    result: List[Optional[int]] = []
    for idx in range(expected):
        if idx >= len(parts) or parts[idx] == "":
            result.append(None)
        else:
            result.append(_to_int(parts[idx]))
    return result


def _sheet_rows(ws) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def _parse_dept_sheet(rows: List[List[Any]], year_month: str) -> List[Tuple]:
    result = []
    for row in rows[1:]:
        if not row or not _cell_str(row[0] if len(row) > 0 else None):
            continue
        result.append(
            (
                year_month,
                _normalize_dept_name(row[0]),
                _to_int(row[1] if len(row) > 1 else None),
                _to_float(row[2] if len(row) > 2 else None),
                _to_float(row[3] if len(row) > 3 else None),
                _to_float(row[4] if len(row) > 4 else None),
                _to_float(row[5] if len(row) > 5 else None),
                _to_int(row[6] if len(row) > 6 else None),
                _to_int(row[7] if len(row) > 7 else None),
                _to_int(row[8] if len(row) > 8 else None),
                _to_int(row[9] if len(row) > 9 else None),
            )
        )
    return result


def _parse_person_sheet(rows: List[List[Any]], year_month: str, dept_name: str) -> List[Tuple]:
    result = []
    std_dept = _normalize_dept_name(dept_name)
    for row in rows[1:]:
        if not row:
            continue
        name = _cell_str(row[2] if len(row) > 2 else None)
        if not name:
            continue
        type_counts = _split_counts(row[3] if len(row) > 3 else None, 5)
        patrol_counts = _split_counts(row[7] if len(row) > 7 else None, 3)
        result.append(
            (
                year_month,
                std_dept,
                name,
                _to_int(row[0] if len(row) > 0 else None),
                _to_float(row[1] if len(row) > 1 else None),
                type_counts[0],
                type_counts[1],
                type_counts[2],
                type_counts[3],
                type_counts[4],
                _to_float(row[4] if len(row) > 4 else None),
                _to_float(row[5] if len(row) > 5 else None),
                _to_float(row[6] if len(row) > 6 else None),
                patrol_counts[0],
                patrol_counts[1],
                patrol_counts[2],
                _to_int(row[8] if len(row) > 8 else None),
                _to_float(row[9] if len(row) > 9 else None),
                _to_float(row[10] if len(row) > 10 else None),
                _to_int(row[11] if len(row) > 11 else None),
                _to_int(row[12] if len(row) > 12 else None),
                _to_int(row[13] if len(row) > 13 else None),
                _to_int(row[14] if len(row) > 14 else None),
                _to_int(row[15] if len(row) > 15 else None),
            )
        )
    return result


def _parse_order_sheet(rows: List[List[Any]], year_month: str) -> List[Tuple]:
    result = []
    seen = set()
    # 同月 Excel 可能出现重复工单号；后写覆盖先写。
    for row in rows[1:]:
        if not row:
            continue
        order_no = _cell_str(row[3] if len(row) > 3 else None)
        if not order_no:
            continue
        item = (
            year_month,
            _normalize_dept_name(row[0] if len(row) > 0 else None) or None,
            _cell_str(row[1] if len(row) > 1 else None) or None,
            _cell_str(row[2] if len(row) > 2 else None) or None,
            order_no,
            _cell_str(row[4] if len(row) > 4 else None) or None,
            _cell_str(row[5] if len(row) > 5 else None) or None,
            _cell_str(row[6] if len(row) > 6 else None) or None,
            _cell_str(row[7] if len(row) > 7 else None) or None,
            _cell_str(row[8] if len(row) > 8 else None) or None,
            _cell_str(row[9] if len(row) > 9 else None) or None,
            _cell_str(row[10] if len(row) > 10 else None) or None,
            _to_datetime(row[11] if len(row) > 11 else None),
            _cell_str(row[12] if len(row) > 12 else None) or None,
            _cell_str(row[13] if len(row) > 13 else None) or None,
            _cell_str(row[14] if len(row) > 14 else None) or None,
            _to_datetime(row[15] if len(row) > 15 else None),
            _to_datetime(row[16] if len(row) > 16 else None),
            _cell_str(row[17] if len(row) > 17 else None) or None,
            _to_datetime(row[18] if len(row) > 18 else None),
            _cell_str(row[19] if len(row) > 19 else None) or None,
            _cell_str(row[20] if len(row) > 20 else None) or None,
            _to_datetime(row[21] if len(row) > 21 else None),
            _to_datetime(row[22] if len(row) > 22 else None),
            _cell_str(row[23] if len(row) > 23 else None) or None,
            _cell_str(row[24] if len(row) > 24 else None) or None,
            _cell_str(row[25] if len(row) > 25 else None) or None,
            _to_datetime(row[26] if len(row) > 26 else None),
            _normalize_dept_name(row[27] if len(row) > 27 else None) or None,
        )
        if order_no in seen:
            for idx, old in enumerate(result):
                if old[4] == order_no:
                    result[idx] = item
                    break
        else:
            seen.add(order_no)
            result.append(item)
    return result


def _parse_workbook(content: bytes, year_month: str) -> Dict[str, List[Tuple]]:
    # 每次解析前刷新一次科室标准名，避免长期缓存过期。
    _load_standard_lsys(force_refresh=True)
    wb = load_workbook(io.BytesIO(content), data_only=True)
    dept_rows: List[Tuple] = []
    person_rows: List[Tuple] = []
    order_rows: List[Tuple] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = _sheet_rows(ws)
        if not rows:
            continue
        name = (sheet_name or "").strip()
        if name == "科室统计数据":
            dept_rows.extend(_parse_dept_sheet(rows, year_month))
        elif name.endswith("服务绩效"):
            dept_name = name[: -len("服务绩效")].strip() or name
            person_rows.extend(_parse_person_sheet(rows, year_month, dept_name))
        elif name == "工单列表":
            order_rows.extend(_parse_order_sheet(rows, year_month))
        else:
            logger.info("忽略未知工作表: %s", name)

    return {
        "dept": dept_rows,
        "person": person_rows,
        "order": order_rows,
    }


def _replace_month_data(year_month: str, parsed: Dict[str, List[Tuple]]) -> None:
    """同月整表覆盖：先删该月旧数据，再写入本次解析结果（支持源端数据更新）。"""
    db.execute_update("DELETE FROM mashangban_dept_monthly WHERE report_month=%s", (year_month,))
    db.execute_update("DELETE FROM mashangban_person_monthly WHERE report_month=%s", (year_month,))
    db.execute_update("DELETE FROM mashangban_work_orders WHERE report_month=%s", (year_month,))

    if parsed["dept"]:
        affected = db.execute_many(
            """
            INSERT INTO mashangban_dept_monthly (
              report_month, dept_name, order_count, total_service_hours, avg_service_hours,
              avg_accept_hours, avg_arrive_hours, pending_accept, pending_arrive,
              processing, pending_confirm
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            parsed["dept"],
        )
        if affected < 0:
            raise HTTPException(status_code=500, detail="写入科室统计失败")

    if parsed["person"]:
        affected = db.execute_many(
            """
            INSERT INTO mashangban_person_monthly (
              report_month, dept_name, employee_name, service_count, total_service_hours,
              type_simple, type_normal, type_complex, type_hard, type_improve,
              avg_service_hours, avg_accept_hours, avg_arrive_hours,
              patrol_factory, patrol_newtech, patrol_follow,
              patrol_count, patrol_total_hours, patrol_avg_hours,
              rate_excellent, rate_good, rate_normal, rate_poor, rate_bad
            ) VALUES (
              %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s
            )
            """,
            parsed["person"],
        )
        if affected < 0:
            raise HTTPException(status_code=500, detail="写入人员绩效失败")

    if parsed["order"]:
        affected = db.execute_many(
            """
            INSERT INTO mashangban_work_orders (
              report_month, dept_name, operator_name, order_type, order_no, operator_phone,
              workshop_name, machine_name, workpiece_name, work_no, assembly_no, drawing_no,
              created_at_src, order_desc, order_status, assigner_name, assigned_at,
              operator_finished_at, schedule_color, process_finished_at,
              process_engineer_name, process_engineer_phone, process_accepted_at,
              process_scanned_at, process_order_status, rating_score, rating_label,
              rated_at, department
            ) VALUES (
              %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s
            )
            """,
            parsed["order"],
        )
        if affected < 0:
            raise HTTPException(status_code=500, detail="写入工单明细失败")


LATEST_EXCEL_NAME = "latest.xlsx"


def _keep_only_latest_excel(content: bytes) -> Path:
    """磁盘只保留一份最新 Excel，清理历史月份文件，避免长期堆积。"""
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    latest_path = UPLOAD_DIR / LATEST_EXCEL_NAME
    latest_path.write_bytes(content)
    for path in UPLOAD_DIR.iterdir():
        if not path.is_file():
            continue
        if path.name == LATEST_EXCEL_NAME:
            continue
        if path.suffix.lower() in {".xlsx", ".xls", ".tmp"}:
            try:
                path.unlink()
            except Exception as exc:
                logger.warning("清理旧码上办 Excel 失败 %s: %s", path, exc)
    return latest_path


@router.post("/push/report")
async def push_report(
    yearMonth: str = Form(...),
    file: UploadFile = File(...),
):
    """pusher 推送码上办月度 Excel：同月数据覆盖入库，磁盘仅保留最新一份 Excel。"""
    year_month = _validate_year_month(yearMonth)
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="文件为空")

    safe_name = Path(file.filename or f"mashangban_{year_month}.xlsx").name

    try:
        parsed = _parse_workbook(content, year_month)
        if not parsed["dept"] and not parsed["person"] and not parsed["order"]:
            raise HTTPException(status_code=400, detail="Excel 未解析到有效数据")
        _replace_month_data(year_month, parsed)
        _keep_only_latest_excel(content)
    except HTTPException:
        db.execute_update(
            """
            INSERT INTO mashangban_import_log
              (report_month, file_name, file_size, dept_rows, person_rows, order_rows, status, message)
            VALUES (%s,%s,%s,0,0,0,'error',%s)
            """,
            (year_month, safe_name, len(content), "解析或入库失败"),
        )
        raise
    except Exception as exc:
        logger.exception("码上办月报入库失败")
        db.execute_update(
            """
            INSERT INTO mashangban_import_log
              (report_month, file_name, file_size, dept_rows, person_rows, order_rows, status, message)
            VALUES (%s,%s,%s,0,0,0,'error',%s)
            """,
            (year_month, safe_name, len(content), str(exc)[:480]),
        )
        raise HTTPException(status_code=500, detail=f"解析失败: {exc}") from exc

    dept_n = len(parsed["dept"])
    person_n = len(parsed["person"])
    order_n = len(parsed["order"])
    db.execute_update(
        """
        INSERT INTO mashangban_import_log
          (report_month, file_name, file_size, dept_rows, person_rows, order_rows, status, message)
        VALUES (%s,%s,%s,%s,%s,%s,'ok',%s)
        """,
        (
            year_month,
            LATEST_EXCEL_NAME,
            len(content),
            dept_n,
            person_n,
            order_n,
            f"dept={dept_n}, person={person_n}, order={order_n}; overwritten month={year_month}",
        ),
    )
    return {
        "success": True,
        "status": "ok",
        "yearMonth": year_month,
        "deptRows": dept_n,
        "personRows": person_n,
        "orderRows": order_n,
        "overwritten": True,
        "file": f"/api/mashangban/uploads/{LATEST_EXCEL_NAME}",
    }


@router.get("/months")
def list_months():
    rows = db.execute_query(
        """
        SELECT report_month AS yearMonth,
               MAX(imported_at) AS latestImport,
               SUM(CASE WHEN status='ok' THEN 1 ELSE 0 END) AS okCount
        FROM mashangban_import_log
        GROUP BY report_month
        ORDER BY report_month DESC
        """
    )
    return {"success": True, "items": rows}


@router.get("/dept")
def list_dept(yearMonth: str = Query(...)):
    year_month = _validate_year_month(yearMonth)
    rows = db.execute_query(
        """
        SELECT report_month AS yearMonth, dept_name AS deptName, order_count AS orderCount,
               total_service_hours AS totalServiceHours, avg_service_hours AS avgServiceHours,
               avg_accept_hours AS avgAcceptHours, avg_arrive_hours AS avgArriveHours,
               pending_accept AS pendingAccept, pending_arrive AS pendingArrive,
               processing, pending_confirm AS pendingConfirm, updated_at AS updatedAt
        FROM mashangban_dept_monthly
        WHERE report_month=%s
        ORDER BY order_count DESC, dept_name ASC
        """,
        (year_month,),
    )
    return {"success": True, "yearMonth": year_month, "items": rows}


@router.get("/person")
def list_person(
    yearMonth: str = Query(...),
    dept: Optional[str] = Query(None),
):
    year_month = _validate_year_month(yearMonth)
    sql = """
        SELECT report_month AS yearMonth, dept_name AS deptName, employee_name AS employeeName,
               service_count AS serviceCount, total_service_hours AS totalServiceHours,
               type_simple AS typeSimple, type_normal AS typeNormal, type_complex AS typeComplex,
               type_hard AS typeHard, type_improve AS typeImprove,
               avg_service_hours AS avgServiceHours, avg_accept_hours AS avgAcceptHours,
               avg_arrive_hours AS avgArriveHours,
               patrol_factory AS patrolFactory, patrol_newtech AS patrolNewtech,
               patrol_follow AS patrolFollow, patrol_count AS patrolCount,
               patrol_total_hours AS patrolTotalHours, patrol_avg_hours AS patrolAvgHours,
               rate_excellent AS rateExcellent, rate_good AS rateGood, rate_normal AS rateNormal,
               rate_poor AS ratePoor, rate_bad AS rateBad, updated_at AS updatedAt
        FROM mashangban_person_monthly
        WHERE report_month=%s
    """
    params: List[Any] = [year_month]
    dept_name = (dept or "").strip()
    if dept_name:
        sql += " AND dept_name=%s"
        params.append(dept_name)
    sql += " ORDER BY dept_name ASC, service_count DESC, employee_name ASC"
    rows = db.execute_query(sql, tuple(params))
    return {"success": True, "yearMonth": year_month, "items": rows}


@router.get("/orders")
def list_orders(
    yearMonth: str = Query(...),
    dept: Optional[str] = Query(None),
    engineer: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    orderType: Optional[str] = Query(None),
    operator: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    pageSize: int = Query(50, ge=1, le=2000),
):
    year_month = _validate_year_month(yearMonth)
    where = ["report_month=%s"]
    params: List[Any] = [year_month]
    if (dept or "").strip():
        where.append("dept_name=%s")
        params.append(dept.strip())
    if (engineer or "").strip():
        where.append("process_engineer_name=%s")
        params.append(engineer.strip())
    if (status or "").strip():
        where.append("order_status=%s")
        params.append(status.strip())
    if (orderType or "").strip():
        where.append("order_type=%s")
        params.append(orderType.strip())
    if (operator or "").strip():
        where.append("operator_name=%s")
        params.append(operator.strip())
    if (keyword or "").strip():
        like = f"%{keyword.strip()}%"
        where.append(
            "(order_no LIKE %s OR workpiece_name LIKE %s OR work_no LIKE %s OR order_desc LIKE %s OR process_engineer_name LIKE %s OR operator_name LIKE %s)"
        )
        params.extend([like, like, like, like, like, like])
    where_sql = " AND ".join(where)
    total = db.execute_scalar(
        f"SELECT COUNT(1) FROM mashangban_work_orders WHERE {where_sql}",
        tuple(params),
    ) or 0
    offset = (page - 1) * pageSize
    rows = db.execute_query(
        f"""
        SELECT report_month AS yearMonth, dept_name AS deptName, operator_name AS operatorName,
               order_type AS orderType, order_no AS orderNo, operator_phone AS operatorPhone,
               workshop_name AS workshopName, machine_name AS machineName,
               workpiece_name AS workpieceName, work_no AS workNo, assembly_no AS assemblyNo,
               drawing_no AS drawingNo, created_at_src AS createdAt, order_desc AS orderDesc,
               order_status AS orderStatus, assigner_name AS assignerName, assigned_at AS assignedAt,
               operator_finished_at AS operatorFinishedAt, schedule_color AS scheduleColor,
               process_finished_at AS processFinishedAt, process_engineer_name AS processEngineerName,
               process_engineer_phone AS processEngineerPhone, process_accepted_at AS processAcceptedAt,
               process_scanned_at AS processScannedAt, process_order_status AS processOrderStatus,
               rating_score AS ratingScore, rating_label AS ratingLabel, rated_at AS ratedAt,
               department
        FROM mashangban_work_orders
        WHERE {where_sql}
        ORDER BY created_at_src DESC, order_no DESC
        LIMIT %s OFFSET %s
        """,
        tuple(params + [pageSize, offset]),
    )
    return {
        "success": True,
        "yearMonth": year_month,
        "page": page,
        "pageSize": pageSize,
        "total": int(total),
        "items": rows,
    }


@router.get("/uploads/{filename}")
def download_uploaded(filename: str):
    """仅提供最新 Excel；历史月份文件不会长期保留。"""
    safe_name = Path(filename).name
    if safe_name not in {LATEST_EXCEL_NAME, "latest"}:
        # 兼容旧链接：任意月份文件名也回落到 latest.xlsx
        safe_name = LATEST_EXCEL_NAME
    if safe_name == "latest":
        safe_name = LATEST_EXCEL_NAME
    path = UPLOAD_DIR / safe_name
    if not path.exists():
        raise HTTPException(status_code=404, detail="暂无最新 Excel")
    from fastapi.responses import FileResponse

    return FileResponse(
        path,
        filename=LATEST_EXCEL_NAME,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
